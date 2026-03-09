from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional, Dict
import json
import os
import datetime
from scheduler_engine import ShiftScheduler
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from fastapi.responses import FileResponse
import tempfile
import subprocess
import shutil

app = FastAPI()

# DATABASE — SQLite backend (shared with planilla system)
import sys
import sqlite3

# Add planillas dir to path so we can import database module
_planillas_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas")
if os.path.exists(_planillas_dir):
    sys.path.insert(0, os.path.abspath(_planillas_dir))
    import database as plan_db  # Initializes planilla.db with all tables
    import planilla as pl_module
    import generador_boletas as gb_module
    import horario_db

DB_FILE_LEGACY = "database.json"  # JSON original, kept for migration reference

def _get_conn():
    """Get connection to the shared planilla.db"""
    return plan_db.get_conn()

def load_db():
    """Lee datos de SQLite y devuelve dict compatible con la estructura JSON original."""
    conn = _get_conn()

    # Empleados
    rows = conn.execute("SELECT * FROM horario_empleados WHERE activo=1").fetchall()
    employees = []
    for r in rows:
        emp = {
            "name": r["nombre"],
            "gender": r["genero"],
            "can_do_night": bool(r["puede_nocturno"]),
            "allow_no_rest": bool(r["allow_no_rest"]),
            "forced_libres": bool(r["forced_libres"]),
            "forced_quebrado": bool(r["forced_quebrado"]),
            "is_jefe_pista": bool(r["es_jefe_pista"]),
            "strict_preferences": bool(r["strict_preferences"]) if "strict_preferences" in r.keys() else False,
            "fixed_shifts": json.loads(r["turnos_fijos"]) if r["turnos_fijos"] else {},
        }
        employees.append(emp)

    # Config
    cfg_row = conn.execute("SELECT * FROM horario_config WHERE id=1").fetchone()
    config = {}
    if cfg_row:
        config = {
            "night_mode": cfg_row["night_mode"],
            "fixed_night_person": cfg_row["fixed_night_person"],
            "allow_long_shifts": bool(cfg_row["allow_long_shifts"]),
            "use_refuerzo": bool(cfg_row["use_refuerzo"]),
            "refuerzo_type": cfg_row["refuerzo_type"],
            "allow_collision_quebrado": bool(cfg_row["allow_collision_quebrado"]),
            "collision_peak_priority": cfg_row["collision_peak_priority"],
            "sunday_cycle_index": cfg_row["sunday_cycle_index"] or 0,
            "sunday_rotation_queue": json.loads(cfg_row["sunday_rotation_queue"]) if cfg_row["sunday_rotation_queue"] else None,
        }

    # History log
    hist_rows = conn.execute("SELECT * FROM horarios_generados ORDER BY id").fetchall()
    history_log = []
    for r in hist_rows:
        entry = {
            "name": r["nombre"],
            "schedule": json.loads(r["horario"]),
            "daily_tasks": json.loads(r["tareas"]) if r["tareas"] else {},
            "timestamp": r["timestamp"],
        }
        meta = json.loads(r["metadata"]) if r["metadata"] else {}
        entry.update(meta)
        history_log.append(entry)

    # Last result: use the most recent history entry or empty
    last_result = {}
    if history_log:
        last_entry = history_log[-1]
        last_result = {
            "status": "Success",
            "schedule": last_entry.get("schedule", {}),
            "daily_tasks": last_entry.get("daily_tasks", {}),
            "metadata": {k: v for k, v in last_entry.items() if k not in ("name", "schedule", "daily_tasks", "timestamp")},
        }

    conn.close()
    return {
        "employees": employees,
        "config": config,
        "history_log": history_log,
        "last_result": last_result,
    }


def save_db(data):
    """Guarda dict de vuelta a SQLite, manteniendo compatibilidad."""
    conn = _get_conn()

    # Guardar empleados
    if "employees" in data:
        # Desactivar todos primero
        conn.execute("UPDATE horario_empleados SET activo=0")
        for emp in data["employees"]:
            existing = conn.execute(
                "SELECT id FROM horario_empleados WHERE nombre=?", (emp["name"],)
            ).fetchone()
            if existing:
                conn.execute("""
                    UPDATE horario_empleados SET
                        genero=?, puede_nocturno=?, allow_no_rest=?, forced_libres=?,
                        forced_quebrado=?, es_jefe_pista=?, strict_preferences=?, turnos_fijos=?, activo=1
                    WHERE nombre=?
                """, (
                    emp.get("gender", "M"),
                    1 if emp.get("can_do_night", True) else 0,
                    1 if emp.get("allow_no_rest", False) else 0,
                    1 if emp.get("forced_libres", False) else 0,
                    1 if emp.get("forced_quebrado", False) else 0,
                    1 if emp.get("is_jefe_pista", False) else 0,
                    1 if emp.get("strict_preferences", False) else 0,
                    json.dumps(emp.get("fixed_shifts", {}), ensure_ascii=False),
                    emp["name"],
                ))
            else:
                conn.execute("""
                    INSERT INTO horario_empleados
                    (nombre, genero, puede_nocturno, allow_no_rest, forced_libres,
                     forced_quebrado, es_jefe_pista, strict_preferences, turnos_fijos, activo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """, (
                    emp["name"], emp.get("gender", "M"),
                    1 if emp.get("can_do_night", True) else 0,
                    1 if emp.get("allow_no_rest", False) else 0,
                    1 if emp.get("forced_libres", False) else 0,
                    1 if emp.get("forced_quebrado", False) else 0,
                    1 if emp.get("is_jefe_pista", False) else 0,
                    1 if emp.get("strict_preferences", False) else 0,
                    json.dumps(emp.get("fixed_shifts", {}), ensure_ascii=False),
                ))

    # Guardar config
    if "config" in data:
        cfg = data["config"]
        conn.execute("DELETE FROM horario_config")
        conn.execute("""
            INSERT INTO horario_config
            (id, night_mode, fixed_night_person, allow_long_shifts, use_refuerzo,
             refuerzo_type, allow_collision_quebrado, collision_peak_priority,
             sunday_cycle_index, sunday_rotation_queue)
            VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cfg.get("night_mode", "rotation"),
            cfg.get("fixed_night_person"),
            1 if cfg.get("allow_long_shifts", False) else 0,
            1 if cfg.get("use_refuerzo", False) else 0,
            cfg.get("refuerzo_type", "diurno"),
            1 if cfg.get("allow_collision_quebrado", False) else 0,
            cfg.get("collision_peak_priority", "pm"),
            cfg.get("sunday_cycle_index", 0),
            json.dumps(cfg.get("sunday_rotation_queue")) if cfg.get("sunday_rotation_queue") else None,
        ))

    # Guardar history_log (reescribir completo)
    if "history_log" in data:
        conn.execute("DELETE FROM horarios_generados")
        for entry in data["history_log"]:
            meta = {k: v for k, v in entry.items()
                    if k not in ("name", "schedule", "daily_tasks", "timestamp")}
            conn.execute("""
                INSERT INTO horarios_generados (nombre, horario, tareas, metadata, timestamp)
                VALUES (?, ?, ?, ?, ?)
            """, (
                entry.get("name", "SIN NOMBRE"),
                json.dumps(entry.get("schedule", {}), ensure_ascii=False),
                json.dumps(entry.get("daily_tasks", {}), ensure_ascii=False),
                json.dumps(meta, ensure_ascii=False),
                entry.get("timestamp", datetime.datetime.now().isoformat()),
            ))

    conn.commit()
    conn.close()

# INITIALIZATION (Seed if empty)
db = load_db()
if not db.get("employees"):
    default_employees = [
        {"name": "Jeison", "gender": "M", "can_do_night": False, "is_jefe_pista": True, "fixed_shifts": {"Vie": "J_06-16", "Lun": "J_06-16", "Mar": "J_06-16", "Mié": "J_06-16", "Jue": "J_06-16", "Sáb": "T1_05-13", "Dom": "OFF"}},
        {"name": "Eligio", "gender": "M", "can_do_night": True},
        {"name": "Maikel", "gender": "M", "can_do_night": True},
        {"name": "Jensy", "gender": "F", "can_do_night": False},
        {"name": "Ileana", "gender": "F", "can_do_night": False},
        {"name": "Steven", "gender": "M", "can_do_night": True},
        {"name": "Randall", "gender": "M", "can_do_night": True},
        {"name": "Angel", "gender": "M", "can_do_night": True},
        {"name": "Keilor", "gender": "M", "can_do_night": True}
    ]
    db["employees"] = default_employees
    db["config"] = {
        "night_mode": "rotation", 
        "fixed_night_person": "Eligio"
    }
    save_db(db)

# MODELS
class Employee(BaseModel):
    name: str
    gender: str = "M"
    can_do_night: bool = True
    allow_no_rest: bool = False
    forced_libres: bool = False
    forced_quebrado: bool = False
    is_jefe_pista: bool = False
    strict_preferences: bool = False
    fixed_shifts: Dict[str, str] = {} 

class Config(BaseModel):
    night_mode: str = "rotation"
    fixed_night_person: Optional[str] = None
    allow_long_shifts: bool = False
    use_refuerzo: bool = False
    refuerzo_type: str = "diurno"
    allow_collision_quebrado: bool = False
    collision_peak_priority: str = "pm"
    sunday_cycle_index: int = 0  # Legacy, kept for backwards compat
    sunday_rotation_queue: Optional[List[str]] = None

class SolverRequest(BaseModel):
    employees: List[Employee]
    config: Config

class HistoryEntry(BaseModel):
    name: str
    schedule: Dict[str, Dict[str, str]]
    daily_tasks: Dict[str, Dict[str, Optional[str]]] = {}
    next_sunday_cycle_index: Optional[int] = None  # Legacy
    next_sunday_rotation_queue: Optional[List[str]] = None
    timestamp: str = "" 

# ENDPOINTS
@app.get("/api/employees")
def get_employees():
    # Map from the unified Planilla Database format to the legacy generator format
    unified_emps = plan_db.get_empleados(solo_activos=True)
    legacy_emps = []
    
    for e in unified_emps:
        try:
            fixed_shifts = json.loads(e.get("turnos_fijos", "{}")) if e.get("turnos_fijos") else {}
        except:
            fixed_shifts = {}
            
        legacy_emps.append({
            "name": e.get("nombre", ""),
            "gender": e.get("genero", "M"),
            "can_do_night": bool(e.get("puede_nocturno", 1)),
            "allow_no_rest": bool(e.get("allow_no_rest", 0)),
            "forced_libres": bool(e.get("forced_libres", 0)),
            "forced_quebrado": bool(e.get("forced_quebrado", 0)),
            "is_jefe_pista": bool(e.get("es_jefe_pista", 0)),
            "strict_preferences": bool(e.get("strict_preferences", 0)),
            "fixed_shifts": fixed_shifts
        })
        
    return legacy_emps

@app.post("/api/employees")
def update_employees(employees: List[Employee]):
    db = load_db()
    db["employees"] = [e.dict() for e in employees]
    save_db(db)
    return {"status": "Updated"}

@app.get("/api/config")
def get_config():
    db = load_db()
    return db.get("config", {})

@app.post("/api/config")
def update_config(config: Config):
    db = load_db()
    db["config"] = config.dict()
    save_db(db)
    return {"status": "Updated"}

@app.post("/api/solve")
def solve_schedule(request: SolverRequest):
    db = load_db()
    
    # Get History (List of past schedules)
    history_list = db.get("history_log", [])
    if not isinstance(history_list, list):
        history_list = []
        
    employees_data = [e.dict() for e in request.employees]
    config_data = request.config.dict()
    
    # Instantiate Scheduler
    scheduler = ShiftScheduler(employees_data, config_data, history_data=history_list)
    result = scheduler.solve()
    
    # Save last result for Export (if feasible or even partial)
    # We save it to DB so export_excel can read it
    if result.get("schedule"):
        db["last_result"] = result
        save_db(db)
    
    return result

@app.get("/api/history")
def get_history():
    db = load_db()
    return db.get("history_log", [])

@app.post("/api/history")
def save_history(entry: HistoryEntry):
    db = load_db()
    history_list = db.get("history_log", [])
    if not isinstance(history_list, list): history_list = []
    
    # Add new entry
    new_record = entry.dict()
    if not new_record.get("timestamp"):
        new_record["timestamp"] = datetime.datetime.now().isoformat()
        
    history_list.append(new_record)
    
    # Limit history size if needed, user requested manual deletion so maybe high limit
    if len(history_list) > 50:
        history_list = history_list[-50:]
        
    db["history_log"] = history_list
    
    # Update Sunday Rotation Queue if present (persist logic state)
    if "config" not in db: db["config"] = {}
    if entry.next_sunday_rotation_queue is not None:
        db["config"]["sunday_rotation_queue"] = entry.next_sunday_rotation_queue
    elif entry.next_sunday_cycle_index is not None:
        # Legacy fallback
        db["config"]["sunday_cycle_index"] = entry.next_sunday_cycle_index
        
    save_db(db)
    return {"status": "Saved", "history_len": len(history_list)}

@app.delete("/api/history/{index}")
def delete_history_item(index: int):
    db = load_db()
    history_list = db.get("history_log", [])
    
    if 0 <= index < len(history_list):
        history_list.pop(index)
        db["history_log"] = history_list
        save_db(db)
        return {"status": "Deleted"}
    raise HTTPException(status_code=404, detail="Index out of bounds")

@app.patch("/api/history/{index}")
def update_history_item(index: int, entry: HistoryEntry):
    db = load_db()
    history_list = db.get("history_log", [])
    
    if 0 <= index < len(history_list):
        # Update specific fields
        history_list[index]["schedule"] = entry.schedule
        history_list[index]["daily_tasks"] = entry.daily_tasks
        db["history_log"] = history_list
        save_db(db)
        return {"status": "Updated"}
    raise HTTPException(status_code=404, detail="Index out of bounds")

def format_shift_code(code: str) -> str:
    """
    Converts internal shift code (e.g., 'T12_14-22') to readable 12h format 
    (e.g., '02:00 PM - 10:00 PM').
    """
    if not code or code == "OFF": return "LIBRE"
    if code == "VAC": return "VACACIONES"
    
    # Check for split shift first (e.g. Q1_05-11+17-20)
    if "+" in code:
        parts = code.split("_")
        if len(parts) > 1:
            times = parts[1].split("+") # ['05-11', '17-20']
            readable_times = []
            for t in times:
                readable_times.append(_format_time_range(t))
            return " / ".join(readable_times)
    
    # Standard Shift (Code_Start-End)
    parts = code.split("_")
    if len(parts) > 1:
        return _format_time_range(parts[1])
        
    return code

def _format_time_range(time_range: str) -> str:
    # time_range ex: "14-22" or "06-16"
    try:
        start_s, end_s = time_range.split("-")
        start_h = int(start_s)
        end_h = int(end_s)
        
        # Handle > 24 (next day)
        start_dt = datetime.time(start_h % 24, 0)
        end_dt = datetime.time(end_h % 24, 0)
        
        start_str = start_dt.strftime("%I:%M %p")
        end_str = end_dt.strftime("%I:%M %p")
        
        return f"{start_str} - {end_str}"
    except:
        return time_range

@app.get("/api/validation_rules")
def get_validation_rules():
    db = load_db()
    employees = db.get("employees", [])
    
    # Calculate standard_mode identical to scheduler_engine
    from scheduler_engine import DAYS
    active_count = 0
    for e in employees:
        if e.get('is_refuerzo', False) or e.get('name') == 'Refuerzo':
            continue
        fixed = e.get('fixed_shifts', {})
        all_absent = all(fixed.get(d) in ['VAC', 'PERM'] for d in DAYS)
        if not all_absent:
            active_count += 1
            
    standard_mode = active_count >= 10
    
    from scheduler_engine import SHIFTS, HOURS, coverage_bounds
    
    # Precompute coverage matrix bounds: { "Dom": { "5": 2, "6": 2... } }
    bounds = {}
    soft_bounds = {}  # Desired optimal coverage (for yellow warnings)
    for d in DAYS:
        bounds[d] = {}
        soft_bounds[d] = {}
        for h in HOURS:
            mn, mx = coverage_bounds(h, d, standard_mode)
            bounds[d][h] = mn
            
            # Soft targets (desired, not strictly enforced)
            if d == "Dom" and standard_mode:
                # Exact puzzle — no slack, soft = hard
                soft_bounds[d][h] = mn
            elif d == "Dom":
                soft_bounds[d][h] = mn  # No specific soft targets on short-staffed Sunday
            else:
                # Weekdays: peak hours have soft target of 4
                if 7 <= h <= 10 or 17 <= h <= 19:
                    soft_bounds[d][h] = 4
                elif h == 6:
                    soft_bounds[d][h] = 3
                else:
                    soft_bounds[d][h] = mn  # Same as hard
            
    # Shift sets list conversion
    shift_sets = {s: list(hours) for s, hours in SHIFTS.items()}
    
    shift_options = [
        {"code": "AUTO", "label": "Auto"},
        {"code": "VAC", "label": "VACACIONES"},
        {"code": "PERM", "label": "PERMISO"},
        {"code": "OFF", "label": "LIBRE"},
    ]
    shift_hours = {
        "OFF": 0, "VAC": 0, "PERM": 0
    }
    
    for s, hours in SHIFTS.items():
        if s in ["OFF", "VAC", "PERM", "AUTO"]: continue
        shift_hours[s] = len(hours)
        if s == "N_22-05":
            label = "Noche"
        else:
            parts = s.split("_")
            label = parts[1] if len(parts) > 1 else s
            if "+" in label: # For Q1_05-11+17-20 -> Q1
                label = parts[0]
        shift_options.append({"code": s, "label": label})
    
    return {
        "shift_sets": shift_sets,
        "shift_options": shift_options,
        "shift_hours": shift_hours,
        "bounds": bounds,
        "soft_bounds": soft_bounds,
        "standard_mode": standard_mode,
        "active_employees": active_count
    }

@app.get("/api/export_excel")
def export_excel(history_index: Optional[int] = None):
    db = load_db()
    
    target_schedule = {}
    target_tasks = {}
    
    if history_index is not None:
        # Export from History
        history_list = db.get("history_log", [])
        if 0 <= history_index < len(history_list):
            entry = history_list[history_index]
            target_schedule = entry.get("schedule", {})
            target_tasks = entry.get("daily_tasks", {})
    else:
        # Export Last Result
        last_result = db.get("last_result", {})
        target_schedule = last_result.get("schedule", {})
        target_tasks = last_result.get("daily_tasks", {}) 
    
    if not target_schedule:
        pass

    from openpyxl.styles import Border, Side
    from scheduler_engine import DAYS, SHIFTS

    # Try to load VBA template for format propagation macros
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "formato_template.xlsm")
    use_vba = os.path.exists(template_path)
    
    if use_vba:
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active
        if ws.title != "Horario":
            ws.title = "Horario"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horario"
    
    # ========================
    # COLOR PALETTE for employees (rotating, distinct colors)
    # ========================
    palette = [
        "D6E4F0",  # Light Blue
        "E2EFDA",  # Light Green
        "FCE4D6",  # Light Peach
        "EDEDED",  # Light Gray
        "D9E2F3",  # Lavender
        "FFF2CC",  # Light Yellow
        "E2D9F3",  # Light Purple
        "D5F5E3",  # Mint
        "FADBD8",  # Light Pink
        "D4EFDF",  # Sage
        "F9E79F",  # Gold
        "AED6F1",  # Sky Blue
    ]
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ========================
    # SHEET 1: HORARIO PRINCIPAL
    # ========================
    headers = ["Colaborador"] + DAYS + ["Horas"]
    ws.append(headers)
    
    # Header Style
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Employee Data Rows
    emp_names = list(target_schedule.keys())
    
    # Precalculate colors per employee for consistency across all sections
    emp_colors = {name: palette[idx % len(palette)] for idx, name in enumerate(emp_names)}
    
    for idx, name in enumerate(emp_names):
        shifts = target_schedule[name]
        row_data = [name]
        total_hours = 0
        
        for d in DAYS:
            s_code = shifts.get(d, "OFF")
            readable_shift = format_shift_code(s_code)
            row_data.append(readable_shift)
            
            # Calculate hours
            hours = len(SHIFTS.get(s_code, set()))
            total_hours += hours
            
        row_data.append(total_hours)
        ws.append(row_data)
        
        # Style the row
        current_row = ws.max_row
        color = emp_colors[name]
        name_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            
            if col_idx == 1:
                # Name cell: bold, colored, left-aligned — THIS IS THE FORMAT KEY
                cell.fill = name_fill
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(vertical="center", horizontal="left")
            else:
                # Schedule cells inherit the name's color
                light_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = light_fill
                
                # Highlight special shifts
                s_code = shifts.get(DAYS[col_idx - 2], "OFF") if col_idx <= 8 else None
                if s_code == "OFF":
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.font = Font(color="999999", italic=True)
                elif s_code == "VAC":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif s_code == "N_22-05":
                    cell.fill = PatternFill(start_color="2F2F5F", end_color="2F2F5F", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif col_idx == len(row_data):
                    # Hours column
                    cell.font = Font(bold=True, size=11)
        
    # Column Widths
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, 9):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(9)].width = 8
    
    # ========================
    # FORMATO SECTION (Column K) — Visual legend linked to both tables
    # ========================
    formato_col = 11  # Column K
    
    # Header
    fmt_header = ws.cell(row=1, column=formato_col, value="FORMATO")
    fmt_header.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    fmt_header.font = Font(bold=True, color="FFFFFF", size=11)
    fmt_header.alignment = Alignment(horizontal="center", vertical="center")
    fmt_header.border = thin_border
    
    # One cell per employee with the same format as their row in both tables
    for idx, name in enumerate(emp_names):
        fmt_row = idx + 2  # Row 2 onwards (row 1 is header)
        color = emp_colors[name]
        fmt_cell = ws.cell(row=fmt_row, column=formato_col, value=name)
        fmt_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fmt_cell.font = Font(bold=True, size=11)
        fmt_cell.alignment = Alignment(vertical="center", horizontal="left")
        fmt_cell.border = thin_border
    
    # LIBRE format entry (after all employees)
    libre_row = len(emp_names) + 2
    libre_cell = ws.cell(row=libre_row, column=formato_col, value="LIBRE")
    libre_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    libre_cell.font = Font(color="999999", italic=True, size=11)
    libre_cell.alignment = Alignment(vertical="center", horizontal="left")
    libre_cell.border = thin_border
    
    ws.column_dimensions[openpyxl.utils.get_column_letter(formato_col)].width = 18

    # ========================
    # SUB-TABLE: OBLIGACIONES / LIMPIEZA
    # ========================
    # Leave 2 blank rows as separator
    separator_row = ws.max_row + 2
    
    # Title row
    ws.cell(row=separator_row, column=1, value="OBLIGACIONES / LIMPIEZA")
    title_cell = ws.cell(row=separator_row, column=1)
    title_cell.font = Font(bold=True, size=13, color="2F5496")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=separator_row, start_column=1, end_row=separator_row, end_column=9)
    
    # Task Headers
    task_header_row = separator_row + 1
    task_headers = ["Colaborador"] + DAYS
    for col_idx, header in enumerate(task_headers, 1):
        cell = ws.cell(row=task_header_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Task Data Rows
    for idx, name in enumerate(emp_names):
        emp_tasks = target_tasks.get(name, {})
        task_row_num = task_header_row + 1 + idx
        color = emp_colors[name]
        name_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Name cell
        name_cell = ws.cell(row=task_row_num, column=1, value=name)
        name_cell.fill = name_fill
        name_cell.font = Font(bold=True, size=10)
        name_cell.alignment = Alignment(vertical="center", horizontal="left")
        name_cell.border = thin_border
        
        for d_idx, d in enumerate(DAYS):
            task = emp_tasks.get(d)
            col = d_idx + 2
            
            if task:
                # Clean up task text for Excel
                task_text = task
                if "↑AM" in task_text or "↓PM" in task_text:
                    task_text = task_text.replace("↑AM", "(AM)").replace("↓PM", "(PM)")
                
                cell = ws.cell(row=task_row_num, column=col, value=task_text)
                
                # Color-code by task type
                if "Baños" in task:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(color="B45309", bold=True, size=10)
                elif "Tanques" in task:
                    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                    cell.font = Font(color="1D4ED8", bold=True, size=10)
                elif "Oficina" in task:
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    cell.font = Font(color="BE185D", bold=True, size=10)
                else:
                    cell.font = Font(size=10)
            else:
                cell = ws.cell(row=task_row_num, column=col, value="—")
                cell.font = Font(color="CCCCCC", size=10)
            
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = thin_border

    # Save to temp
    suffix = ".xlsm" if use_vba else ".xlsx"
    filename = "horario" + suffix
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    wb.save(tmp.name)
    tmp.close()
    
    return FileResponse(tmp.name, filename=filename)

# ==============================================================================
# PLANILLAS API ENDPOINTS (Shared with new unified Web UI)
# ==============================================================================

class PlanillaEmpleado(BaseModel):
    nombre: str
    tipo_pago: str
    salario_fijo: Optional[float] = None
    cedula: Optional[str] = None
    correo: Optional[str] = None
    telefono: Optional[str] = None
    fecha_inicio: Optional[str] = None
    aplica_seguro: Optional[int] = 1
    genero: Optional[str] = 'M'
    puede_nocturno: Optional[int] = 1
    forced_libres: Optional[int] = 0
    forced_quebrado: Optional[int] = 0
    allow_no_rest: Optional[int] = 0
    es_jefe_pista: Optional[int] = 0
    strict_preferences: Optional[int] = 0
    turnos_fijos: Optional[str] = "{}"

class PlanillaVacacion(BaseModel):
    empleado_id: int
    fecha_inicio: str
    fecha_fin: str
    dias: float
    fecha_reingreso: Optional[str] = None
    notas: Optional[str] = None

class PlanillaTarifas(BaseModel):
    tarifa_diurna: float
    tarifa_nocturna: float
    tarifa_mixta: float
    seguro: float

class PlanillaMes(BaseModel):
    anio: int
    mes: int

class PlanillaSemana(BaseModel):
    mes_id: int
    viernes: str

class PlanillaBoletasRequest(BaseModel):
    semana_nombre: str

class PlanillaImportarHorario(BaseModel):
    horario_id: int
    semana_nombre: str
    sync_empleados: bool = True

@app.get("/api/planillas/empleados")
def get_planillas_empleados(solo_activos: bool = True):
    return plan_db.get_empleados(solo_activos)

@app.post("/api/planillas/empleados")
def add_planilla_empleado(emp: PlanillaEmpleado):
    ok, msg = plan_db.add_empleado(
        emp.nombre, emp.tipo_pago, emp.salario_fijo, 
        cedula=emp.cedula, correo=emp.correo, 
        telefono=emp.telefono, fecha_inicio=emp.fecha_inicio,
        aplica_seguro=emp.aplica_seguro, genero=emp.genero,
        puede_nocturno=emp.puede_nocturno,
        forced_libres=emp.forced_libres, forced_quebrado=emp.forced_quebrado,
        allow_no_rest=emp.allow_no_rest, es_jefe_pista=emp.es_jefe_pista,
        strict_preferences=emp.strict_preferences, turnos_fijos=emp.turnos_fijos
    )
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"status": "success", "message": msg}

@app.put("/api/planillas/empleados/{emp_id}")
def update_planilla_empleado(emp_id: int, emp: PlanillaEmpleado):
    plan_db.update_empleado(
        emp_id, nombre=emp.nombre, tipo_pago=emp.tipo_pago, salario_fijo=emp.salario_fijo,
        cedula=emp.cedula, correo=emp.correo, 
        telefono=emp.telefono, fecha_inicio=emp.fecha_inicio,
        aplica_seguro=emp.aplica_seguro, genero=emp.genero,
        puede_nocturno=emp.puede_nocturno,
        forced_libres=emp.forced_libres, forced_quebrado=emp.forced_quebrado,
        allow_no_rest=emp.allow_no_rest, es_jefe_pista=emp.es_jefe_pista,
        strict_preferences=emp.strict_preferences, turnos_fijos=emp.turnos_fijos
    )
    return {"status": "success"}

@app.delete("/api/planillas/empleados/{emp_id}")
def delete_planilla_empleado(emp_id: int):
    plan_db.delete_empleado(emp_id)
    return {"status": "success"}

@app.get("/api/planillas/vacaciones/{emp_id}")
def get_planillas_vacaciones(emp_id: int):
    vacs = plan_db.get_vacaciones(emp_id)
    emp_data = next((e for e in plan_db.get_empleados(solo_activos=False) if e["id"] == emp_id), None)
    if emp_data:
        fi = emp_data.get("fecha_inicio")
        acumulados = plan_db.calcular_dias_vacaciones(fi) if fi else 0
        tomados = plan_db.total_dias_vacaciones_tomados(emp_id)
        return {
            "registros": vacs,
            "acumulados": acumulados,
            "tomados": tomados,
            "disponibles": max(0, acumulados - tomados)
        }
    return {"registros": vacs}

@app.post("/api/planillas/vacaciones")
def add_planilla_vacacion(vac: PlanillaVacacion):
    plan_db.add_vacacion(
        vac.empleado_id, vac.fecha_inicio, vac.fecha_fin, vac.dias,
        fecha_reingreso=vac.fecha_reingreso, notas=vac.notas
    )
    return {"status": "success"}

@app.delete("/api/planillas/vacaciones/{vac_id}")
def delete_planilla_vacacion(vac_id: int):
    plan_db.delete_vacacion(vac_id)
    return {"status": "success"}

@app.put("/api/planillas/vacaciones/{vac_id}")
def update_planilla_vacacion(vac_id: int, vac: PlanillaVacacion):
    plan_db.update_vacacion(
        vac_id, vac.fecha_inicio, vac.fecha_fin, vac.dias,
        fecha_reingreso=vac.fecha_reingreso, notas=vac.notas
    )
    return {"status": "success"}

# ------------------------------------------------------------------------------
# LIQUIDACIÓN — Cálculo automático según Ley Costarricense
# ------------------------------------------------------------------------------
@app.get("/api/planillas/liquidacion/{emp_id}")
def calcular_liquidacion(emp_id: int):
    """Calcula todos los rubros de liquidación para un empleado.
    
    - Vacaciones: días disponibles × tarifa_diurna × 8hrs
    - Aguinaldo proporcional: salarios brutos del período ÷ 12
    - Cesantía (Art. 29 CT CR): según tabla de antigüedad
    - Preaviso: 1 mes de salario promedio si >1 año
    """
    from datetime import date, timedelta
    import math

    emp_data = next((e for e in plan_db.get_empleados(solo_activos=False) if e["id"] == emp_id), None)
    if not emp_data:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    nombre = emp_data.get("nombre", "")
    fi = emp_data.get("fecha_inicio")
    tarifas = plan_db.get_tarifas()
    tarifa_diurna = tarifas.get("tarifa_diurna", 0)

    # ── 1. VACACIONES ──
    acumulados = plan_db.calcular_dias_vacaciones(fi) if fi else 0
    tomados = plan_db.total_dias_vacaciones_tomados(emp_id)
    vac_dias = max(0, acumulados - tomados)
    vac_monto = vac_dias * tarifa_diurna * 8  # jornada ordinaria diurna

    # ── 2. ANTIGÜEDAD ──
    antiguedad_dias = 0
    antiguedad_anios = 0
    if fi:
        try:
            inicio = datetime.strptime(fi, "%Y-%m-%d").date()
            antiguedad_dias = (date.today() - inicio).days
            antiguedad_anios = antiguedad_dias / 365.25
        except:
            pass

    # ── 3. SALARIO PROMEDIO MENSUAL (de planillas reales) ──
    anio_actual = date.today().year
    salario_total_anual = 0.0
    semanas_con_datos = 0

    for anio in [anio_actual - 1, anio_actual]:
        meses = plan_db.get_meses_del_anio(anio)
        mes_activo = plan_db.get_mes_activo()
        if mes_activo and mes_activo['anio'] == anio:
            meses.append(mes_activo)

        for mes_data in meses:
            archivo = mes_data['archivo']
            archivo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas", archivo)
            if not os.path.exists(archivo_path):
                continue
            try:
                wb = openpyxl.load_workbook(archivo_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    if sheet_name.startswith("Semana "):
                        ws = wb[sheet_name]
                        for row in range(5, ws.max_row + 1):
                            name_cell = ws.cell(row=row, column=1).value
                            if name_cell and isinstance(name_cell, str) and name_cell.strip() == nombre:
                                # Find salario bruto column
                                for col in range(2, min(ws.max_column + 1, 20)):
                                    header = ws.cell(row=4, column=col).value
                                    if header and isinstance(header, str) and "bruto" in header.lower():
                                        val = ws.cell(row=row, column=col).value
                                        if val and isinstance(val, (int, float)):
                                            salario_total_anual += val
                                            semanas_con_datos += 1
                                        break
                wb.close()
            except:
                pass

    # Promedio mensual = total / meses con datos (aprox 4 semanas = 1 mes)
    meses_con_datos = max(1, semanas_con_datos / 4.33)
    salario_promedio_mensual = salario_total_anual / meses_con_datos if meses_con_datos > 0 else 0

    # ── 4. AGUINALDO PROPORCIONAL ──
    # En CR: aguinaldo = salarios de dic anterior a nov actual ÷ 12
    aguinaldo_monto = salario_total_anual / 12.0 if salario_total_anual > 0 else 0

    # ── 5. CESANTÍA (Art. 29 Código de Trabajo CR) ──
    # Solo aplica a despido sin justa causa
    cesantia_dias = 0
    if antiguedad_dias >= 90:  # Mínimo 3 meses
        if antiguedad_dias < 180:       # 3-6 meses
            cesantia_dias = 7
        elif antiguedad_dias < 365:     # 6-12 meses
            cesantia_dias = 14
        else:
            # Más de 1 año: tabla progresiva
            anios_completos = int(antiguedad_anios)
            if anios_completos == 1:
                cesantia_dias = 19.5
            elif anios_completos == 2:
                cesantia_dias = 19.5 + 19.5
            elif anios_completos == 3:
                cesantia_dias = 19.5 + 19.5 + 20
            elif anios_completos == 4:
                cesantia_dias = 19.5 + 19.5 + 20 + 20
            elif anios_completos == 5:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5
            elif anios_completos == 6:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5 + 21
            elif anios_completos >= 7:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5 + 21 + 21.24
            if anios_completos >= 8:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5 + 21 + 21.24 + 21.5
            # Máximo legal: ~22 días por año, tope 8 años

    salario_diario = salario_promedio_mensual / 30 if salario_promedio_mensual > 0 else 0
    cesantia_monto = cesantia_dias * salario_diario

    # ── 6. PREAVISO ──
    # >1 año: 1 mes; 3-6 meses: 1 semana; 6-12 meses: 15 días
    preaviso_monto = 0
    if antiguedad_dias >= 365:
        preaviso_monto = salario_promedio_mensual  # 1 mes
    elif antiguedad_dias >= 180:
        preaviso_monto = salario_promedio_mensual / 2  # 15 días
    elif antiguedad_dias >= 90:
        preaviso_monto = salario_promedio_mensual / 4.33  # 1 semana

    return {
        "emp_id": emp_id,
        "nombre": nombre,
        "fecha_inicio": fi,
        "antiguedad_anios": round(antiguedad_anios, 1),
        "tarifa_diurna": tarifa_diurna,
        "vacaciones_dias": vac_dias,
        "vacaciones_monto": round(vac_monto, 2),
        "aguinaldo_monto": round(aguinaldo_monto, 2),
        "cesantia_dias": round(cesantia_dias, 1),
        "cesantia_monto": round(cesantia_monto, 2),
        "preaviso_monto": round(preaviso_monto, 2),
        "salario_promedio_mensual": round(salario_promedio_mensual, 2),
        "total_despido": round(vac_monto + aguinaldo_monto + cesantia_monto + preaviso_monto, 2),
        "total_renuncia": round(vac_monto + aguinaldo_monto, 2),
    }


def calcular_aguinaldo(anio: int):
    """Calcula aguinaldo basado en la tabla salarios_mensuales de la BD."""
    MES_NOMBRES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                   7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

    desglose_rows = plan_db.get_salarios_anio_desglose(anio)
    if not desglose_rows:
        return {"status": "error", "message": f"No hay salarios registrados para {anio}", "data": []}

    # Group by employee
    emp_data = {}
    for row in desglose_rows:
        eid = row['empleado_id']
        if eid not in emp_data:
            emp_data[eid] = {
                'nombre': row['empleado_nombre'],
                'total_bruto': 0.0,
                'desglose': []
            }
        mes_label = MES_NOMBRES.get(row['mes'], f'Mes {row["mes"]}')
        emp_data[eid]['total_bruto'] += row['total_bruto']
        emp_data[eid]['desglose'].append({
            'mes': mes_label,
            'bruto': round(row['total_bruto'], 2)
        })

    # Get employee info from DB
    emps = plan_db.get_empleados(solo_activos=False)
    emp_lookup = {e['id']: e for e in emps}

    results = []
    total_aguinaldo = 0.0
    meses_set = set()

    for eid, data in emp_data.items():
        emp = emp_lookup.get(eid, {})
        sal_total = data['total_bruto']
        aguinaldo = sal_total / 12.0 if sal_total > 0 else 0.0
        total_aguinaldo += aguinaldo
        for d in data['desglose']:
            meses_set.add(d['mes'])
        results.append({
            "id": eid,
            "nombre": data['nombre'],
            "cedula": emp.get("cedula") or "—",
            "salario_anual": round(sal_total, 2),
            "aguinaldo": round(aguinaldo, 2),
            "desglose_mensual": data['desglose']
        })

    # Also include employees with no salary data
    for emp in emps:
        if emp['id'] not in emp_data:
            results.append({
                "id": emp['id'],
                "nombre": emp['nombre'],
                "cedula": emp.get("cedula") or "—",
                "salario_anual": 0,
                "aguinaldo": 0,
                "desglose_mensual": []
            })

    return {
        "status": "success",
        "data": results,
        "total_aguinaldo": round(total_aguinaldo, 2),
        "meses_evaluados": len(meses_set)
    }


@app.get("/api/planillas/aguinaldo/{anio}")
def get_aguinaldo(anio: int):
    return calcular_aguinaldo(anio)


# ------------------------------------------------------------------------------
# PLANILLAS - GUARDAR SALARIOS
# ------------------------------------------------------------------------------
class SalarioSemanal(BaseModel):
    empleado_id: int
    empleado_nombre: str
    anio: int
    mes: int
    semana: int
    salario_bruto: float

class SalariosLote(BaseModel):
    salarios: list[SalarioSemanal]

@app.post("/api/planillas/salarios")
def guardar_salarios(req: SalariosLote):
    for s in req.salarios:
        plan_db.guardar_salario_semanal(
            s.empleado_id, s.empleado_nombre,
            s.anio, s.mes, s.semana, s.salario_bruto
        )
    return {"status": "success", "count": len(req.salarios)}

@app.post("/api/planillas/sincronizar-aguinaldo/{anio}")
def sincronizar_aguinaldo_anio(anio: int):
    """
    Escanea los Excel (cerrados y activo) para el año dado, lee todas las hojas 
    de semanas y guarda los salarios brutos en la DB para reconstruir el historial.
    """
    meses_a_sincronizar = plan_db.get_meses_del_anio(anio)
    mes_activo = plan_db.get_mes_activo()
    if mes_activo and mes_activo["anio"] == anio:
        if not any(m["archivo"] == mes_activo["archivo"] for m in meses_a_sincronizar):
            meses_a_sincronizar.append(mes_activo)

    # Deduplicar por archivo — evita procesar el mismo Excel dos veces si
    # quedaron registros duplicados de meses en la base de datos.
    seen_archivos = set()
    meses_unicos = []
    for m in meses_a_sincronizar:
        if m["archivo"] not in seen_archivos:
            seen_archivos.add(m["archivo"])
            meses_unicos.append(m)
    meses_a_sincronizar = meses_unicos

    if not meses_a_sincronizar:
        return {"status": "error", "message": f"No se encontraron meses para el año {anio} para sincronizar."}
    
    count = 0
    skipped = []
    errores = []
    tarifas = plan_db.get_tarifas()
    emps = plan_db.get_empleados(solo_activos=False)
    emp_map = {e["nombre"]: e["id"] for e in emps}

    for mes in meses_a_sincronizar:
        archivo_path = os.path.join(_planillas_dir, mes["archivo"])
        if not os.path.exists(archivo_path):
            # Archivo no existe: ignorar silenciosamente (puede ser un mes
            # registrado por error en la DB sin Excel real).
            skipped.append(mes["archivo"])
            continue
            
        try:
            wb = openpyxl.load_workbook(archivo_path, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith("Semana "):
                    try:
                        sem_num = int(sheet_name.split(" ")[-1])
                    except:
                        continue
                    
                    ws = wb[sheet_name]
                    bruto_col = 12 # Default based on rellenar_horas_en_excel logic and template
                    
                    for r in range(5, ws.max_row + 1):
                        name_cell = ws.cell(row=r, column=1).value
                        if name_cell and isinstance(name_cell, str) and name_cell.strip() in emp_map:
                            nombre = name_cell.strip()
                            bruto_val = ws.cell(row=r, column=bruto_col).value
                            if bruto_val and isinstance(bruto_val, (int, float)):
                                plan_db.guardar_salario_semanal(
                                    emp_map[nombre], nombre, anio, mes["mes"], sem_num, bruto_val
                                )
                                count += 1
            wb.close()
        except Exception as e:
            errores.append(f"Error procesando {mes['archivo']}: {str(e)}")

    msg = f"Sincronizados {count} registros de salarios."
    if skipped:
        print(f"Sincronización: {len(skipped)} archivos omitidos (no existen): {skipped}")
    if errores:
        msg += f" ({len(errores)} errores, revisa la consola para detalles)"
        print("Errores de sincronización:", errores)
        
    return {"status": "success", "message": msg, "count": count}


# ------------------------------------------------------------------------------
# PLANILLAS - TARIFAS
# ------------------------------------------------------------------------------
@app.get("/api/planillas/tarifas")
def get_planillas_tarifas():
    return plan_db.get_tarifas()

@app.post("/api/planillas/tarifas")
def update_planillas_tarifas(t: PlanillaTarifas):
    plan_db.set_tarifas(t.tarifa_diurna, t.tarifa_nocturna, t.tarifa_mixta, t.seguro)
    return {"status": "success"}

# ------------------------------------------------------------------------------
# PLANILLAS - MESES & SEMANAS
# ------------------------------------------------------------------------------
@app.get("/api/planillas/meses")
def get_todos_meses():
    meses = plan_db.get_todos_meses()
    for m in meses:
        m["semanas"] = plan_db.get_semanas_del_mes(m["id"])
    return meses

@app.get("/api/planillas/meses/activo")
def get_mes_activo():
    mes = plan_db.get_mes_activo()
    if not mes:
        return {"mes": None, "semanas": []}
    semanas = plan_db.get_semanas_del_mes(mes["id"])
    return {"mes": mes, "semanas": semanas}

@app.post("/api/planillas/meses")
def create_mes(m: PlanillaMes):
    emps = plan_db.get_empleados()
    if not emps:
        raise HTTPException(status_code=400, detail="Agregue empleados antes de crear la planilla.")
    
    # Crear Excel del Mes
    meses_nombres = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    mes_name = meses_nombres.get(m.mes, str(m.mes))
    
    # Create nested folder structure: Planillas {Anio} / {Mes}
    folder_year = f"Planillas {m.anio}"
    folder_month = mes_name
    rel_folder_path = os.path.join(folder_year, folder_month)
    abs_folder_path = os.path.join(_planillas_dir, rel_folder_path)
    os.makedirs(abs_folder_path, exist_ok=True)
    
    archivo_name = f"Planilla_{mes_name}_{m.anio}.xlsx"
    archivo_rel_path = os.path.join(rel_folder_path, archivo_name).replace("\\", "/")
    archivo_path = os.path.join(abs_folder_path, archivo_name)
    
    # Init Excel
    tarifas = plan_db.get_tarifas()
    pl_module.TARIFA_DIURNA = tarifas["tarifa_diurna"]
    pl_module.TARIFA_NOCTURNA = tarifas["tarifa_nocturna"]
    pl_module.TARIFA_MIXTA = tarifas["tarifa_mixta"]
    
    wb = openpyxl.Workbook()
    pl_module.SAMPLE_EMPS = []
    pl_module.crear_catalogo(wb)
    
    ws_cat = wb["Catalogo"]
    tipo_map = {"tarjeta": "Tarjeta", "efectivo": "Efectivo", "fijo": "Salario Fijo"}
    for i, emp in enumerate(emps):
        r = 5 + i
        pl_module.sc(ws_cat, r, 1, emp["nombre"], pl_module.f_data, pl_module.ROW_1 if i % 2 == 0 else pl_module.ROW_2, pl_module.al_l, pl_module.borde)
        pl_module.sc(ws_cat, r, 2, tipo_map.get(emp["tipo_pago"], emp["tipo_pago"]),
                     pl_module.f_data, pl_module.ROW_1 if i % 2 == 0 else pl_module.ROW_2, pl_module.al_c, pl_module.borde)
        if emp["salario_fijo"]:
            pl_module.sc(ws_cat, r, 3, emp["salario_fijo"],
                         pl_module.f_data, pl_module.ROW_1 if i % 2 == 0 else pl_module.ROW_2, pl_module.al_c, pl_module.borde, pl_module.MONEY)
                         
    pl_module.crear_resumen_mensual(wb)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(archivo_path)
    
    # Registrar en DB
    mes_db = plan_db.crear_mes(m.anio, m.mes, archivo_rel_path)
    return {"status": "success", "mes": mes_db}

@app.post("/api/planillas/meses/{mes_id}/cerrar")
def cerrar_mes(mes_id: int):
    plan_db.cerrar_mes(mes_id)
    return {"status": "success"}

@app.post("/api/planillas/semanas")
def agregar_semana(s: PlanillaSemana):
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo or mes_activo["id"] != s.mes_id:
        raise HTTPException(status_code=400, detail="Mes inválido o cerrado.")
        
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=400, detail=f"No se encontró el Excel: {mes_activo['archivo']}")
        
    viernes_date = datetime.datetime.strptime(s.viernes, "%Y-%m-%d").date()
    
    tarifas = plan_db.get_tarifas()
    pl_module.TARIFA_DIURNA = tarifas["tarifa_diurna"]
    pl_module.TARIFA_NOCTURNA = tarifas["tarifa_nocturna"]
    pl_module.TARIFA_MIXTA = tarifas["tarifa_mixta"]
    
    wb = openpyxl.load_workbook(archivo_path)
    empleados = pl_module.leer_catalogo(wb)
    total_emp = sum(len(v) for v in empleados.values())
    if total_emp == 0:
        raise HTTPException(status_code=400, detail="No hay empleados en el Excel.")
        
    num = pl_module.contar_semanas(wb) + 1
    sem_num = pl_module.num_semana_anual(viernes_date)
    
    nombre_hoja, gran_row, section_totals = pl_module.crear_hoja_semanal(
        wb, num, viernes_date, empleados, seguro=tarifas["seguro"])
    pl_module.crear_resumen_semanal(wb, nombre_hoja, sem_num, viernes_date)
    pl_module.crear_resumen_mensual(wb)
    pl_module.crear_dashboard(wb)
    
    wb.save(archivo_path)
    
    # Save to DB
    plan_db.add_semana(s.mes_id, sem_num, s.viernes)
    semanas = plan_db.get_semanas_del_mes(s.mes_id)
    return {"status": "success", "semanas": semanas}

@app.delete("/api/planillas/semanas/{semana_id}")
def delete_semana(semana_id: int):
    """Elimina una semana de cualquier mes (activo o cerrado).
    
    Busca el mes al que pertenece la semana entre TODOS los meses registrados,
    no solo el activo. Esto permite editar el historial de planillas anteriores.
    """
    # Buscar la semana y su mes en todos los meses registrados
    all_months = plan_db.get_todos_meses()
    target_sem = None
    target_mes = None

    for mes in all_months:
        semanas = plan_db.get_semanas_del_mes(mes["id"])
        found = next((s for s in semanas if s["id"] == semana_id), None)
        if found:
            target_sem = found
            target_mes = mes
            break

    if not target_sem or not target_mes:
        raise HTTPException(status_code=404, detail="Semana no encontrada")

    num_semana = target_sem["num_semana"]
    plan_db.delete_semana(semana_id)

    # Eliminar hojas del Excel si el archivo existe
    archivo_path = os.path.join(_planillas_dir, target_mes["archivo"])
    if os.path.exists(archivo_path):
        try:
            wb = openpyxl.load_workbook(archivo_path)
            sheet_sem = f"Semana {num_semana}"
            sheet_res = f"Res. Sem. {num_semana}"
            if sheet_sem in wb.sheetnames: wb.remove(wb[sheet_sem])
            if sheet_res in wb.sheetnames: wb.remove(wb[sheet_res])
            pl_module.crear_resumen_mensual(wb)
            pl_module.crear_dashboard(wb)
            wb.save(archivo_path)
        except Exception as e:
            print(f"Advertencia al modificar Excel: {e}")

    return {"status": "success", "semanas": plan_db.get_semanas_del_mes(target_mes["id"])}


@app.delete("/api/planillas/meses/{mes_id}")
def delete_mes_historial(mes_id: int):
    """Elimina un mes cerrado y todos sus registros. Rechaza meses activos."""
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")
    if not mes_target.get("cerrado"):
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar meses cerrados")

    # Intentar borrar el Excel físico (no fatal si falla)
    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if os.path.exists(archivo_path):
        try:
            os.remove(archivo_path)
        except Exception as e:
            print(f"Advertencia: no se pudo eliminar el archivo Excel: {e}")

    plan_db.delete_mes(mes_id)
    return {"status": "success", "message": f"Mes {mes_id} eliminado"}


@app.post("/api/planillas/meses/{mes_id}/semanas")
def agregar_semana_a_mes_historico(mes_id: int, s: PlanillaSemana):
    """Agrega una semana a cualquier mes (activo o cerrado), para edición del historial.
    
    A diferencia del endpoint /api/planillas/semanas, este no exige que el mes
    sea el activo, permitiendo gestionar semanas en planillas anteriores.
    """
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")

    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=400, detail=f"No se encontró el Excel: {mes_target['archivo']}")

    viernes_date = datetime.datetime.strptime(s.viernes, "%Y-%m-%d").date()

    tarifas = plan_db.get_tarifas()
    pl_module.TARIFA_DIURNA = tarifas["tarifa_diurna"]
    pl_module.TARIFA_NOCTURNA = tarifas["tarifa_nocturna"]
    pl_module.TARIFA_MIXTA = tarifas["tarifa_mixta"]

    wb = openpyxl.load_workbook(archivo_path)
    empleados = pl_module.leer_catalogo(wb)
    total_emp = sum(len(v) for v in empleados.values())
    if total_emp == 0:
        raise HTTPException(status_code=400, detail="No hay empleados en el Excel.")

    num = pl_module.contar_semanas(wb) + 1
    sem_num = pl_module.num_semana_anual(viernes_date)

    nombre_hoja, gran_row, section_totals = pl_module.crear_hoja_semanal(
        wb, num, viernes_date, empleados, seguro=tarifas["seguro"])
    pl_module.crear_resumen_semanal(wb, nombre_hoja, sem_num, viernes_date)
    pl_module.crear_resumen_mensual(wb)
    pl_module.crear_dashboard(wb)

    wb.save(archivo_path)

    plan_db.add_semana(mes_id, sem_num, s.viernes)
    semanas = plan_db.get_semanas_del_mes(mes_id)
    return {"status": "success", "semanas": semanas}

@app.get("/api/planillas/horarios-disponibles")
def get_horarios_disponibles():
    horarios = horario_db.get_horarios_generados()
    return {"status": "success", "horarios": horarios}

@app.post("/api/planillas/semanas/importar")
def importar_horario_semana(req: PlanillaImportarHorario):
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo:
        raise HTTPException(status_code=400, detail="No hay un mes activo")
        
    horario = horario_db.get_horario_por_id(req.horario_id)
    if not horario:
        raise HTTPException(status_code=404, detail="Horario no encontrado")
        
    if req.sync_empleados:
        horario_db.sincronizar_empleados_a_planilla()
        
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    ok, msg, hours_data = horario_db.rellenar_horas_en_excel(
        archivo_path, req.semana_nombre, horario["horario"]
    )
    
    if not ok:
        raise HTTPException(status_code=400, detail=msg)

    # Store salaries in DB
    try:
        tarifas = plan_db.get_tarifas()
        emps = plan_db.get_empleados(solo_activos=False)
        emp_map = {e["nombre"]: e["id"] for e in emps}
        
        # Get week number from semantic name "Semana X"
        try:
            sem_num = int(req.semana_nombre.split(" ")[-1])
        except:
            sem_num = 0

        for emp_nombre, h in hours_data.items():
            if emp_nombre in emp_map:
                bruto = (h["diurnas"] * tarifas["tarifa_diurna"]) + \
                        (h["mixtas"] * tarifas["tarifa_mixta"]) + \
                        (h["nocturnas"] * tarifas["tarifa_nocturna"]) + \
                        (h["extra"] * (tarifas["tarifa_diurna"] * 1.5)) # Extra typically 1.5x
                
                plan_db.guardar_salario_semanal(
                    emp_map[emp_nombre], emp_nombre,
                    mes_activo["anio"], mes_activo["mes"], sem_num, bruto
                )
    except Exception as e:
        print(f"Error saving salaries: {e}")

    return {"status": "success", "message": msg}

@app.post("/api/planillas/boletas/generar")
def generar_boletas(req: PlanillaBoletasRequest):
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo:
        raise HTTPException(status_code=400, detail="No active month to generate boletas.")
        
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    logo_path = os.path.join(_planillas_dir, "logo.png")
    
    ok, msg, out_dir = gb_module.generar_boletas_semana(archivo_path, req.semana_nombre, logo_path)
    
    if ok:
        # Open folder dynamically on Windows
        if os.name == 'nt':
            os.startfile(out_dir)
        return {"status": "success", "message": msg, "out_dir": out_dir}
    else:
        raise HTTPException(status_code=400, detail=msg)

@app.get("/api/planillas/excel/abrir")
def abrir_excel_activo():
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo:
        raise HTTPException(status_code=400, detail="No hay mes activo.")
    
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=404, detail="No se encontró el archivo Excel.")
        
    if os.name == 'nt':
        os.startfile(archivo_path)
    return {"status": "success"}

@app.get("/api/planillas/excel/abrir/{mes_id}")
def abrir_excel_por_id(mes_id: int):
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")
        
    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=404, detail="Archivo Excel no encontrado")
        
    if os.name == 'nt':
        os.startfile(archivo_path)
    return {"status": "success"}

# ==============================================================================
# UTILIDADES — GENERADOR DE DOCUMENTOS WORD
# ==============================================================================
import docx_generator

class DocPrestamo(BaseModel):
    emp_id: int
    monto_total: float
    pago_semanal: float

class AmoDatoItem(BaseModel):
    fecha: Optional[str] = None
    monto: Optional[float] = None
    minutos: Optional[int] = None

class DocAmonestacion(BaseModel):
    emp_id: int
    tipo: str  # "faltantes", "tardanzas", "conductas"
    datos: List[AmoDatoItem] = []

class DocVacacionesReq(BaseModel):
    emp_id: int
    tipo: str  # 'total' or 'parcial'
    fecha_inicio: str
    fecha_reingreso: str

class DocLiquidacion(BaseModel):
    emp_id: int
    vacaciones_dias: float
    vacaciones_monto: float
    aguinaldo_monto: float
    cesantia_monto: float = 0
    preaviso_monto: float = 0
    total_pagar: float
    modo_pago: str  # "Abonos" o "Total"

def _get_emp_info(emp_id: int):
    emps = plan_db.get_empleados(solo_activos=False)
    emp = next((e for e in emps if e["id"] == emp_id), None)
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    return emp.get("nombre", ""), emp.get("cedula", "")

@app.post("/api/utilidades/prestamo")
def generar_doc_prestamo(req: DocPrestamo):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas/logo.png")
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../")
    path = docx_generator.generar_prestamo(nombre, cedula, req.monto_total, req.pago_semanal, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/amonestacion")
def generar_doc_amonestacion(req: DocAmonestacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas/logo.png")
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../")
    datos_dict = [d.dict(exclude_none=True) for d in req.datos]
    path = docx_generator.generar_amonestacion(nombre, cedula, req.tipo, datos_dict, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/vacaciones")
def generar_doc_vacaciones(req: DocVacacionesReq):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas/logo.png")
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../")
    path = docx_generator.generar_vacaciones(nombre, cedula, req.tipo, req.fecha_inicio, req.fecha_reingreso, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/despido")
def generar_doc_despido(req: DocLiquidacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas/logo.png")
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../")
    path = docx_generator.generar_liquidacion("Despido", nombre, cedula, req.vacaciones_dias, req.vacaciones_monto, req.aguinaldo_monto, req.cesantia_monto, req.preaviso_monto, req.total_pagar, req.modo_pago, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/renuncia")
def generar_doc_renuncia(req: DocLiquidacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas/logo.png")
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../")
    path = docx_generator.generar_liquidacion("Renuncia", nombre, cedula, req.vacaciones_dias, req.vacaciones_monto, req.aguinaldo_monto, 0, 0, req.total_pagar, req.modo_pago, logo, base)
    return {"status": "success", "path": path}

class DocRecomendacion(BaseModel):
    emp_id: int
    puesto: str
    texto_adicional: str = ""

@app.post("/api/utilidades/recomendacion")
def generar_doc_recomendacion(req: DocRecomendacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    emps = plan_db.get_empleados(solo_activos=False)
    emp = next((e for e in emps if e['id'] == req.emp_id), None)
    fecha_inicio = emp.get("fecha_inicio", "") if emp else ""
    logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../planillas/logo.png")
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../")
    path = docx_generator.generar_recomendacion(nombre, cedula, req.puesto, fecha_inicio, req.texto_adicional, logo, base)
    return {"status": "success", "path": path}

# ==============================================================================
# Serve Frontend
# ==============================================================================
frontend_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../frontend")
if os.path.exists(frontend_path):
    # Mount frontend directory for static files (React/Vue/JS/CSS)
    # Using 'html=True' lets it serve index.html for the root automatically.
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")
else:
    print(f"Warning: Frontend path {frontend_path} not found")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)