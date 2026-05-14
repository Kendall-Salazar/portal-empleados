
import openpyxl
import os
import sys

def find_special_chars(path):
    sys.stdout.reconfigure(encoding='utf-8')
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit(): continue
        ws = wb[sheet_name]
        print(f"--- {sheet_name} ---")
        for row in ws.iter_rows(min_row=10, max_row=80, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    # Check for common arrow characters or anything non-ascii
                    if any(ord(c) > 127 for c in cell):
                        print(f"Found special: {cell} (hex: {[hex(ord(c)) for c in cell]})")

if __name__ == "__main__":
    find_special_chars(r"C:\Users\kenda\Downloads\horario 2026 (2).xlsx")
