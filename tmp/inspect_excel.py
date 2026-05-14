
import sys
import os
import openpyxl
import traceback

def inspect_excel(path):
    try:
        if not os.path.exists(path):
            print(f"File not found: {path}")
            return

        # Use a safer way to print unicode
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')

        wb = openpyxl.load_workbook(path, data_only=True)
        # Fix: look for numeric sheets or others
        target_sheets = [s for s in wb.sheetnames if s.isdigit()]
        if not target_sheets:
            target_sheets = wb.sheetnames[:5]

        for sheet_name in target_sheets:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            # Print only up to 150 rows to avoid too much data
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=150, values_only=True), 1):
                if any(v is not None for v in row):
                    # Clean up values for safe printing
                    row_str = ", ".join(str(v).strip() if v is not None else "None" for v in row)
                    # Use ASCII representation for things that still fail if needed, but utf-8 should work
                    print(f"Row {row_idx}: {row_str}")
    except Exception:
        traceback.print_exc()

if __name__ == "__main__":
    excel_path = r"C:\Users\kenda\Downloads\horario 2026 (2).xlsx"
    inspect_excel(excel_path)
