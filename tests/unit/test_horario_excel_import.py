"""Parser e import Excel de horarios al historial (layouts split_friday y linear)."""
import datetime
import json
import os
import sys
import openpyxl
import pytest

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_PLAN = os.path.join(_ROOT, "planillas")
_BACKEND = os.path.join(_ROOT, "backend")
for _p in (_PLAN, _BACKEND):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import horario_excel_import as hei  # noqa: E402


def _workbook_split_friday(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "10"
    ws["A1"] = "SEMANA 10"
    ws["B2"] = datetime.datetime(2026, 2, 28)
    ws["C2"] = datetime.datetime(2026, 3, 1)
    ws["D2"] = datetime.datetime(2026, 3, 2)
    ws["E2"] = datetime.datetime(2026, 3, 3)
    ws["F2"] = datetime.datetime(2026, 3, 4)
    ws["G2"] = datetime.datetime(2026, 3, 5)
    ws["J2"] = datetime.datetime(2026, 2, 27)
    ws["A3"] = "Colaborador"
    ws["B3"] = "Sáb"
    ws["C3"] = "Dom"
    ws["D3"] = "Lun"
    ws["E3"] = "Mar"
    ws["F3"] = "Mié"
    ws["G3"] = "Jue"
    ws["I3"] = "Colaborador"
    ws["J3"] = "Vie"
    ws["A4"] = "Ana Perez"
    ws["B4"] = "LIBRE"
    ws["C4"] = "VACACIONES"
    ws["J4"] = "LIBRE"
    ws["A5"] = "Carlos"
    ws["B5"] = "XYZ_NO_EXISTE"
    wb.save(path)
    wb.close()


def _workbook_linear_vie_jue(path: str, with_formato: bool) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "11"
    ws["A1"] = "SEMANA 11"
    d0 = datetime.datetime(2026, 2, 27)
    for i in range(7):
        ws.cell(row=2, column=2 + i, value=d0 + datetime.timedelta(days=i))
    ws["A3"] = "Colaborador"
    days = ["Vie", "Sáb", "Dom", "Lun", "Mar", "Mié", "Jue"]
    for i, day in enumerate(days):
        ws.cell(row=3, column=2 + i, value=day)
    if with_formato:
        ws["J3"] = "FORMATO"
    ws["A4"] = "Bob"
    ws["B4"] = "LIBRE"
    ws["C4"] = "PERMISO"
    wb.save(path)
    wb.close()


def _workbook_linear_strings_dates(path: str) -> None:
    """Como hoja 13: fechas como texto dd/mm/yyyy, sin columna FORMATO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "13"
    ws["A1"] = "Semana 13"
    dates = [
        "27/02/2026",
        "28/02/2026",
        "01/03/2026",
        "02/03/2026",
        "03/03/2026",
        "04/03/2026",
        "05/03/2026",
    ]
    for i, ds in enumerate(dates):
        ws.cell(row=2, column=2 + i, value=ds)
    ws["A3"] = "Colaborador"
    for i, day in enumerate(["Vie", "Sáb", "Dom", "Lun", "Mar", "Mié", "Jue"]):
        ws.cell(row=3, column=2 + i, value=day)
    ws["A4"] = "Eva"
    ws["B4"] = "LIBRE"
    wb.save(path)
    wb.close()


class TestHorarioExcelImportLayouts:
    def test_split_friday_week_dates_and_shifts(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _workbook_split_friday(str(p))
        drafts = hei.parse_workbook_sheets(str(p), ["10"])
        assert len(drafts) == 1
        d = drafts[0]
        assert not d["errors"], d["errors"]
        assert d["week_dates"]["Vie"] == "27/02/2026"
        assert d["week_dates"]["Jue"] == "05/03/2026"
        assert d["schedule"]["Ana Perez"]["Sáb"] == "OFF"
        assert d["schedule"]["Ana Perez"]["Dom"] == "VAC"
        assert d["schedule"]["Ana Perez"]["Vie"] == "OFF"
        assert any("Sin código" in w for w in d["warnings"])

    def test_linear_with_formato(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _workbook_linear_vie_jue(str(p), with_formato=True)
        drafts = hei.parse_workbook_sheets(str(p), ["11"])
        assert not drafts[0]["errors"]
        d = drafts[0]
        assert d["week_dates"]["Vie"] == "27/02/2026"
        assert d["schedule"]["Bob"]["Vie"] == "OFF"
        assert d["schedule"]["Bob"]["Sáb"] == "PERM"

    def test_linear_string_dates_no_formato(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _workbook_linear_strings_dates(str(p))
        drafts = hei.parse_workbook_sheets(str(p), ["13"])
        assert not drafts[0]["errors"]
        assert drafts[0]["week_dates"]["Mié"] == "04/03/2026"

    def test_missing_sheet(self, tmp_path):
        p = tmp_path / "h.xlsx"
        _workbook_split_friday(str(p))
        drafts = hei.parse_workbook_sheets(str(p), ["no_existe"])
        assert drafts[0]["errors"]


class TestHorarioExcelPreviewEndpoint:
    def test_preview_lists_sheetnames_without_parsing_rows(self, tmp_path):
        from fastapi.testclient import TestClient

        p = tmp_path / "h.xlsx"
        _workbook_split_friday(str(p))
        import main as app_main  # noqa: E402

        client = TestClient(app_main.app)
        with open(p, "rb") as f:
            res = client.post(
                "/api/history/import-horario-excel/preview",
                files={"file": ("h.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"sheets": "[]"},
            )
        assert res.status_code == 200
        data = res.json()
        assert "10" in data["sheetnames"]
        assert data["drafts"] == []

    def test_preview_returns_drafts(self, tmp_path):
        from fastapi.testclient import TestClient

        p = tmp_path / "h.xlsx"
        _workbook_split_friday(str(p))
        import main as app_main  # noqa: E402

        client = TestClient(app_main.app)
        with open(p, "rb") as f:
            res = client.post(
                "/api/history/import-horario-excel/preview",
                files={"file": ("h.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"sheets": json.dumps(["10"])},
            )
        assert res.status_code == 200
        data = res.json()
        assert len(data["drafts"]) == 1
        assert data["drafts"][0]["sheet"] == "10"
        assert data["drafts"][0]["schedule"]
