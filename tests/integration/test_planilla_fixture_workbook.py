"""
Integración: workbook con tres semanas (extras por tipo, feriado, feriado+jornada),
empleados con salario fijo, rellenar_horas_en_excel y comprobación de fórmulas sin #REF.

Artefacto: este módulo fuerza `PYTEST_PLANILLA_ARTIFACT=1` (fixture autouse) y copia
`tests planillas/planilla_casos_manual.xlsx` para inspección local. En CI el archivo
sigue ignorado por git (`.gitignore`).
"""
from __future__ import annotations

import os
import shutil
import sqlite3
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest

_ROOT = Path(__file__).resolve().parents[2]
_PLAN = _ROOT / "planillas"
if str(_PLAN) not in sys.path:
    sys.path.insert(0, str(_PLAN))

import horario_db as hd  # noqa: E402
import planilla_layout as pllay  # noqa: E402
from planilla import crear_hoja_semanal, num_semana_anual  # noqa: E402

DIAS = hd.DIAS_SEMANA_PLANILLA

# Siempre generar Excel en `tests planillas/` al correr estos tests (puede anularse
# quitando el fixture o poniendo PYTEST_PLANILLA_ARTIFACT=0 antes de pytest).
@pytest.fixture(autouse=True)
def _activar_artifact_planilla(monkeypatch):
    monkeypatch.setenv("PYTEST_PLANILLA_ARTIFACT", "1")


def _artifact_save_enabled() -> bool:
    """True si hay que copiar el xlsx a `tests planillas/` (mayúsculas o minúsculas)."""
    v = os.environ.get("PYTEST_PLANILLA_ARTIFACT") or os.environ.get(
        "pytest_planilla_artifact", ""
    )
    return str(v).strip().lower() in ("1", "true", "yes", "on")


def _all_off():
    return {d: "OFF" for d in DIAS}


def _patch_test_db(tmp_path, monkeypatch):
    """SQLite en archivo + tarifas fijas; no toca planilla.db del proyecto."""
    import database as db_mod

    db_path = tmp_path / "planilla_fixture.db"
    conn = sqlite3.connect(str(db_path), isolation_level=None)
    conn.row_factory = sqlite3.Row
    conn.executescript(
        """
        CREATE TABLE empleados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            tipo_pago TEXT,
            salario_fijo REAL,
            periodo_salario_fijo TEXT,
            aplica_seguro INTEGER DEFAULT 1
        );
        CREATE TABLE prestamos (
            id INTEGER PRIMARY KEY,
            empleado_id INTEGER,
            pago_semanal REAL,
            saldo REAL,
            estado TEXT
        );
        INSERT INTO empleados (nombre, tipo_pago, salario_fijo, periodo_salario_fijo, aplica_seguro)
        VALUES
            ('Alice ExtraD', 'tarjeta', NULL, 'mensual', 1),
            ('Bob ExtraM', 'tarjeta', NULL, 'mensual', 1),
            ('Carol ExtraN', 'tarjeta', NULL, 'mensual', 1),
            ('Diana Efe', 'efectivo', NULL, 'mensual', 1),
            ('Eva Fijo Mensual', 'fijo', 433000.0, 'mensual', 1),
            ('Fran Fijo Semanal', 'fijo', 50000.0, 'semanal', 1);
        """
    )
    conn.commit()
    conn.close()

    def get_conn():
        c = sqlite3.connect(str(db_path), isolation_level=None)
        c.row_factory = sqlite3.Row
        c.execute("PRAGMA foreign_keys = ON")
        c.execute("PRAGMA busy_timeout = 30000")
        return c

    def get_tarifas():
        return {
            "tarifa_diurna": 1000.0,
            "tarifa_mixta": 900.0,
            "tarifa_nocturna": 1100.0,
            "seguro": 1000.0,
            "seguro_modo": "porcentual",
            "seguro_valor": 0.1,
            "pagar_horas_extra": 1,
        }

    monkeypatch.setattr(db_mod, "get_conn", get_conn)
    monkeypatch.setattr(db_mod, "get_tarifas", get_tarifas)


def _assert_no_ref_errors_in_sheet(ws, label: str):
    bad = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=20):
        for c in row:
            v = c.value
            if v is None:
                continue
            s = str(v).upper()
            if "#REF!" in s or s == "#REF":
                bad.append((c.coordinate, v))
    assert not bad, f"#REF en {ws.title} ({label}): {bad[:12]}"


def _assert_workbook_semanas_sin_ref(wb: openpyxl.Workbook):
    for name in wb.sheetnames:
        if not name.startswith("Semana "):
            continue
        _assert_no_ref_errors_in_sheet(wb[name], "planilla semanal")
        res = f"Res. Sem. {name.split()[-1]}"
        if res in wb.sheetnames:
            _assert_no_ref_errors_in_sheet(wb[res], "resumen semanal")


@pytest.fixture
def empleados_catalogo():
    return {
        "tarjeta": ["Alice ExtraD", "Bob ExtraM", "Carol ExtraN"],
        "efectivo": ["Diana Efe"],
        "fijo": ["Eva Fijo Mensual", "Fran Fijo Semanal"],
    }


def test_planilla_tres_semanas_rellenar_y_sin_ref(tmp_path, monkeypatch, empleados_catalogo):
    _patch_test_db(tmp_path, monkeypatch)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Semana 1: una persona por tipo de extra (preview alineado con rellenar).
    v1 = date(2026, 1, 2)
    prev1 = {
        "Alice ExtraD": {**_all_off(), "Lun": "J_08-18"},
        "Bob ExtraM": {**_all_off(), "Mar": "D4_13-22"},
        "Carol ExtraN": {**_all_off(), "Mié": "N_22-05"},
        "Diana Efe": _all_off(),
        "Eva Fijo Mensual": _all_off(),
        "Fran Fijo Semanal": _all_off(),
    }
    crear_hoja_semanal(
        wb,
        num_semana_anual(v1),
        v1,
        empleados_catalogo,
        0,
        {},
        [],
        horario_preview=prev1,
    )
    h1 = {
        "Alice ExtraD": {**_all_off(), "Lun": "J_08-18"},
        "Bob ExtraM": {**_all_off(), "Mar": "D4_13-22"},
        "Carol ExtraN": {**_all_off(), "Mié": "N_22-05"},
        "Diana Efe": _all_off(),
        "Eva Fijo Mensual": _all_off(),
        "Fran Fijo Semanal": _all_off(),
    }

    # Semana 2: feriado en lunes (sin jornada → horas estándar en fila ★).
    v2 = date(2026, 1, 9)
    prev2 = {k: _all_off() for k in prev1}
    crear_hoja_semanal(
        wb,
        num_semana_anual(v2),
        v2,
        empleados_catalogo,
        0,
        {},
        [{"date": "2026-01-05", "name": "Feriado test"}],
        horario_preview=prev2,
    )
    h2 = dict(prev2)

    # Semana 3: feriado con jornada que genera extras el mismo día.
    v3 = date(2026, 1, 16)
    prev3 = {
        "Alice ExtraD": {**_all_off(), "Lun": "J_08-18"},
        "Bob ExtraM": _all_off(),
        "Carol ExtraN": _all_off(),
        "Diana Efe": _all_off(),
        "Eva Fijo Mensual": _all_off(),
        "Fran Fijo Semanal": _all_off(),
    }
    crear_hoja_semanal(
        wb,
        num_semana_anual(v3),
        v3,
        empleados_catalogo,
        0,
        {},
        [{"date": "2026-01-19", "name": "Feriado trabajo"}],
        horario_preview=prev3,
    )
    h3 = {
        "Alice ExtraD": {**_all_off(), "Lun": "J_08-18"},
        "Bob ExtraM": _all_off(),
        "Carol ExtraN": _all_off(),
        "Diana Efe": _all_off(),
        "Eva Fijo Mensual": _all_off(),
        "Fran Fijo Semanal": _all_off(),
    }

    out_path = tmp_path / "planilla_casos.xlsx"
    try:
        wb.save(out_path)
        wb.close()

        for vn, hn, hol in (
            ("Semana 1", h1, []),
            ("Semana 2", h2, [{"date": "2026-01-05", "name": "Feriado test"}]),
            ("Semana 3", h3, [{"date": "2026-01-19", "name": "Feriado trabajo"}]),
        ):
            ok, msg, _ = hd.rellenar_horas_en_excel(str(out_path), vn, hn, holidays=hol)
            assert ok, msg

        wb2 = openpyxl.load_workbook(out_path, data_only=False)
        _assert_workbook_semanas_sin_ref(wb2)

        # Resumen semanal: tarjeta + salarios fijos en sección propia.
        for sem in ("1", "2", "3"):
            wres = wb2[f"Res. Sem. {sem}"]
            n_alice = sum(
                1
                for r in range(1, wres.max_row + 1)
                if wres.cell(r, 1).value == "Alice ExtraD"
            )
            assert n_alice >= 1, f"Res. Sem. {sem} sin fila de Alice (resumen vacío o roto)"
            for nombre in ("Eva Fijo Mensual", "Fran Fijo Semanal"):
                n_f = sum(
                    1
                    for r in range(1, wres.max_row + 1)
                    if wres.cell(r, 1).value == nombre
                )
                assert n_f >= 1, f"Res. Sem. {sem} sin fila de salario fijo ({nombre})"

        # Hoja semanal: bloque SALARIO FIJO con ambos nombres.
        for sn in ("Semana 1", "Semana 2", "Semana 3"):
            wsem = wb2[sn]
            hdr_monto = any(
                str(wsem.cell(r, 2).value or "").strip() == "Monto a pagar"
                for r in range(1, wsem.max_row + 1)
            )
            assert hdr_monto, f"{sn}: no se encontró cabecera de sección salario fijo"
            for nombre in ("Eva Fijo Mensual", "Fran Fijo Semanal"):
                assert any(
                    wsem.cell(r, 1).value == nombre for r in range(1, wsem.max_row + 1)
                ), f"{sn}: falta empleado fijo {nombre}"

        # Semana 3: feriado en Lun con jornada → horas ★ feriado + grilla ordinaria/extra.
        ws3 = wb2["Semana 3"]
        hora_alice = None
        for rr in range(5, ws3.max_row + 1):
            if (
                ws3.cell(rr, 1).value == "Alice ExtraD"
                and ws3.cell(rr, 2).value == "Hrs. Diurnas"
            ):
                hora_alice = rr
                break
        assert hora_alice is not None
        sc_a = pllay.scan_jornada_block_rows(ws3, hora_alice)
        fer_r = sc_a.get("fer")
        assert fer_r is not None
        col_lun = 6
        assert ws3.cell(fer_r, col_lun).value == hd.HORAS_FERIADO_SIN_LABOR
        assert float(ws3.cell(hora_alice, col_lun).value) == 8.0
        r_ed = sc_a.get("ed")
        assert r_ed is not None
        assert float(ws3.cell(r_ed, col_lun).value) == 2.0

        wb2.close()

        if _artifact_save_enabled():
            dest_dir = _ROOT / "tests planillas"
            dest_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy(out_path, dest_dir / "planilla_casos_manual.xlsx")
    finally:
        if out_path.exists() and not _artifact_save_enabled():
            out_path.unlink(missing_ok=True)
