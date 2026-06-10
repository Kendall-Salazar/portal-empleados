from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import json
import os
import datetime
from scheduler_engine import ShiftScheduler
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from fastapi.responses import FileResponse
import tempfile
import subprocess
import shutil
import re
import unicodedata

app = FastAPI()

# DATABASE — SQLite backend (shared with planilla system)
import sys
import sqlite3

_backend_dir = os.path.dirname(os.path.abspath(__file__))


def _get_resource_root():
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.abspath(sys._MEIPASS)
    return os.path.abspath(os.path.join(_backend_dir, ".."))


def _get_runtime_root():
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.abspath(os.path.join(_backend_dir, ".."))


def _prefer_runtime_path(*parts):
    runtime_path = os.path.join(_runtime_root, *parts)
    resource_path = os.path.join(_resource_root, *parts)
    if os.path.exists(runtime_path):
        return runtime_path
    return resource_path


_resource_root = _get_resource_root()
_runtime_root = _get_runtime_root()
_resource_planillas_dir = os.path.join(_resource_root, "planillas")
_runtime_planillas_dir = os.path.join(_runtime_root, "planillas")
_planillas_dir = (
    _runtime_planillas_dir
    if os.path.exists(_runtime_planillas_dir) or not os.path.exists(_resource_planillas_dir)
    else _resource_planillas_dir
)
_frontend_dir = _prefer_runtime_path("frontend")
_template_path = _prefer_runtime_path("backend", "formato_template.xlsm")

os.makedirs(_planillas_dir, exist_ok=True)

# Add source planillas dir to path for local runs; frozen builds rely on bundled imports.
if os.path.exists(_resource_planillas_dir):
    sys.path.insert(0, os.path.abspath(_resource_planillas_dir))

import database as plan_db  # Initializes cronos.db with all tables
import planilla as pl_module
import generador_boletas as gb_module
import horario_db
import prestamo_sync

# Import routers
from routes import empleados_router, horarios_router, planillas_router, config_router
from routes.horarios import save_history

# Import shared models and helpers (deduplicated from routes/)
from routes.shared_models import (
    Employee, Config, SolverRequest, HistoryEntry,
    PlanillaPermiso, DescontarPermisosRequest, SyncVacPermRequest,
    GeneratorParamFlags, GeneratorEmployeeUpdate, GeneratorParamsBatchPut,
    GeneratorSyncRrhhRequest, HorarioExcelImportItem, HorarioExcelImportConfirm,
    ValidationRulesRequest, ImageExportRequest,
    FolderCreate, FolderAddEntries,
)
from routes.helpers import (
    _get_conn, _history_sqlite_row_to_log_entry, load_db, save_db,
    _normalize_special_days, _parse_date_like, _parse_timestamp,
    _infer_week_start_from_name, _extract_history_anchor, _history_entry_display_name,
    _prepare_history_for_solver, _build_validation_rules_impl, _HISTORY_LIST_ORDER,
)

DB_FILE_LEGACY = "database.json"  # JSON original, kept for migration reference
EXPORT_DIR = os.path.join(_runtime_root, "export_horarios")
os.makedirs(EXPORT_DIR, exist_ok=True)

# Preview del último /api/solve exitoso (no escribe SQLite). Evita save_db en generar.
_last_generated_preview: Optional[dict] = None

def _sanitize_export_stem(value: str, fallback: str = "horario") -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"[<>:\"/\\\\|?*\x00-\x1F]+", " ", normalized)
    normalized = re.sub(r"\s+", "_", normalized.strip())
    normalized = re.sub(r"_+", "_", normalized).strip("._-")
    return normalized or fallback

def _build_export_filename_parts(filename: str, default_ext: str = ".png", fallback: str = "horario"):
    safe_name = os.path.basename(str(filename or "")).strip()
    stem, ext = os.path.splitext(safe_name)
    ext = (ext or default_ext).lower()
    if not re.fullmatch(r"\.[a-z0-9]+", ext):
        ext = default_ext
    return _sanitize_export_stem(stem or safe_name, fallback), ext

def _get_export_base_name(entry: Optional[dict], fallback: str = "horario") -> str:
    if not isinstance(entry, dict):
        return fallback
    return _sanitize_export_stem(entry.get("name"), fallback)


def _normalize_inventory_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def _matches_inventory_alias(normalized_value: str, alias: str) -> bool:
    if not normalized_value:
        return False
    return (
        normalized_value == alias
        or normalized_value.startswith(f"{alias} ")
        or f" {alias} " in f" {normalized_value} "
    )


def _detect_inventory_header_row(ws, target_fields, max_scan_rows=30, max_scan_cols=30):
    normalized_aliases = {
        field: [_normalize_inventory_header(alias) for alias in aliases]
        for field, aliases in target_fields.items()
    }

    best_row = None
    best_header_map = {}
    best_score = 0

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_header_map = {}
        matched_fields = set()

        for col_idx in range(1, min(ws.max_column, max_scan_cols) + 1):
            normalized_value = _normalize_inventory_header(ws.cell(row=row_idx, column=col_idx).value)
            if not normalized_value:
                continue

            for field, aliases in normalized_aliases.items():
                if field in matched_fields:
                    continue
                if any(_matches_inventory_alias(normalized_value, alias) for alias in aliases):
                    row_header_map[col_idx] = field
                    matched_fields.add(field)
                    break

        row_score = len(matched_fields)
        if row_score > best_score:
            best_score = row_score
            best_row = row_idx
            best_header_map = row_header_map

    return best_row, best_header_map


def _upsert_history_horario_import(
    conn,
    name: str,
    schedule: dict,
    week_dates: dict,
    daily_tasks: Optional[dict],
):
    """Inserta o actualiza por metadata.week_dates['Vie'] (misma cadena DD/MM/YYYY)."""
    daily_tasks = daily_tasks or {}
    vie = week_dates.get("Vie")
    if not vie:
        raise HTTPException(status_code=400, detail="week_dates debe incluir Vie")

    row_id = None
    rows = conn.execute(
        "SELECT id, metadata FROM horarios_generados WHERE IFNULL(deleted, 0) = 0"
    ).fetchall()
    for row in rows:
        meta = {}
        if row["metadata"]:
            try:
                meta = json.loads(row["metadata"])
            except (json.JSONDecodeError, TypeError):
                meta = {}
        wd = meta.get("week_dates") or {}
        if isinstance(wd, dict) and wd.get("Vie") == vie:
            row_id = row["id"]
            break

    horario_json = json.dumps(schedule, ensure_ascii=False)
    tareas_json = json.dumps(daily_tasks, ensure_ascii=False)
    # Usar la fecha real del Viernes como timestamp para orden cronológico correcto
    vie_str = week_dates.get("Vie", "")
    try:
        vie_date = datetime.datetime.strptime(vie_str, "%d/%m/%Y")
        ts = vie_date.replace(hour=12, minute=0, second=0).isoformat()
    except (ValueError, TypeError):
        ts = datetime.datetime.now().isoformat()

    if row_id:
        row = conn.execute(
            "SELECT metadata FROM horarios_generados WHERE id = ? AND IFNULL(deleted, 0) = 0",
            (row_id,),
        ).fetchone()
        existing_meta = {}
        if row and row["metadata"]:
            try:
                existing_meta = json.loads(row["metadata"])
            except (json.JSONDecodeError, TypeError):
                existing_meta = {}
        existing_meta["week_dates"] = week_dates
        metadata_json = json.dumps(existing_meta, ensure_ascii=False)
        conn.execute(
            """
            UPDATE horarios_generados
            SET nombre = ?, horario = ?, tareas = ?, metadata = ?, timestamp = ?
            WHERE id = ?
            """,
            (name, horario_json, tareas_json, metadata_json, ts, row_id),
        )
        return row_id, "updated"

    metadata_json = json.dumps(
        {"week_dates": week_dates, "special_days": {}},
        ensure_ascii=False,
    )
    cursor = conn.execute(
        """
        INSERT INTO horarios_generados (nombre, horario, tareas, metadata, timestamp, deleted, deleted_at)
        VALUES (?, ?, ?, ?, ?, 0, NULL)
        """,
        (name, horario_json, tareas_json, metadata_json, ts),
    )
    return cursor.lastrowid, "inserted"


@app.post("/api/history/import-horario-excel/preview")
async def import_horario_excel_preview(
    file: UploadFile = File(...),
    sheets: str = Form("[]"),
):
    import horario_excel_import as hei

    suffix = os.path.splitext(file.filename or "")[1].lower()
    if suffix not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=400, detail="Solo se admiten .xlsx o .xlsm")

    raw = await file.read()
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo demasiado grande (máx. 25 MB)")

    fd, path = tempfile.mkstemp(suffix=suffix or ".xlsx")
    os.close(fd)
    try:
        with open(path, "wb") as f:
            f.write(raw)
        try:
            want = json.loads(sheets) if (sheets or "").strip() else []
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=f"sheets debe ser JSON array: {exc}") from exc
        if not isinstance(want, list):
            raise HTTPException(status_code=400, detail="sheets debe ser un array JSON")
        if not want:
            sns = hei.list_sheet_names(path)
            return {"sheetnames": sns, "drafts": []}

        drafts = hei.parse_workbook_sheets(path, [str(s) for s in want])
        _, inv_collisions = hei.build_inverse_shift_map()
        return {
            "sheetnames": None,
            "drafts": drafts,
            "inverse_map_warnings": inv_collisions[:30],
        }
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


@app.post("/api/history/import-horario-excel/confirm")
def import_horario_excel_confirm(body: HorarioExcelImportConfirm):
    import horario_excel_import as hei

    days_order = hei.DAYS_ORDER
    results = []
    conn = _get_conn()
    try:
        for item in body.items:
            wd = item.week_dates or {}
            missing = [d for d in days_order if d not in wd]
            if missing:
                raise HTTPException(
                    status_code=400,
                    detail=f"week_dates incompleto para {item.name!r}: faltan {missing}",
                )
            row_id, action = _upsert_history_horario_import(
                conn,
                item.name,
                item.schedule,
                wd,
                item.daily_tasks,
            )
            results.append({"row_id": row_id, "action": action, "name": item.name})
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()

    global _last_generated_preview
    _last_generated_preview = None
    return {"status": "ok", "results": results}


# ═══════════════════════════════════════════════════════════
# FOLDERS — carpetas para agrupar horarios del año
# ═══════════════════════════════════════════════════════════

class FolderCreate(BaseModel):
    name: str

class FolderAddEntries(BaseModel):
    entry_ids: List[int]

@app.get("/api/folders")
def list_folders():
    """Lista todas las carpetas activas con conteo de entradas."""
    conn = _get_conn()
    rows = conn.execute("""
        SELECT f.*, COUNT(fe.id) AS entry_count
        FROM folders f
        LEFT JOIN folder_entries fe ON fe.folder_id = f.id
        WHERE f.deleted = 0
        GROUP BY f.id
        ORDER BY f.name DESC
    """).fetchall()
    conn.close()
    return [{
        "id": r["id"],
        "name": r["name"],
        "created_at": r["created_at"],
        "entry_count": r["entry_count"]
    } for r in rows]

@app.post("/api/folders")
def create_folder(body: FolderCreate):
    """Crea una carpeta nueva."""
    conn = _get_conn()
    now = datetime.datetime.now().isoformat()
    conn.execute(
        "INSERT INTO folders (name, created_at) VALUES (?, ?)",
        (body.name.strip(), now)
    )
    folder_id = conn.lastrowid
    conn.commit()
    conn.close()
    return {"id": folder_id, "name": body.name.strip(), "created_at": now, "entry_count": 0}

@app.delete("/api/folders/{folder_id}")
def delete_folder(folder_id: int, purge: bool = False):
    """Elimina lógico (papelera 7d) o físico si purge=true."""
    conn = _get_conn()
    row = conn.execute("SELECT id FROM folders WHERE id = ? AND deleted = 0", (folder_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "Carpeta no encontrada")
    if purge:
        # Eliminación física: borrar entradas + carpeta
        conn.execute("DELETE FROM folder_entries WHERE folder_id = ?", (folder_id,))
        conn.execute("DELETE FROM folders WHERE id = ?", (folder_id,))
    else:
        now = datetime.datetime.now().isoformat()
        conn.execute("UPDATE folders SET deleted = 1, deleted_at = ? WHERE id = ?", (now, folder_id))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

@app.get("/api/folders/trash")
def list_folder_trash():
    """Lista carpetas en papelera."""
    conn = _get_conn()
    rows = conn.execute("""
        SELECT f.*, COUNT(fe.id) AS entry_count
        FROM folders f
        LEFT JOIN folder_entries fe ON fe.folder_id = f.id
        WHERE f.deleted = 1
        GROUP BY f.id
        ORDER BY f.deleted_at DESC
    """).fetchall()
    conn.close()
    return [{
        "id": r["id"],
        "name": r["name"],
        "created_at": r["created_at"],
        "deleted_at": r["deleted_at"],
        "entry_count": r["entry_count"]
    } for r in rows]

@app.post("/api/folders/{folder_id}/restore")
def restore_folder(folder_id: int):
    """Restaura carpeta de la papelera."""
    conn = _get_conn()
    conn.execute("UPDATE folders SET deleted = 0, deleted_at = NULL WHERE id = ?", (folder_id,))
    conn.commit()
    conn.close()
    return {"status": "restored"}

@app.post("/api/folders/{folder_id}/entries")
def add_folder_entries(folder_id: int, body: FolderAddEntries):
    """Agrega horarios a una carpeta."""
    conn = _get_conn()
    folder = conn.execute("SELECT id FROM folders WHERE id = ? AND deleted = 0", (folder_id,)).fetchone()
    if not folder:
        conn.close()
        raise HTTPException(404, "Carpeta no encontrada")
    added = 0
    for eid in body.entry_ids:
        # Verificar que el horario existe y no está en la carpeta ya
        existing = conn.execute(
            "SELECT id FROM folder_entries WHERE folder_id = ? AND history_entry_id = ?",
            (folder_id, eid)
        ).fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO folder_entries (folder_id, history_entry_id) VALUES (?, ?)",
                (folder_id, eid)
            )
            added += 1
    conn.commit()
    conn.close()
    return {"status": "ok", "added": added}

@app.delete("/api/folders/{folder_id}/entries/{entry_id}")
def remove_folder_entry(folder_id: int, entry_id: int):
    """Quita un horario de la carpeta."""
    conn = _get_conn()
    conn.execute(
        "DELETE FROM folder_entries WHERE folder_id = ? AND history_entry_id = ?",
        (folder_id, entry_id)
    )
    conn.commit()
    conn.close()
    return {"status": "removed"}

@app.get("/api/folders/{folder_id}/entries")
def get_folder_entries(folder_id: int):
    """Lista horarios dentro de una carpeta (orden cronológico ascendente)."""
    conn = _get_conn()
    rows = conn.execute("""
        SELECT h.id, h.nombre, h.horario, h.tareas, h.metadata, h.timestamp
        FROM folder_entries fe
        JOIN horarios_generados h ON h.id = fe.history_entry_id
        WHERE fe.folder_id = ? AND h.deleted = 0
        ORDER BY h.timestamp ASC
    """, (folder_id,)).fetchall()
    conn.close()
    return [{
        "db_id": r["id"],
        "name": r["nombre"],
        "schedule": json.loads(r["horario"]) if r["horario"] else {},
        "daily_tasks": json.loads(r["tareas"]) if r["tareas"] else {},
        "metadata": json.loads(r["metadata"]) if r["metadata"] else {},
        "timestamp": r["timestamp"] or "",
    } for r in rows]

@app.get("/api/folders/{folder_id}/export-excel")
def export_folder_excel(folder_id: int):
    """Exporta todos los horarios de una carpeta como Excel multi-sheet."""
    from openpyxl.styles import Border, Side
    from scheduler_engine import DAYS

    conn = _get_conn()
    folder = conn.execute("SELECT name FROM folders WHERE id = ? AND deleted = 0", (folder_id,)).fetchone()
    if not folder:
        conn.close()
        raise HTTPException(404, "Carpeta no encontrada")

    rows = conn.execute("""
        SELECT h.id, h.nombre, h.horario, h.tareas, h.metadata, h.timestamp
        FROM folder_entries fe
        JOIN horarios_generados h ON h.id = fe.history_entry_id
        WHERE fe.folder_id = ? AND h.deleted = 0
        ORDER BY h.timestamp ASC
    """, (folder_id,)).fetchall()
    conn.close()

    if not rows:
        raise HTTPException(400, "La carpeta está vacía")

    import openpyxl
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for r in rows:
        sheet_name = r["nombre"][:31]  # Excel max sheet name length
        ws = wb.create_sheet(title=sheet_name)
        schedule = json.loads(r["horario"]) if r["horario"] else {}
        tasks = json.loads(r["tareas"]) if r["tareas"] else {}

        employees = sorted(schedule.keys())
        # Header row
        headers = ["Empleado"] + DAYS + ["Horas"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = openpyxl.styles.Font(bold=True, size=10)
            cell.border = thin_border

        for ri, emp in enumerate(employees, 2):
            ws.cell(row=ri, column=1, value=emp).border = thin_border
            total_hours = 0
            for di, day in enumerate(DAYS, 2):
                shift = (schedule.get(emp) or {}).get(day, "")
                cell = ws.cell(row=ri, column=di, value=shift)
                cell.border = thin_border
                # Calculate hours if shift is working
                shift_upper = shift.upper() if shift else ""
                if shift_upper and shift_upper not in ("OFF", "VAC", "PERM", ""):
                    try:
                        from scheduler_engine import get_shift_hours_set
                        hrs = get_shift_hours_set(shift_upper)
                        if hrs:
                            total_hours += len(hrs)
                    except Exception:
                        pass
            ws.cell(row=ri, column=len(DAYS) + 2, value=total_hours).border = thin_border

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    folder_name = folder["name"]
    filename = f"HORARIOS {folder_name}.xlsx"
    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "export_horarios")
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return {"filename": filename, "path": filepath, "sheets": len(rows)}

@app.post("/api/save-history-with-folder-check")
def save_history_with_folder_check(entry: HistoryEntry):
    """Igual que save_history pero auto-asigna a carpeta del año si existe."""
    result = save_history(entry)
    if result.get("status") != "Saved":
        return result

    # Inferir año desde week_dates o timestamp
    year = None
    wd = (entry.week_dates or {})
    for v in wd.values():
        if v and "/" in str(v):
            parts = str(v).split("/")
            if len(parts) == 3:
                try:
                    year = int(parts[2])
                    break
                except ValueError:
                    pass
    if not year and entry.timestamp:
        try:
            year = datetime.datetime.fromisoformat(entry.timestamp).year
        except Exception:
            pass

    if year:
        conn = _get_conn()
        folder = conn.execute(
            "SELECT id FROM folders WHERE name = ? AND deleted = 0", (str(year),)
        ).fetchone()
        if folder:
            # Obtener el id del horario recién guardado
            row = conn.execute(
                "SELECT id FROM horarios_generados WHERE deleted = 0 ORDER BY id DESC LIMIT 1"
            ).fetchone()
            if row:
                conn.execute(
                    "INSERT OR IGNORE INTO folder_entries (folder_id, history_entry_id) VALUES (?, ?)",
                    (folder["id"], row["id"])
                )
            conn.commit()
        conn.close()

    return result


def _soft_delete_history_row(conn, row_id: int) -> bool:
    now = datetime.datetime.now().isoformat()
    cur = conn.execute(
        "UPDATE horarios_generados SET deleted = 1, deleted_at = ? WHERE id = ? AND deleted = 0",
        (now, row_id),
    )
    return cur.rowcount > 0


@app.delete("/api/history/entry/{row_id}")
def delete_history_by_db_id(row_id: int):
    """Mueve a la papelera por id de fila (robusto ante orden o deduplicación en el cliente)."""
    conn = _get_conn()
    if _soft_delete_history_row(conn, row_id):
        conn.commit()
        conn.close()
        return {"status": "Trashed", "message": "Movido a la papelera"}
    conn.close()
    raise HTTPException(status_code=404, detail="Entrada no encontrada")


def _patch_history_row(conn, row_id: int, entry: HistoryEntry) -> bool:
    row = conn.execute(
        "SELECT id, nombre, horario, tareas, metadata FROM horarios_generados WHERE id = ? AND deleted = 0",
        (row_id,),
    ).fetchone()
    if not row:
        return False

    horario_json = json.dumps(entry.schedule, ensure_ascii=False)
    tareas_json = json.dumps(entry.daily_tasks or {}, ensure_ascii=False)

    existing_meta = json.loads(row["metadata"]) if row["metadata"] else {}
    normalized_special_days = _normalize_special_days(entry.special_days)
    if normalized_special_days:
        existing_meta["special_days"] = normalized_special_days
    else:
        existing_meta.pop("special_days", None)
    if entry.week_dates is not None:
        existing_meta["week_dates"] = entry.week_dates
    
    # Merge metadata (holiday_days, etc)
    if entry.metadata:
        for key, value in entry.metadata.items():
            existing_meta[key] = value

    metadata_json = json.dumps(existing_meta, ensure_ascii=False)

    conn.execute(
        """
        UPDATE horarios_generados
        SET horario = ?, tareas = ?, metadata = ?
        WHERE id = ?
        """,
        (horario_json, tareas_json, metadata_json, row_id),
    )
    return True


@app.patch("/api/history/entry/{row_id}")
def update_history_item_by_db_id(row_id: int, entry: HistoryEntry):
    """Actualiza historial por id de fila."""
    conn = _get_conn()
    if _patch_history_row(conn, row_id, entry):
        conn.commit()
        conn.close()
        return {"status": "Updated"}
    conn.close()
    raise HTTPException(status_code=404, detail="Entrada no encontrada")


def _reassign_history_tasks_for_row(conn, row_id: int) -> dict:
    row = conn.execute(
        "SELECT id, nombre, horario, tareas, metadata FROM horarios_generados WHERE id = ? AND deleted = 0",
        (row_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Index out of bounds")

    schedule = json.loads(row["horario"]) if row["horario"] else {}
    if not isinstance(schedule, dict) or not schedule:
        raise HTTPException(status_code=400, detail="El historial no tiene un horario válido")

    employees_rows = conn.execute("SELECT * FROM horario_empleados WHERE activo=1").fetchall()
    
    # Map from database columns (nombre, es_jefe_pista, etc) to ShiftScheduler keys (name, is_jefe_pista, etc)
    employees_data = []
    for e in employees_rows:
        e = dict(e)
        try:
            fixed_shifts = json.loads(e["turnos_fijos"]) if e["turnos_fijos"] else {}
        except:
            fixed_shifts = {}
        employees_data.append({
            "name": e["nombre"],
            "gender": e.get("genero", "M"),
            "can_do_night": bool(e.get("puede_nocturno", 1)),
            "allow_no_rest": bool(e.get("allow_no_rest", 0)),
            "forced_libres": bool(e.get("forced_libres", 0)),
            "forced_quebrado": bool(e.get("forced_quebrado", 0)),
            "is_jefe_pista": bool(e.get("es_jefe_pista", 0)),
            "is_practicante": bool(e.get("es_practicante", 0)),
            "strict_preferences": bool(e.get("strict_preferences", 0)),
            "fixed_shifts": fixed_shifts
        })

    employee_names = {emp["name"] for emp in employees_data}
    for missing_name in sorted(name for name in schedule.keys() if name not in employee_names):
        employees_data.append({
            "name": missing_name, 
            "gender": "M", 
            "can_do_night": 1,
            "is_jefe_pista": 0,
            "fixed_shifts": {}
        })

    config_row = conn.execute("SELECT * FROM horario_config WHERE id=1").fetchone()
    config_data = dict(config_row) if config_row else {}
    # cleaning_tasks y jefe_config vienen de la DB (columna TEXT → string), parseamos a dict
    if "cleaning_tasks" in config_data and isinstance(config_data["cleaning_tasks"], str):
        try:
            config_data["cleaning_tasks"] = json.loads(config_data["cleaning_tasks"]) if config_data["cleaning_tasks"] else {}
        except json.JSONDecodeError:
            config_data["cleaning_tasks"] = {}
    if "jefe_config" in config_data and isinstance(config_data["jefe_config"], str):
        try:
            config_data["jefe_config"] = json.loads(config_data["jefe_config"]) if config_data["jefe_config"] else {}
        except json.JSONDecodeError:
            config_data["jefe_config"] = {}
    config_data["use_refuerzo"] = "Refuerzo" in schedule
    existing_meta = json.loads(row["metadata"]) if row["metadata"] else {}
    special_days = _normalize_special_days(existing_meta.get("special_days", {}))
    config_data["special_days"] = special_days

    scheduler = ShiftScheduler(employees_data, config_data, history_data=[])
    try:
        daily_tasks = scheduler.assign_tasks(schedule)
    except Exception as e:
        print(f"Error in assign_tasks: {e}")
        # Fallback to empty tasks instead of crashing
        daily_tasks = {}


    existing_meta["daily_tasks"] = daily_tasks
    if special_days:
        existing_meta["special_days"] = special_days
    else:
        existing_meta.pop("special_days", None)

    conn.execute(
        "UPDATE horarios_generados SET tareas = ?, metadata = ? WHERE id = ?",
        (json.dumps(daily_tasks, ensure_ascii=False), json.dumps(existing_meta, ensure_ascii=False), row_id),
    )

    return {
        "status": "Updated",
        "daily_tasks": daily_tasks,
        "special_days": special_days,
    }


@app.patch("/api/history/entry/{row_id}/task")
def update_history_task(row_id: int, employee_name: str, day: str, task: str = None):
    """Update or remove a specific cleaning task in a history entry."""
    conn = plan_db.get_conn()
    row = conn.execute(
        "SELECT tareas, metadata FROM horarios_generados WHERE id = ? AND deleted = 0",
        (row_id,),
    ).fetchone()
    
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Historial no encontrado")

    try:
        daily_tasks = json.loads(row["tareas"]) if row["tareas"] else {}
        metadata = json.loads(row["metadata"]) if row["metadata"] else {}
    except:
        daily_tasks = {}
        metadata = {}

    if not isinstance(daily_tasks, dict):
        daily_tasks = {}

    if employee_name not in daily_tasks:
        daily_tasks[employee_name] = {}
    
    if task and task.strip():
        daily_tasks[employee_name][day] = task.strip()
    else:
        # Remove task
        if day in daily_tasks[employee_name]:
            daily_tasks[employee_name][day] = None

    metadata["daily_tasks"] = daily_tasks

    conn.execute(
        "UPDATE horarios_generados SET tareas = ?, metadata = ? WHERE id = ?",
        (json.dumps(daily_tasks, ensure_ascii=False), json.dumps(metadata, ensure_ascii=False), row_id),
    )
    conn.commit()
    conn.close()
    return {"status": "Updated", "daily_tasks": daily_tasks}


@app.post("/api/history/entry/{row_id}/reassign_tasks")
def reassign_history_tasks_by_db_id(row_id: int):
    """Recalcular tareas de limpieza por id de fila."""
    conn = _get_conn()
    try:
        out = _reassign_history_tasks_for_row(conn, row_id)
        conn.commit()
        return out
    finally:
        conn.close()


def format_shift_code(code: str) -> str:
    """
    Converts internal shift code (e.g., 'T12_14-22') to readable 12h format 
    (e.g., '02:00 PM - 10:00 PM').
    """
    normalized = horario_db.normalize_manual_shift_code(code) or code

    if not normalized or normalized == "OFF": return "LIBRE"
    if normalized == "VAC": return "VACACIONES"
    if normalized == "PERM": return "PERMISO"
    
    # Check for split shift first (e.g. Q1_05-11+17-20)
    if "+" in normalized:
        parts = normalized.split("_")
        if len(parts) > 1:
            times = parts[1].split("+") # ['05-11', '17-20']
            readable_times = []
            for t in times:
                readable_times.append(_format_time_range(t))
            return " / ".join(readable_times)
    
    # Standard Shift (Code_Start-End)
    parts = normalized.split("_")
    if len(parts) > 1:
        return _format_time_range(parts[1])
        
    return normalized

def _format_time_range(time_range: str) -> str:
    # time_range ex: "14-22" or "06-16"
    try:
        start_s, end_s = time_range.split("-")
        start_h = int(start_s)
        end_h = int(end_s)
        
        # Handle > 24 (next day)
        start_dt = datetime.time(start_h % 24, 0)
        end_dt = datetime.time(end_h % 24, 0)
        
        start_str = start_dt.strftime("%I:%M %p")
        end_str = end_dt.strftime("%I:%M %p")
        
        return f"{start_str} - {end_str}"
    except (ValueError, TypeError, AttributeError):
        return time_range


EXCEL_EMPLOYEE_PALETTE = [
    "4D93D9",
    "FF0000",
    "61CBF3",
    "663300",
    "D86DCD",
    "153D64",
    "8ED973",
    "D9EAD3",
    "BFBFBF",
    "F1A983",
    "F9E79F",
    "D6E4F0",
]
EXCEL_EMPLOYEE_COLOR_MAP = {
    "Angel": "4D93D9",
    "Eligio": "FF0000",
    "Ileana": "61CBF3",
    "Jeison": "663300",
    "Jensy": "D86DCD",
    "Keilor": "153D64",
    "Maikel": "8ED973",
    "Natanael": "FF0000",
    "Alejandro": "D9EAD3",
    "Randall": "BFBFBF",
    "Steven": "F1A983",
    "Tomas": "F9E79F",
    "Refuerzo": "D6E4F0",
}
EXCEL_EMPLOYEE_FONT_COLOR_MAP = {
    "Eligio": "FFFFFF",
    "Natanael": "FFFFFF",
}
EXCEL_LIBRE_FILL = "FFFF00"


def _excel_font_color_for_fill(hex_color: str) -> str:
    color = (hex_color or "").strip().lstrip("#")
    if len(color) == 8:
        color = color[2:]
    if len(color) != 6:
        return "000000"

    try:
        red = int(color[0:2], 16)
        green = int(color[2:4], 16)
        blue = int(color[4:6], 16)
    except ValueError:
        return "000000"

    brightness = (red * 299 + green * 587 + blue * 114) / 1000
    return "FFFFFF" if brightness < 145 else "000000"


def _normalize_excel_task_text(task_text: str) -> str:
    if task_text is None:
        return ""

    text = str(task_text).strip()
    replacements = {
        "â†‘AM": "↑",
        "â†“PM": "↓",
        "↑AM": "↑",
        "↓PM": "↓",
        "↑ AM": "↑",
        "↓ PM": "↓",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return " ".join(text.split())


def _build_excel_task_text(task_text: str, shift_code: str) -> str:
    text = _normalize_excel_task_text(task_text)
    if not text:
        return ""

    if "↑" not in text and "↓" not in text:
        hours = sorted(horario_db.get_shift_hours_set(shift_code))
        if hours:
            text = f"{text} {'↑' if hours[0] < 12 else '↓'}"

    return text


def _task_style_key(task_text: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(task_text or ""))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower()
    if "banos" in normalized:
        return "banos"
    if "tanques" in normalized:
        return "tanques"
    if "oficina" in normalized:
        return "oficina"
    if "calibracion" in normalized:
        return "calibracion"
    if "canos" in normalized:
        return "canos"
    return ""


@app.get("/api/export_excel")
def export_excel(history_index: Optional[int] = None, history_db_id: Optional[int] = None):
    db = load_db()

    target_schedule = {}
    target_tasks = {}
    selected_entry = None

    if history_db_id is not None:
        # Lookup directo por db_id — robusto ante cambios de orden en el historial
        conn = _get_conn()
        row = conn.execute(
            "SELECT * FROM horarios_generados WHERE id = ? AND IFNULL(deleted, 0) = 0",
            (history_db_id,)
        ).fetchone()
        conn.close()
        if row:
            selected_entry = _history_sqlite_row_to_log_entry(dict(row))
            if selected_entry:
                target_schedule = selected_entry.get("schedule", {})
                target_tasks = selected_entry.get("daily_tasks", {})
    elif history_index is not None:
        history_list = db.get("history_log", [])
        if 0 <= history_index < len(history_list):
            selected_entry = history_list[history_index]
            target_schedule = selected_entry.get("schedule", {})
            target_tasks = selected_entry.get("daily_tasks", {})
    else:
        global _last_generated_preview
        target_schedule = {}
        target_tasks = {}
        preview_used = False
        if _last_generated_preview and _last_generated_preview.get("schedule"):
            target_schedule = _last_generated_preview.get("schedule") or {}
            target_tasks = _last_generated_preview.get("daily_tasks") or {}
            preview_used = True
        if not target_schedule:
            last_result = db.get("last_result", {})
            target_schedule = last_result.get("schedule", {})
            target_tasks = last_result.get("daily_tasks", {})
        history_list = db.get("history_log", [])
        if preview_used:
            meta = (_last_generated_preview or {}).get("metadata") or {}
            wd = meta.get("week_dates")
            if wd or meta:
                selected_entry = {"week_dates": wd, "metadata": meta}
        elif history_list:
            selected_entry = history_list[-1]
    
    if not target_schedule:
        raise HTTPException(status_code=404, detail="No hay horario generado para exportar")

    from openpyxl.styles import Border, Side
    from scheduler_engine import DAYS

    template_path = _template_path
    use_vba = os.path.exists(template_path)
    
    if use_vba:
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active
        if ws.title != "Horario":
            ws.title = "Horario"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horario"
    
    palette = EXCEL_EMPLOYEE_PALETTE
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    week_dates = None
    if selected_entry:
        week_dates = selected_entry.get("week_dates")
        if not week_dates:
            meta = selected_entry.get("metadata", {})
            if isinstance(meta, str):
                import json as _json
                try:
                    meta = _json.loads(meta)
                except _json.JSONDecodeError:
                    if history_index is not None:
                        print(f"export_excel: metadata JSON invalido en history_index={history_index}")
                    meta = {}
            week_dates = meta.get("week_dates")
    current_row = 1

    # --- Fila de título: nombre del horario (A1:H1 unificada) ---
    schedule_title = (selected_entry or {}).get("name", "") if selected_entry else ""
    if schedule_title:
        title_cell = ws.cell(row=current_row, column=1, value=schedule_title)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        title_cell.font = Font(bold=True, color="2F5496", size=24)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 36
        current_row += 1

    # --- Fila de fechas ---
    if week_dates:
        date_font = Font(bold=True, color="2F5496", size=10)
        for day_idx, day in enumerate(DAYS, start=2):
            cell = ws.cell(row=current_row, column=day_idx, value=week_dates.get(day, ""))
            cell.number_format = "d/m/yyyy"
            cell.font = date_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
    
    header_row = current_row
    headers = ["Colaborador"] + DAYS + ["Horas"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    emp_names = list(target_schedule.keys())
    emp_colors = {}
    fallback_index = 0
    for name in emp_names:
        if name in EXCEL_EMPLOYEE_COLOR_MAP:
            emp_colors[name] = EXCEL_EMPLOYEE_COLOR_MAP[name]
        else:
            emp_colors[name] = palette[fallback_index % len(palette)]
            fallback_index += 1
    first_emp_row = header_row + 1
    
    for idx, name in enumerate(emp_names):
        shifts = target_schedule[name]
        row_data = [name]
        total_hours = 0
        
        for d in DAYS:
            s_code = shifts.get(d, "OFF")
            readable_shift = format_shift_code(s_code)
            row_data.append(readable_shift)
            
            # Calculate hours
            hours = len(horario_db.get_shift_hours_set(s_code))
            total_hours += hours
            
        row_data.append(total_hours)
        current_data_row = first_emp_row + idx
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=current_data_row, column=col_idx, value=value)

        color = emp_colors[name]
        row_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row_font_color = EXCEL_EMPLOYEE_FONT_COLOR_MAP.get(name, _excel_font_color_for_fill(color))
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_data_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            
            if col_idx == 1:
                cell.fill = row_fill
                cell.font = Font(bold=True, size=11, color=row_font_color)
                cell.alignment = Alignment(vertical="center", horizontal="left")
            else:
                cell.fill = row_fill
                cell.font = Font(
                    bold=(col_idx == len(row_data)),
                    size=11 if col_idx == len(row_data) else 10,
                    color=row_font_color,
                )
                
                s_code = shifts.get(DAYS[col_idx - 2], "OFF") if col_idx <= 8 else None
                if s_code == "OFF":
                    cell.fill = PatternFill(start_color=EXCEL_LIBRE_FILL, end_color=EXCEL_LIBRE_FILL, fill_type="solid")
                    cell.font = Font(color="999999", italic=True, size=11 if col_idx == len(row_data) else 10)
                elif s_code == "VAC":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif s_code == "PERM":
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    cell.font = Font(color="9A3412", bold=True)
        
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, 9):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(9)].width = 8
    
    formato_col = 11  # Column K
    
    fmt_header = ws.cell(row=header_row, column=formato_col, value="FORMATO")
    fmt_header.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    fmt_header.font = Font(bold=True, color="FFFFFF", size=11)
    fmt_header.alignment = Alignment(horizontal="center", vertical="center")
    fmt_header.border = thin_border
    
    for idx, name in enumerate(emp_names):
        fmt_row = first_emp_row + idx
        color = emp_colors[name]
        fmt_cell = ws.cell(row=fmt_row, column=formato_col, value=name)
        fmt_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fmt_cell.font = Font(bold=True, size=11, color=_excel_font_color_for_fill(color))
        fmt_cell.alignment = Alignment(vertical="center", horizontal="left")
        fmt_cell.border = thin_border
    
    libre_row = first_emp_row + len(emp_names)
    libre_cell = ws.cell(row=libre_row, column=formato_col, value="LIBRE")
    libre_cell.fill = PatternFill(start_color=EXCEL_LIBRE_FILL, end_color=EXCEL_LIBRE_FILL, fill_type="solid")
    libre_cell.font = Font(color="999999", italic=True, size=11)
    libre_cell.alignment = Alignment(vertical="center", horizontal="left")
    libre_cell.border = thin_border
    
    ws.column_dimensions[openpyxl.utils.get_column_letter(formato_col)].width = 18

    separator_row = ws.max_row + 2
    
    ws.cell(row=separator_row, column=1, value="OBLIGACIONES / LIMPIEZA")
    cleaning_title_cell = ws.cell(row=separator_row, column=1)
    cleaning_title_cell.font = Font(bold=True, size=24, color="2F5496")
    cleaning_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=separator_row, start_column=1, end_row=separator_row, end_column=8)
    ws.row_dimensions[separator_row].height = 36
    
    task_header_row = separator_row + 1
    task_headers = ["Colaborador"] + DAYS
    for col_idx, header in enumerate(task_headers, 1):
        cell = ws.cell(row=task_header_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    for idx, name in enumerate(emp_names):
        emp_tasks = target_tasks.get(name, {})
        task_row_num = task_header_row + 1 + idx
        color = emp_colors[name]
        name_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        name_cell = ws.cell(row=task_row_num, column=1, value=name)
        name_cell.fill = name_fill
        name_cell.font = Font(bold=True, size=10, color=_excel_font_color_for_fill(color))
        name_cell.alignment = Alignment(vertical="center", horizontal="left")
        name_cell.border = thin_border
        
        for d_idx, d in enumerate(DAYS):
            task = emp_tasks.get(d)
            shift_code = target_schedule.get(name, {}).get(d, "OFF")
            col = d_idx + 2
            
            if task:
                task_text = _build_excel_task_text(task, shift_code)
                cell = ws.cell(row=task_row_num, column=col, value=task_text)
                style_key = _task_style_key(task_text)
                
                if style_key == "banos":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.font = Font(color="B45309", bold=True, size=10)
                elif style_key == "tanques":
                    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                    cell.font = Font(color="1D4ED8", bold=True, size=10)
                elif style_key == "oficina":
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    cell.font = Font(color="BE185D", bold=True, size=10)
                elif style_key == "calibracion":
                    cell.fill = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")
                    cell.font = Font(color="6B21A8", bold=True, size=10)
                elif style_key == "canos":
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    cell.font = Font(color="065F46", bold=True, size=10)
                else:
                    cell.font = Font(size=10)
            else:
                cell = ws.cell(row=task_row_num, column=col, value="—")
                cell.font = Font(color="CCCCCC", size=10)
            
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = thin_border

    # Save to temp
    suffix = ".xlsm" if use_vba else ".xlsx"
    export_base_name = "horario"
    if selected_entry:
        fallback = f"historial_{history_index + 1}" if history_index is not None else "horario"
        export_base_name = _get_export_base_name(selected_entry, fallback)
    filename = f"{export_base_name}{suffix}"
    
    # Save a local backup to EXPORT_DIR
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    local_filename = f"{export_base_name}_{timestamp}{suffix}"
    local_path = os.path.join(EXPORT_DIR, local_filename)
    wb.save(local_path)
    
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    wb.save(tmp.name)
    tmp.close()
    wb.close()
    
    return FileResponse(
        tmp.name,
        filename=filename,
        headers={"X-Export-Local-Path": local_path},
    )


@app.post("/api/open_export_folder")
def open_export_folder():
    """Abre la carpeta de exportaciones en el explorador de archivos."""
    if not os.path.isdir(EXPORT_DIR):
        os.makedirs(EXPORT_DIR, exist_ok=True)
    if os.name == "nt":
        os.startfile(EXPORT_DIR)
    return {"status": "ok", "path": EXPORT_DIR}


@app.post("/api/export_image")
def export_image(req: ImageExportRequest):
    import base64
    try:
        # data:image/png;base64,...
        header, encoded = req.image_data.split(",", 1)
        data = base64.b64decode(encoded)
        
        stem, ext = _build_export_filename_parts(req.filename, default_ext=".png", fallback="horario")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{stem}_{timestamp}{ext}"
        file_path = os.path.join(EXPORT_DIR, filename)
        
        with open(file_path, "wb") as f:
            f.write(data)
            
        return {"status": "success", "file": filename, "path": file_path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==============================================================================
# PLANILLAS API ENDPOINTS (Shared with new unified Web UI)
# ==============================================================================

class PlanillaEmpleado(BaseModel):
    nombre: str
    tipo_pago: str
    salario_fijo: Optional[float] = None
    # semanal | quincenal | mensual — el monto salario_fijo es el pago de ese período (solo tipo fijo)
    periodo_salario_fijo: Optional[str] = "mensual"
    cedula: Optional[str] = None
    correo: Optional[str] = None
    telefono: Optional[str] = None
    fecha_inicio: Optional[str] = None
    aplica_seguro: Optional[int] = 1
    genero: Optional[str] = 'M'
    puede_nocturno: Optional[int] = 1
    forced_libres: Optional[int] = 0
    forced_quebrado: Optional[int] = 0
    allow_no_rest: Optional[int] = 0
    es_jefe_pista: Optional[int] = 0
    es_practicante: Optional[int] = 0
    strict_preferences: Optional[int] = 0
    activo: Optional[int] = 1
    incluir_en_horario: Optional[int] = 1
    turnos_fijos: Optional[str] = "{}"
    pref_plantilla_id: Optional[int] = None


class PlanillaEmpleadoUpdate(BaseModel):
    """PUT parcial: solo los campos enviados se persisten (no pisa horario si se omiten)."""

    model_config = ConfigDict(extra="forbid")

    nombre: str
    tipo_pago: str
    salario_fijo: Optional[float] = None
    periodo_salario_fijo: Optional[str] = None
    cedula: Optional[str] = None
    correo: Optional[str] = None
    telefono: Optional[str] = None
    fecha_inicio: Optional[str] = None
    aplica_seguro: Optional[int] = None
    genero: Optional[str] = None
    puede_nocturno: Optional[int] = None
    activo: Optional[int] = None
    incluir_en_horario: Optional[int] = None
    forced_libres: Optional[int] = None
    forced_quebrado: Optional[int] = None
    allow_no_rest: Optional[int] = None
    es_jefe_pista: Optional[int] = None
    es_practicante: Optional[int] = None
    strict_preferences: Optional[int] = None
    turnos_fijos: Optional[str] = None
    pref_plantilla_id: Optional[int] = None


class PlanillaPrefPlantillaCreate(BaseModel):
    nombre: str
    descripcion: str = ""
    activa: int = 1
    turnos_fijos: str = "{}"
    strict_preferences: int = 0
    allow_no_rest: int = 0
    forced_libres: int = 0
    forced_quebrado: int = 0

class PlanillaPrefPlantillaUpdate(BaseModel):
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    activa: Optional[int] = None
    turnos_fijos: Optional[str] = None
    strict_preferences: Optional[int] = None
    allow_no_rest: Optional[int] = None
    forced_libres: Optional[int] = None
    forced_quebrado: Optional[int] = None

class PlanillaVacacion(BaseModel):
    empleado_id: int
    fecha_inicio: str
    fecha_fin: str
    dias: float
    fecha_reingreso: Optional[str] = None
    notas: Optional[str] = None
    solo_pago: bool = False
    # Categorización del registro: 'periodo' (default), 'ajuste_historico'
    # (días gozados antes de usar el sistema) o 'descuento_permiso'.
    tipo: Optional[str] = None
    # Año al que pertenece el período (default = año de fecha_inicio).
    anio_periodo: Optional[int] = None

class PlanillaTarifas(BaseModel):
    tarifa_diurna: float
    tarifa_nocturna: float
    tarifa_mixta: float
    seguro_modo: str = "porcentual"  # porcentual | fijo
    seguro_valor: float  # tasa decimal (ej. 0.1067) o monto fijo en colones por semana
    pagar_horas_extra: bool = True  # si False, el bruto de la planilla no incluye pago por filas de extras

class PlanillaMes(BaseModel):
    anio: int
    mes: int

class PlanillaSemana(BaseModel):
    mes_id: int
    viernes: str

class PlanillaBoletasRequest(BaseModel):
    semana_nombre: str
    tipo_bono: str = "otros"
    valor_sticker: float = 150.0

class PlanillaImportarHorario(BaseModel):
    horario_id: int
    semana_nombre: str
    sync_empleados: bool = True

@app.get("/api/planillas/empleados")
def get_planillas_empleados(solo_activos: bool = True):
    return plan_db.get_empleados(solo_activos)

@app.post("/api/planillas/empleados")
def add_planilla_empleado(emp: PlanillaEmpleado):
    ok, msg = plan_db.add_empleado(
        emp.nombre, emp.tipo_pago, emp.salario_fijo,
        cedula=emp.cedula, correo=emp.correo,
        telefono=emp.telefono, fecha_inicio=emp.fecha_inicio,
        aplica_seguro=emp.aplica_seguro, genero=emp.genero,
        puede_nocturno=emp.puede_nocturno,
        forced_libres=emp.forced_libres, forced_quebrado=emp.forced_quebrado,
        allow_no_rest=emp.allow_no_rest, es_jefe_pista=emp.es_jefe_pista,
        es_practicante=emp.es_practicante,
        strict_preferences=emp.strict_preferences, turnos_fijos=emp.turnos_fijos,
        pref_plantilla_id=emp.pref_plantilla_id,
        incluir_en_horario=emp.incluir_en_horario,
        periodo_salario_fijo=emp.periodo_salario_fijo,
    )
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"status": "success", "message": msg}

@app.put("/api/planillas/empleados/{emp_id}")
def update_planilla_empleado(emp_id: int, emp: PlanillaEmpleadoUpdate):
    data = emp.model_dump(exclude_unset=True)
    activo_val = data.pop("activo", None) if "activo" in data else None
    pref_extra: Dict = {}
    if "pref_plantilla_id" in data:
        pref_extra["pref_plantilla_id"] = data.pop("pref_plantilla_id")
    plan_db.update_empleado(emp_id, **data, **pref_extra)

    if activo_val is not None:
        exist = plan_db.get_conn().execute("SELECT activo FROM empleados WHERE id=?", (emp_id,)).fetchone()
        if exist:
            current_state = exist["activo"]
            new_state = 1 if activo_val else 0
            if current_state == 1 and new_state == 0:
                plan_db.remove_empleado(emp_id)
            elif current_state == 0 and new_state == 1:
                plan_db.reactivar_empleado(emp_id)

    return {"status": "success"}

@app.get("/api/planillas/pref-plantillas")
def list_pref_plantillas(solo_activas: bool = False):
    return plan_db.list_pref_plantillas(solo_activas=solo_activas)

@app.post("/api/planillas/pref-plantillas")
def create_pref_plantilla(body: PlanillaPrefPlantillaCreate):
    try:
        new_id = plan_db.create_pref_plantilla(
            body.nombre,
            descripcion=body.descripcion,
            activa=body.activa,
            turnos_fijos=body.turnos_fijos,
            strict_preferences=body.strict_preferences,
            allow_no_rest=body.allow_no_rest,
            forced_libres=body.forced_libres,
            forced_quebrado=body.forced_quebrado,
        )
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Ya existe una plantilla con ese nombre")
    return {"status": "success", "id": new_id}

@app.get("/api/planillas/pref-plantillas/{plantilla_id}")
def get_pref_plantilla(plantilla_id: int):
    row = plan_db.get_pref_plantilla(plantilla_id)
    if not row:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    return row

@app.put("/api/planillas/pref-plantillas/{plantilla_id}")
def update_pref_plantilla(plantilla_id: int, body: PlanillaPrefPlantillaUpdate):
    if not plan_db.get_pref_plantilla(plantilla_id):
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    patch = body.model_dump(exclude_unset=True)
    if not patch:
        return {"status": "success"}
    try:
        plan_db.update_pref_plantilla(plantilla_id, **patch)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Ya existe una plantilla con ese nombre")
    return {"status": "success"}

@app.delete("/api/planillas/pref-plantillas/{plantilla_id}")
def delete_pref_plantilla(plantilla_id: int):
    if not plan_db.get_pref_plantilla(plantilla_id):
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    plan_db.delete_pref_plantilla(plantilla_id)
    return {"status": "success"}


# ------------------------------------------------------------------------------
# PANEL PARÁMETROS DEL GENERADOR (matriz colaboradores + batch JSON)
# ------------------------------------------------------------------------------
@app.get("/api/generator/employee-params")
def get_generator_employee_params(week_start: Optional[str] = None):
    return plan_db.get_generator_employee_params(week_start)


@app.put("/api/generator/employee-params")
def put_generator_employee_params(body: GeneratorParamsBatchPut):
    payload = []
    for u in body.updates:
        d = {"employee_id": u.employee_id}
        if u.flags is not None:
            d["flags"] = {k: v for k, v in u.flags.model_dump().items() if v is not None}
        if u.shift_preferences is not None:
            d["shift_preferences"] = dict(u.shift_preferences)
        if "pref_plantilla_id" in u.model_fields_set:
            d["pref_plantilla_id"] = u.pref_plantilla_id
        payload.append(d)
    return plan_db.apply_generator_employee_params_batch(payload)


@app.post("/api/generator/sync-rrhh-to-shifts")
def post_generator_sync_rrhh(body: GeneratorSyncRrhhRequest):
    return plan_db.sync_all_rrhh_to_fixed_shifts_for_week(body.week_start)


@app.delete("/api/planillas/empleados/{emp_id}")
def delete_planilla_empleado(emp_id: int):
    plan_db.delete_empleado(emp_id)
    return {"status": "success"}

@app.get("/api/planillas/vacaciones/{emp_id}")
def get_planillas_vacaciones(emp_id: int):
    vacs = plan_db.get_vacaciones(emp_id)
    emp_data = next((e for e in plan_db.get_empleados(solo_activos=False) if e["id"] == emp_id), None)
    if emp_data:
        fi = emp_data.get("fecha_inicio")
        acumulados = plan_db.calcular_dias_vacaciones(fi) if fi else 0
        desglose = plan_db.desglose_vacaciones_empleado(emp_id)
        # 'tomados' (compatible) = todo lo que descuenta saldo
        tomados = desglose["descuento_saldo_total"]
        return {
            "registros": vacs,
            "acumulados": acumulados,
            "tomados": tomados,
            "disponibles": max(0, acumulados - tomados),
            "desglose": desglose,
        }
    return {"registros": vacs}

@app.post("/api/planillas/vacaciones")
def add_planilla_vacacion(vac: PlanillaVacacion):
    plan_db.add_vacacion(
        vac.empleado_id, vac.fecha_inicio, vac.fecha_fin, vac.dias,
        fecha_reingreso=vac.fecha_reingreso, notas=vac.notas, solo_pago=vac.solo_pago,
        tipo=getattr(vac, "tipo", None), anio_periodo=getattr(vac, "anio_periodo", None),
    )
    return {"status": "success"}

@app.delete("/api/planillas/vacaciones/{vac_id}")
def delete_planilla_vacacion(vac_id: int):
    plan_db.delete_vacacion(vac_id)
    return {"status": "success"}

@app.put("/api/planillas/vacaciones/{vac_id}")
def update_planilla_vacacion(vac_id: int, vac: PlanillaVacacion):
    plan_db.update_vacacion(
        vac_id, vac.fecha_inicio, vac.fecha_fin, vac.dias,
        fecha_reingreso=vac.fecha_reingreso, notas=vac.notas, solo_pago=vac.solo_pago,
        tipo=getattr(vac, "tipo", None), anio_periodo=getattr(vac, "anio_periodo", None),
    )
    return {"status": "success"}

# ------------------------------------------------------------------------------
# PERMISOS API
# ------------------------------------------------------------------------------
@app.get("/api/planillas/permisos/{emp_id}")
def get_planillas_permisos(emp_id: int, anio: Optional[int] = None):
    permisos = plan_db.get_permisos_empleado(emp_id, anio=anio)
    if anio:
        conteo = plan_db.get_conteo_permisos_anio(emp_id, anio)
    else:
        conteo = plan_db.get_conteo_permisos_anio(emp_id, datetime.datetime.now().year)
    return {"permisos": permisos, "conteo": conteo}

@app.post("/api/planillas/permisos")
def add_planilla_permiso(perm: PlanillaPermiso):
    permiso_id = plan_db.add_permiso(
        perm.empleado_id, perm.fecha, motivo=perm.motivo, notas=perm.notas,
        fecha_fin=perm.fecha_fin, horas=perm.horas or 0,
    )
    return {"status": "success", "id": permiso_id}

@app.put("/api/planillas/permisos/{permiso_id}")
def update_planilla_permiso(permiso_id: int, perm: PlanillaPermiso):
    plan_db.update_permiso(
        permiso_id, perm.fecha, motivo=perm.motivo, notas=perm.notas,
        fecha_fin=perm.fecha_fin, horas=perm.horas or 0,
    )
    return {"status": "success"}

@app.delete("/api/planillas/permisos/{permiso_id}")
def delete_planilla_permiso(permiso_id: int, restaurar: bool = True):
    plan_db.delete_permiso(permiso_id, restaurar_vacaciones=restaurar)
    return {"status": "success"}

@app.post("/api/planillas/permisos/descontar-vacaciones")
def descontar_permisos_vacaciones(req: DescontarPermisosRequest):
    ok, msg = plan_db.descontar_permisos_de_vacaciones(
        req.empleado_id, req.cantidad, req.anio
    )
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"status": "success", "message": msg}

# ------------------------------------------------------------------------------
# SYNC VACACIONES/PERMISOS → HORARIO (Fixed Shifts)
# ------------------------------------------------------------------------------
@app.post("/api/sync_vac_fixed_shifts")
def sync_vac_fixed_shifts(req: SyncVacPermRequest):
    """Sincroniza vacaciones y permisos activos con los turnos fijos de todos los empleados."""
    emps = plan_db.get_empleados(solo_activos=True)
    updated = []
    for emp in emps:
        shifts = plan_db.sync_vac_perm_to_fixed_shifts(
            emp["nombre"], req.fecha_inicio, req.fecha_fin
        )
        updated.append({"nombre": emp["nombre"], "turnos_fijos": shifts})
    return {"status": "success", "updated": updated}

# ------------------------------------------------------------------------------
# PRÉSTAMOS API
# ------------------------------------------------------------------------------
class PlanillaPrestamo(BaseModel):
    empleado_id: int
    monto_total: float
    pago_semanal: float
    notas: Optional[str] = None

class PrestamoAbono(BaseModel):
    monto: float
    tipo: str = "planilla"
    semana_planilla: Optional[str] = None
    notas: Optional[str] = None
    fecha: Optional[str] = None


def _find_mes_for_sync(mes_id: Optional[int] = None):
    if mes_id is None:
        return plan_db.get_mes_activo()

    meses = plan_db.get_todos_meses()
    return next((m for m in meses if m["id"] == mes_id), None)


def _sync_prestamo_rebajos_mes(mes_id: Optional[int] = None, mes_data: Optional[dict] = None):
    mes = mes_data or _find_mes_for_sync(mes_id)
    if not mes:
        return {
            "status": "noop",
            "message": "No hay un mes para sincronizar.",
            "created": 0,
            "updated": 0,
            "deleted": 0,
            "skipped": 0,
            "affected_prestamos": 0,
        }

    archivo_path = os.path.join(_planillas_dir, mes["archivo"])
    if not os.path.exists(archivo_path):
        return {
            "status": "noop",
            "message": f"No se encontro el Excel del mes {mes['id']}.",
            "created": 0,
            "updated": 0,
            "deleted": 0,
            "skipped": 0,
            "affected_prestamos": 0,
        }

    semanas = plan_db.get_semanas_del_mes(mes["id"])
    summary = prestamo_sync.sync_rebajos_mes(mes, archivo_path, semanas)
    summary["status"] = "success"
    summary["mes_id"] = mes["id"]
    return summary


def _sync_prestamo_rebajos_todos():
    meses = plan_db.get_todos_meses()
    if not meses:
        return {
            "status": "noop",
            "message": "No hay meses registrados para sincronizar.",
            "created": 0,
            "updated": 0,
            "deleted": 0,
            "skipped": 0,
            "affected_prestamos": 0,
            "meses_sincronizados": 0,
        }

    total = {
        "status": "success",
        "created": 0,
        "updated": 0,
        "deleted": 0,
        "skipped": 0,
        "affected_prestamos": 0,
        "meses_sincronizados": 0,
    }
    for mes in meses:
        summary = _sync_prestamo_rebajos_mes(mes_data=mes)
        if summary["status"] != "success":
            continue
        total["created"] += summary["created"]
        total["updated"] += summary["updated"]
        total["deleted"] += summary["deleted"]
        total["skipped"] += summary["skipped"]
        total["affected_prestamos"] += summary["affected_prestamos"]
        total["meses_sincronizados"] += 1

    if total["meses_sincronizados"] == 0:
        total["status"] = "noop"
        total["message"] = "No se encontro ningun Excel de planilla para sincronizar."
    return total


@app.post("/api/planillas/prestamos/sync-planilla-rebajos")
def sync_planilla_rebajos_prestamo(mes_id: Optional[int] = None):
    if mes_id is None:
        return _sync_prestamo_rebajos_todos()
    return _sync_prestamo_rebajos_mes(mes_id=mes_id)


@app.get("/api/planillas/prestamos")
def get_all_prestamos_activos():
    prestamos = plan_db.get_todos_prestamos_activos()
    return prestamos

@app.get("/api/planillas/prestamos/{emp_id}")
def get_prestamos_emp(emp_id: int, solo_activos: bool = False):
    prestamos = plan_db.get_prestamos_empleado(emp_id, solo_activos=solo_activos)
    return prestamos

@app.post("/api/planillas/prestamos")
def add_prestamo(req: PlanillaPrestamo):
    pid = plan_db.add_prestamo(
        req.empleado_id, req.monto_total, req.pago_semanal, notas=req.notas
    )
    return {"status": "success", "id": pid}

@app.delete("/api/planillas/prestamos/{prestamo_id}")
def delete_prestamo(prestamo_id: int):
    plan_db.delete_prestamo(prestamo_id)
    return {"status": "success"}

@app.get("/api/planillas/prestamos/{prestamo_id}/abonos")
def get_abonos_prestamo(prestamo_id: int):
    abonos = plan_db.get_abonos(prestamo_id)
    prestamo = plan_db.get_prestamo(prestamo_id)
    return {"abonos": abonos, "prestamo": prestamo}

@app.post("/api/planillas/prestamos/{prestamo_id}/abono")
def add_abono_prestamo(prestamo_id: int, req: PrestamoAbono):
    abono_id = plan_db.add_abono(
        prestamo_id, req.monto, tipo=req.tipo,
        semana_planilla=req.semana_planilla, notas=req.notas,
        fecha=req.fecha
    )
    prestamo = plan_db.get_prestamo(prestamo_id)
    return {"status": "success", "id": abono_id, "nuevo_saldo": prestamo["saldo"], "estado": prestamo["estado"]}


@app.patch("/api/planillas/abonos/{abono_id}")
def update_abono_nota_endpoint(abono_id: int, body: dict):
    """Actualiza solo el campo notas de un abono."""
    plan_db.update_abono_nota(abono_id, body.get("notas", ""))
    return {"status": "success"}


@app.delete("/api/planillas/abonos/{abono_id}")
def delete_abono_endpoint(abono_id: int):
    """Elimina un abono extraordinario y recalcula el saldo."""
    try:
        prestamo_id = plan_db.delete_abono(abono_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Abono no encontrado")
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e))
    prestamo = plan_db.get_prestamo(prestamo_id)
    return {"status": "success", "nuevo_saldo": prestamo["saldo"], "estado": prestamo["estado"]}

# ------------------------------------------------------------------------------
# LIQUIDACIÓN — Cálculo automático según Ley Costarricense
# ------------------------------------------------------------------------------
@app.get("/api/planillas/liquidacion/{emp_id}")
def calcular_liquidacion(emp_id: int):
    """Calcula todos los rubros de liquidación para un empleado.
    
    - Vacaciones: días disponibles × tarifa_diurna × 8hrs
    - Aguinaldo proporcional: salarios brutos del período ÷ 12
    - Cesantía (Art. 29 CT CR): según tabla de antigüedad
    - Preaviso: 1 mes de salario promedio si >1 año
    """
    from datetime import date, timedelta
    import math

    emp_data = next((e for e in plan_db.get_empleados(solo_activos=False) if e["id"] == emp_id), None)
    if not emp_data:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    nombre = emp_data.get("nombre", "")
    fi = emp_data.get("fecha_inicio")
    tarifas = plan_db.get_tarifas()
    tarifa_diurna = tarifas.get("tarifa_diurna", 0)

    # ── 1. VACACIONES ──
    acumulados = plan_db.calcular_dias_vacaciones(fi) if fi else 0
    tomados = plan_db.total_dias_vacaciones_tomados(emp_id)
    vac_dias = max(0, acumulados - tomados)
    vac_monto = vac_dias * tarifa_diurna * 8  # jornada ordinaria diurna

    # ── 2. ANTIGÜEDAD ──
    antiguedad_dias = 0
    antiguedad_anios = 0
    if fi:
        try:
            inicio = datetime.datetime.strptime(fi, "%Y-%m-%d").date()
            antiguedad_dias = (date.today() - inicio).days
            antiguedad_anios = antiguedad_dias / 365.25
        except (ValueError, TypeError):
            print(f"liquidacion_preview: fecha_inicio inválida para empleado_id={emp_id}: {fi}")

    # ── 3. SALARIO PROMEDIO MENSUAL (de planillas reales) ──
    anio_actual = date.today().year
    salario_total_anual = 0.0
    semanas_con_datos = 0

    for anio in [anio_actual - 1, anio_actual]:
        meses = plan_db.get_meses_del_anio(anio)
        mes_activo = plan_db.get_mes_activo()
        if mes_activo and mes_activo['anio'] == anio:
            meses.append(mes_activo)

        for mes_data in meses:
            archivo = mes_data['archivo']
            archivo_path = os.path.join(_planillas_dir, archivo)
            if not os.path.exists(archivo_path):
                continue
            try:
                wb = openpyxl.load_workbook(archivo_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    if sheet_name.startswith("Semana "):
                        ws = wb[sheet_name]
                        for row in range(5, ws.max_row + 1):
                            name_cell = ws.cell(row=row, column=1).value
                            if name_cell and isinstance(name_cell, str) and name_cell.strip() == nombre:
                                # Find salario bruto column
                                for col in range(2, min(ws.max_column + 1, 20)):
                                    header = ws.cell(row=4, column=col).value
                                    if header and isinstance(header, str) and "bruto" in header.lower():
                                        val = ws.cell(row=row, column=col).value
                                        if val and isinstance(val, (int, float)):
                                            salario_total_anual += val
                                            semanas_con_datos += 1
                                        break
                wb.close()
            except (OSError, ValueError, KeyError):
                print(f"liquidacion_preview: no se pudo leer archivo de planilla {archivo}")

    # Promedio mensual = total / meses con datos (aprox 4 semanas = 1 mes)
    meses_con_datos = max(1, semanas_con_datos / 4.33)
    salario_promedio_mensual = salario_total_anual / meses_con_datos if meses_con_datos > 0 else 0

    # ── 4. AGUINALDO PROPORCIONAL ──
    # En CR: aguinaldo = salarios de dic anterior a nov actual ÷ 12
    aguinaldo_monto = salario_total_anual / 12.0 if salario_total_anual > 0 else 0

    # ── 5. CESANTÍA (Art. 29 Código de Trabajo CR) ──
    # Solo aplica a despido sin justa causa
    cesantia_dias = 0
    if antiguedad_dias >= 90:  # Mínimo 3 meses
        if antiguedad_dias < 180:       # 3-6 meses
            cesantia_dias = 7
        elif antiguedad_dias < 365:     # 6-12 meses
            cesantia_dias = 14
        else:
            # Más de 1 año: tabla progresiva
            anios_completos = int(antiguedad_anios)
            if anios_completos == 1:
                cesantia_dias = 19.5
            elif anios_completos == 2:
                cesantia_dias = 19.5 + 19.5
            elif anios_completos == 3:
                cesantia_dias = 19.5 + 19.5 + 20
            elif anios_completos == 4:
                cesantia_dias = 19.5 + 19.5 + 20 + 20
            elif anios_completos == 5:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5
            elif anios_completos == 6:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5 + 21
            elif anios_completos >= 7:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5 + 21 + 21.24
            if anios_completos >= 8:
                cesantia_dias = 19.5 + 19.5 + 20 + 20 + 20.5 + 21 + 21.24 + 21.5
            # Máximo legal: ~22 días por año, tope 8 años

    salario_diario = salario_promedio_mensual / 30 if salario_promedio_mensual > 0 else 0
    cesantia_monto = cesantia_dias * salario_diario

    # ── 6. PREAVISO ──
    # >1 año: 1 mes; 3-6 meses: 1 semana; 6-12 meses: 15 días
    preaviso_monto = 0
    if antiguedad_dias >= 365:
        preaviso_monto = salario_promedio_mensual  # 1 mes
    elif antiguedad_dias >= 180:
        preaviso_monto = salario_promedio_mensual / 2  # 15 días
    elif antiguedad_dias >= 90:
        preaviso_monto = salario_promedio_mensual / 4.33  # 1 semana

    return {
        "emp_id": emp_id,
        "nombre": nombre,
        "fecha_inicio": fi,
        "antiguedad_anios": round(antiguedad_anios, 1),
        "tarifa_diurna": tarifa_diurna,
        "vacaciones_dias": vac_dias,
        "vacaciones_monto": round(vac_monto, 2),
        "aguinaldo_monto": round(aguinaldo_monto, 2),
        "cesantia_dias": round(cesantia_dias, 1),
        "cesantia_monto": round(cesantia_monto, 2),
        "preaviso_monto": round(preaviso_monto, 2),
        "salario_promedio_mensual": round(salario_promedio_mensual, 2),
        "total_despido": round(vac_monto + aguinaldo_monto + cesantia_monto + preaviso_monto, 2),
        "total_renuncia": round(vac_monto + aguinaldo_monto, 2),
    }


def calcular_aguinaldo(anio: int):
    """Calcula aguinaldo basado en la tabla salarios_mensuales de la BD."""
    MES_NOMBRES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                   7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

    desglose_rows = plan_db.get_salarios_anio_desglose(anio)
    if not desglose_rows:
        return {"status": "error", "message": f"No hay salarios registrados para {anio}", "data": []}

    # Group by employee
    emp_data = {}
    for row in desglose_rows:
        eid = row['empleado_id']
        if eid not in emp_data:
            emp_data[eid] = {
                'nombre': row['empleado_nombre'],
                'total_bruto': 0.0,
                'desglose': []
            }
        mes_label = MES_NOMBRES.get(row['mes'], f'Mes {row["mes"]}')
        emp_data[eid]['total_bruto'] += row['total_bruto']
        emp_data[eid]['desglose'].append({
            'mes': mes_label,
            'bruto': round(row['total_bruto'], 2)
        })

    # Get employee info from DB
    emps = plan_db.get_empleados(solo_activos=False)
    emp_lookup = {e['id']: e for e in emps}

    results = []
    total_aguinaldo = 0.0
    meses_set = set()

    for eid, data in emp_data.items():
        emp = emp_lookup.get(eid, {})
        sal_total = data['total_bruto']
        aguinaldo = sal_total / 12.0 if sal_total > 0 else 0.0
        total_aguinaldo += aguinaldo
        for d in data['desglose']:
            meses_set.add(d['mes'])
        results.append({
            "id": eid,
            "nombre": data['nombre'],
            "cedula": emp.get("cedula") or "—",
            "salario_anual": round(sal_total, 2),
            "aguinaldo": round(aguinaldo, 2),
            "desglose_mensual": data['desglose']
        })

    # Also include employees with no salary data
    for emp in emps:
        if emp['id'] not in emp_data:
            results.append({
                "id": emp['id'],
                "nombre": emp['nombre'],
                "cedula": emp.get("cedula") or "—",
                "salario_anual": 0,
                "aguinaldo": 0,
                "desglose_mensual": []
            })

    return {
        "status": "success",
        "data": results,
        "total_aguinaldo": round(total_aguinaldo, 2),
        "meses_evaluados": len(meses_set)
    }


@app.get("/api/planillas/aguinaldo/{anio}")
def get_aguinaldo(anio: int):
    return calcular_aguinaldo(anio)


# ------------------------------------------------------------------------------
# PLANILLAS - GUARDAR SALARIOS
# ------------------------------------------------------------------------------
class SalarioSemanal(BaseModel):
    empleado_id: int
    empleado_nombre: str
    anio: int
    mes: int
    semana: int
    salario_bruto: float

class SalariosLote(BaseModel):
    salarios: list[SalarioSemanal]

@app.post("/api/planillas/salarios")
def guardar_salarios(req: SalariosLote):
    for s in req.salarios:
        plan_db.guardar_salario_semanal(
            s.empleado_id, s.empleado_nombre,
            s.anio, s.mes, s.semana, s.salario_bruto
        )
    return {"status": "success", "count": len(req.salarios)}

@app.post("/api/planillas/sincronizar-aguinaldo/{anio}")
def sincronizar_aguinaldo_anio(anio: int):
    """
    Escanea los Excel (cerrados y activo) para el año dado, lee todas las hojas 
    de semanas y guarda los salarios brutos en la DB para reconstruir el historial.
    """
    meses_a_sincronizar = plan_db.get_meses_del_anio(anio)
    mes_activo = plan_db.get_mes_activo()
    if mes_activo and mes_activo["anio"] == anio:
        if not any(m["archivo"] == mes_activo["archivo"] for m in meses_a_sincronizar):
            meses_a_sincronizar.append(mes_activo)

    # Deduplicar por archivo — evita procesar el mismo Excel dos veces si
    # quedaron registros duplicados de meses en la base de datos.
    seen_archivos = set()
    meses_unicos = []
    for m in meses_a_sincronizar:
        if m["archivo"] not in seen_archivos:
            seen_archivos.add(m["archivo"])
            meses_unicos.append(m)
    meses_a_sincronizar = meses_unicos

    if not meses_a_sincronizar:
        return {"status": "error", "message": f"No se encontraron meses para el año {anio} para sincronizar."}
    
    count = 0
    skipped = []
    errores = []
    tarifas = plan_db.get_tarifas()
    emps = plan_db.get_empleados(solo_activos=False)
    emp_map = {e["nombre"]: e["id"] for e in emps}

    for mes in meses_a_sincronizar:
        archivo_path = os.path.join(_planillas_dir, mes["archivo"])
        if not os.path.exists(archivo_path):
            # Archivo no existe: ignorar silenciosamente (puede ser un mes
            # registrado por error en la DB sin Excel real).
            skipped.append(mes["archivo"])
            continue
            
        try:
            wb = openpyxl.load_workbook(archivo_path, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith("Semana "):
                    try:
                        sem_num = int(sheet_name.split(" ")[-1])
                    except (ValueError, IndexError):
                        print(f"sync_salarios_historicos: nombre de hoja no estándar '{sheet_name}' en {mes['archivo']}")
                        continue
                    
                    ws = wb[sheet_name]
                    bruto_col = 12 # Default based on rellenar_horas_en_excel logic and template
                    
                    for r in range(5, ws.max_row + 1):
                        name_cell = ws.cell(row=r, column=1).value
                        if name_cell and isinstance(name_cell, str) and name_cell.strip() in emp_map:
                            nombre = name_cell.strip()
                            bruto_val = ws.cell(row=r, column=bruto_col).value
                            if bruto_val and isinstance(bruto_val, (int, float)):
                                plan_db.guardar_salario_semanal(
                                    emp_map[nombre], nombre, anio, mes["mes"], sem_num, bruto_val
                                )
                                count += 1
            wb.close()
        except Exception as e:
            errores.append(f"Error procesando {mes['archivo']}: {str(e)}")

    msg = f"Sincronizados {count} registros de salarios."
    if skipped:
        print(f"Sincronización: {len(skipped)} archivos omitidos (no existen): {skipped}")
    if errores:
        msg += f" ({len(errores)} errores, revisa la consola para detalles)"
        print("Errores de sincronización:", errores)
        
    return {"status": "success", "message": msg, "count": count}


# ------------------------------------------------------------------------------
# PLANILLAS - guardado Excel (archivo bloqueado por Excel/OneDrive)
# ------------------------------------------------------------------------------
def _is_excel_file_lock_error(e: BaseException) -> bool:
    if isinstance(e, PermissionError):
        return True
    if isinstance(e, OSError):
        return getattr(e, "errno", None) in (13, 16) or getattr(e, "winerror", None) in (32, 33)
    return False


def _save_workbook_planilla(wb, path: str) -> None:
    """Guarda el workbook; si el destino está bloqueado, intenta vía archivo temporal."""
    try:
        wb.save(path)
        return
    except (PermissionError, OSError) as e:
        if not _is_excel_file_lock_error(e):
            raise
    dir_name = os.path.dirname(os.path.abspath(path)) or "."
    os.makedirs(dir_name, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(prefix=".xlsx_save_", suffix=".tmp", dir=dir_name)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception as e:
        try:
            if os.path.isfile(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
        raise HTTPException(
            status_code=409,
            detail=(
                "Tienes el excel abierto, necesito que lo cierres, no seas inteligente asintomatico :)"
            ),
        ) from e


# ------------------------------------------------------------------------------
# PLANILLAS - TARIFAS
# ------------------------------------------------------------------------------
@app.get("/api/planillas/tarifas")
def get_planillas_tarifas():
    return plan_db.get_tarifas()

@app.post("/api/planillas/tarifas")
def update_planillas_tarifas(t: PlanillaTarifas):
    modo = (t.seguro_modo or "porcentual").strip().lower()
    if modo not in ("porcentual", "fijo"):
        raise HTTPException(status_code=400, detail="seguro_modo debe ser 'porcentual' o 'fijo'.")
    plan_db.set_tarifas(
        t.tarifa_diurna,
        t.tarifa_nocturna,
        t.tarifa_mixta,
        modo,
        float(t.seguro_valor),
        1 if t.pagar_horas_extra else 0,
    )
    return {"status": "success"}

# ------------------------------------------------------------------------------
# PLANILLAS - MESES & SEMANAS
# ------------------------------------------------------------------------------
@app.get("/api/planillas/meses")
def get_todos_meses():
    meses = plan_db.get_todos_meses()
    for m in meses:
        m["semanas"] = plan_db.get_semanas_del_mes(m["id"])
    return meses

@app.get("/api/planillas/meses/activo")
def get_mes_activo():
    mes = plan_db.get_mes_activo()
    if not mes:
        return {"mes": None, "semanas": []}
    semanas = plan_db.get_semanas_del_mes(mes["id"])
    return {"mes": mes, "semanas": semanas}

@app.post("/api/planillas/meses")
def create_mes(m: PlanillaMes):
    emps = plan_db.get_empleados()
    if not emps:
        raise HTTPException(status_code=400, detail="Agregue empleados antes de crear la planilla.")
    
    # Crear Excel del Mes
    meses_nombres = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    mes_name = meses_nombres.get(m.mes, str(m.mes))
    
    # Create nested folder structure: Planillas {Anio} / {Mes}
    folder_year = f"Planillas {m.anio}"
    folder_month = mes_name
    rel_folder_path = os.path.join(folder_year, folder_month)
    abs_folder_path = os.path.join(_planillas_dir, rel_folder_path)
    os.makedirs(abs_folder_path, exist_ok=True)
    
    archivo_name = f"Planilla_{mes_name}_{m.anio}.xlsx"
    archivo_rel_path = os.path.join(rel_folder_path, archivo_name).replace("\\", "/")
    archivo_path = os.path.join(abs_folder_path, archivo_name)
    
    # Init Excel
    tarifas = plan_db.get_tarifas()
    pl_module.TARIFA_DIURNA = tarifas["tarifa_diurna"]
    pl_module.TARIFA_NOCTURNA = tarifas["tarifa_nocturna"]
    pl_module.TARIFA_MIXTA = tarifas["tarifa_mixta"]
    
    wb = openpyxl.Workbook()
    pl_module.SAMPLE_EMPS = []
    pl_module.crear_catalogo(wb)
    
    ws_cat = wb["Catalogo"]
    tipo_map = {"tarjeta": "Transferencia Bancaria", "efectivo": "Efectivo", "fijo": "Salario Fijo"}
    for i, emp in enumerate(emps):
        r = 5 + i
        pl_module.sc(ws_cat, r, 1, emp["nombre"], pl_module.f_data, pl_module.ROW_1 if i % 2 == 0 else pl_module.ROW_2, pl_module.al_l, pl_module.borde)
        pl_module.sc(ws_cat, r, 2, tipo_map.get(emp["tipo_pago"], emp["tipo_pago"]),
                     pl_module.f_data, pl_module.ROW_1 if i % 2 == 0 else pl_module.ROW_2, pl_module.al_c, pl_module.borde)
        if emp["salario_fijo"]:
            pl_module.sc(ws_cat, r, 3, emp["salario_fijo"],
                         pl_module.f_data, pl_module.ROW_1 if i % 2 == 0 else pl_module.ROW_2, pl_module.al_c, pl_module.borde, pl_module.MONEY)
                         
    pl_module.crear_resumen_mensual(wb)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    _save_workbook_planilla(wb, archivo_path)
    
    # Registrar en DB
    mes_db = plan_db.crear_mes(m.anio, m.mes, archivo_rel_path)
    return {"status": "success", "mes": mes_db}

@app.post("/api/planillas/meses/{mes_id}/cerrar")
def cerrar_mes(mes_id: int):
    plan_db.cerrar_mes(mes_id)
    return {"status": "success"}

@app.post("/api/planillas/semanas")
def agregar_semana(s: PlanillaSemana):
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo or mes_activo["id"] != s.mes_id:
        raise HTTPException(status_code=400, detail="Mes inválido o cerrado.")
        
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=400, detail=f"No se encontró el Excel: {mes_activo['archivo']}")
        
    viernes_date = datetime.datetime.strptime(s.viernes, "%Y-%m-%d").date()
    
    tarifas = plan_db.get_tarifas()
    pl_module.TARIFA_DIURNA = tarifas["tarifa_diurna"]
    pl_module.TARIFA_NOCTURNA = tarifas["tarifa_nocturna"]
    pl_module.TARIFA_MIXTA = tarifas["tarifa_mixta"]
    
    wb = openpyxl.load_workbook(archivo_path)
    empleados = pl_module.leer_catalogo(wb)
    total_emp = sum(len(v) for v in empleados.values())
    if total_emp == 0:
        raise HTTPException(status_code=400, detail="No hay empleados en el Excel.")
        
    num = pl_module.contar_semanas(wb) + 1
    sem_num = pl_module.num_semana_anual(viernes_date)
    
    _cfg_plan = load_db()
    nombre_hoja, gran_row, section_totals = pl_module.crear_hoja_semanal(
        wb, num, viernes_date, empleados, tarifas=tarifas,
        holiday_dates=_cfg_plan.get("config", {}).get("holidays", []),
    )
    pl_module.crear_resumen_semanal(wb, nombre_hoja, sem_num, viernes_date)
    pl_module.crear_resumen_mensual(wb)
    pl_module.crear_dashboard(wb)
    
    _save_workbook_planilla(wb, archivo_path)
    
    # Save to DB
    plan_db.add_semana(s.mes_id, sem_num, s.viernes)
    semanas = plan_db.get_semanas_del_mes(s.mes_id)
    return {"status": "success", "semanas": semanas}

@app.delete("/api/planillas/semanas/{semana_id}")
def delete_semana(semana_id: int):
    """Elimina una semana de cualquier mes (activo o cerrado).
    
    Busca el mes al que pertenece la semana entre TODOS los meses registrados,
    no solo el activo. Esto permite editar el historial de planillas anteriores.
    """
    # Buscar la semana y su mes en todos los meses registrados
    all_months = plan_db.get_todos_meses()
    target_sem = None
    target_mes = None

    for mes in all_months:
        semanas = plan_db.get_semanas_del_mes(mes["id"])
        found = next((s for s in semanas if s["id"] == semana_id), None)
        if found:
            target_sem = found
            target_mes = mes
            break

    if not target_sem or not target_mes:
        raise HTTPException(status_code=404, detail="Semana no encontrada")

    num_semana = target_sem["num_semana"]
    plan_db.delete_semana(semana_id)

    # Eliminar hojas del Excel si el archivo existe
    archivo_path = os.path.join(_planillas_dir, target_mes["archivo"])
    if os.path.exists(archivo_path):
        try:
            wb = openpyxl.load_workbook(archivo_path)
            sheet_sem = f"Semana {num_semana}"
            sheet_res = f"Res. Sem. {num_semana}"
            if sheet_sem in wb.sheetnames: wb.remove(wb[sheet_sem])
            if sheet_res in wb.sheetnames: wb.remove(wb[sheet_res])
            pl_module.crear_resumen_mensual(wb)
            pl_module.crear_dashboard(wb)
            _save_workbook_planilla(wb, archivo_path)
        except HTTPException:
            raise
        except Exception as e:
            print(f"Advertencia al modificar Excel: {e}")

    _sync_prestamo_rebajos_mes(mes_data=target_mes)

    return {"status": "success", "semanas": plan_db.get_semanas_del_mes(target_mes["id"])}


@app.delete("/api/planillas/meses/{mes_id}")
def delete_mes_historial(mes_id: int):
    """Elimina un mes cerrado y todos sus registros. Rechaza meses activos."""
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")
    if not mes_target.get("cerrado"):
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar meses cerrados")

    prestamo_sync.clear_auto_rebajos_mes(mes_id)

    # Intentar borrar el Excel físico (no fatal si falla)
    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if os.path.exists(archivo_path):
        try:
            os.remove(archivo_path)
        except Exception as e:
            print(f"Advertencia: no se pudo eliminar el archivo Excel: {e}")

    plan_db.delete_mes(mes_id)
    return {"status": "success", "message": f"Mes {mes_id} eliminado"}


@app.post("/api/planillas/meses/{mes_id}/semanas")
def agregar_semana_a_mes_historico(mes_id: int, s: PlanillaSemana):
    """Agrega una semana a cualquier mes (activo o cerrado), para edición del historial.
    
    A diferencia del endpoint /api/planillas/semanas, este no exige que el mes
    sea el activo, permitiendo gestionar semanas en planillas anteriores.
    """
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")

    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=400, detail=f"No se encontró el Excel: {mes_target['archivo']}")

    viernes_date = datetime.datetime.strptime(s.viernes, "%Y-%m-%d").date()

    tarifas = plan_db.get_tarifas()
    pl_module.TARIFA_DIURNA = tarifas["tarifa_diurna"]
    pl_module.TARIFA_NOCTURNA = tarifas["tarifa_nocturna"]
    pl_module.TARIFA_MIXTA = tarifas["tarifa_mixta"]

    wb = openpyxl.load_workbook(archivo_path)
    empleados = pl_module.leer_catalogo(wb)
    total_emp = sum(len(v) for v in empleados.values())
    if total_emp == 0:
        raise HTTPException(status_code=400, detail="No hay empleados en el Excel.")

    num = pl_module.contar_semanas(wb) + 1
    sem_num = pl_module.num_semana_anual(viernes_date)

    _cfg_plan_h = load_db()
    nombre_hoja, gran_row, section_totals = pl_module.crear_hoja_semanal(
        wb, num, viernes_date, empleados, tarifas=tarifas,
        holiday_dates=_cfg_plan_h.get("config", {}).get("holidays", []),
    )
    pl_module.crear_resumen_semanal(wb, nombre_hoja, sem_num, viernes_date)
    pl_module.crear_resumen_mensual(wb)
    pl_module.crear_dashboard(wb)

    _save_workbook_planilla(wb, archivo_path)

    plan_db.add_semana(mes_id, sem_num, s.viernes)
    semanas = plan_db.get_semanas_del_mes(mes_id)
    return {"status": "success", "semanas": semanas}

@app.get("/api/planillas/horarios-disponibles")
def get_horarios_disponibles():
    horarios = horario_db.get_horarios_generados()
    return {"status": "success", "horarios": horarios}


@app.put("/api/planillas/horarios/{horario_id}/tareas")
def update_horario_tareas_endpoint(horario_id: int, tareas: Dict[str, Any]):
    """Actualiza las tareas de limpieza de un horario existente."""
    horario = horario_db.get_horario_por_id(horario_id)
    if not horario:
        raise HTTPException(status_code=404, detail="Horario no encontrado")
    horario_db.update_horario_tareas(horario_id, tareas)
    return {"status": "success", "message": "Tareas actualizadas"}


@app.post("/api/planillas/semanas/importar")
def importar_horario_semana(req: PlanillaImportarHorario):
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo:
        raise HTTPException(status_code=400, detail="No hay un mes activo")
        
    horario = horario_db.get_horario_por_id(req.horario_id)
    if not horario:
        raise HTTPException(status_code=404, detail="Horario no encontrado")
        
    if req.sync_empleados:
        horario_db.sincronizar_empleados_a_planilla()
        
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    _cfg = load_db()
    ok, msg, hours_data = horario_db.rellenar_horas_en_excel(
        archivo_path, req.semana_nombre, horario["horario"],
        holidays=_cfg.get("config", {}).get("holidays", []),
    )
    
    if not ok:
        raise HTTPException(status_code=400, detail=msg)

    # Store salaries in DB
    try:
        tarifas = plan_db.get_tarifas()
        emps = plan_db.get_empleados(solo_activos=False)
        emp_map = {e["nombre"]: e["id"] for e in emps}
        
        # Get week number from semantic name "Semana X"
        try:
            sem_num = int(req.semana_nombre.split(" ")[-1])
        except (ValueError, IndexError):
            print(f"importar_horario_semana: semana_nombre no parseable '{req.semana_nombre}', usando 0")
            sem_num = 0

        for emp_nombre, h in hours_data.items():
            if emp_nombre in emp_map:
                bruto = h.get("salario_bruto")
                if bruto is None:
                    td = float(tarifas["tarifa_diurna"])
                    tm = float(tarifas["tarifa_mixta"])
                    tn = float(tarifas["tarifa_nocturna"])
                    bruto = (
                        h["diurnas"] * td
                        + h["mixtas"] * tm
                        + h["nocturnas"] * tn
                        + float(h.get("recargo_feriado") or 0)
                    )
                    if int(tarifas.get("pagar_horas_extra") or 1):
                        ed = float(h.get("extra_diurnas") or 0)
                        em = float(h.get("extra_mixtas") or 0)
                        en = float(h.get("extra_nocturnas") or 0)
                        if ed + em + en == 0:
                            ed = float(h.get("extra") or 0)
                        bruto += ed * td * 1.5 + em * tm * 1.5 + en * tn * 1.5

                plan_db.guardar_salario_semanal(
                    emp_map[emp_nombre], emp_nombre,
                    mes_activo["anio"], mes_activo["mes"], sem_num, bruto
                )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error guardando salarios semanales: {e}")

    return {"status": "success", "message": msg}

@app.post("/api/planillas/meses/{mes_id}/semanas/{semana_nombre}/importar")
def importar_horario_semana_historico(mes_id: int, semana_nombre: str, req: PlanillaImportarHorario):
    """Importa un horario guardado a la hoja de una semana específica de CUALQUIER mes.
    
    A diferencia del endpoint /api/planillas/semanas/importar, este no exige que el mes
    sea el activo, permitiendo rellenar horas en planillas históricas o cerradas.
    """
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")

    horario = horario_db.get_horario_por_id(req.horario_id)
    if not horario:
        raise HTTPException(status_code=404, detail="Horario no encontrado")

    if req.sync_empleados:
        horario_db.sincronizar_empleados_a_planilla()

    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=404, detail=f"No se encontró el Excel: {mes_target['archivo']}")

    _cfg_h = load_db()
    ok, msg, hours_data = horario_db.rellenar_horas_en_excel(
        archivo_path, semana_nombre, horario["horario"],
        holidays=_cfg_h.get("config", {}).get("holidays", []),
    )

    if not ok:
        raise HTTPException(status_code=400, detail=msg)

    # Guardar salarios en DB
    try:
        tarifas = plan_db.get_tarifas()
        emps = plan_db.get_empleados(solo_activos=False)
        emp_map = {e["nombre"]: e["id"] for e in emps}

        try:
            sem_num = int(semana_nombre.split(" ")[-1])
        except (ValueError, IndexError):
            print(f"importar_horario_semana_historico: semana_nombre no parseable '{semana_nombre}', usando 0")
            sem_num = 0

        for emp_nombre, h in hours_data.items():
            if emp_nombre in emp_map:
                bruto = h.get("salario_bruto")
                if bruto is None:
                    td = float(tarifas["tarifa_diurna"])
                    tm = float(tarifas["tarifa_mixta"])
                    tn = float(tarifas["tarifa_nocturna"])
                    bruto = (
                        h["diurnas"] * td
                        + h["mixtas"] * tm
                        + h["nocturnas"] * tn
                        + float(h.get("recargo_feriado") or 0)
                    )
                    if int(tarifas.get("pagar_horas_extra") or 1):
                        ed = float(h.get("extra_diurnas") or 0)
                        em = float(h.get("extra_mixtas") or 0)
                        en = float(h.get("extra_nocturnas") or 0)
                        if ed + em + en == 0:
                            ed = float(h.get("extra") or 0)
                        bruto += ed * td * 1.5 + em * tm * 1.5 + en * tn * 1.5

                plan_db.guardar_salario_semanal(
                    emp_map[emp_nombre], emp_nombre,
                    mes_target["anio"], mes_target["mes"], sem_num, bruto
                )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar salarios semanales: {e}")

    return {"status": "success", "message": msg}


@app.post("/api/planillas/boletas/generar")
def generar_boletas(req: PlanillaBoletasRequest):
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo:
        raise HTTPException(status_code=400, detail="No active month to generate boletas.")

    _sync_prestamo_rebajos_mes(mes_data=mes_activo)

    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    logo_path = os.path.join(_planillas_dir, "logo.png")
    
    ok, msg, out_dir = gb_module.generar_boletas_semana(
        archivo_path, 
        req.semana_nombre, 
        logo_path,
        tipo_bono=req.tipo_bono,
        valor_sticker=req.valor_sticker
    )
    
    if ok:
        # Open folder dynamically on Windows
        if os.name == 'nt':
            os.startfile(out_dir)
        return {"status": "success", "message": msg, "out_dir": out_dir}
    else:
        raise HTTPException(status_code=400, detail=msg)

@app.get("/api/planillas/excel/abrir")
def abrir_excel_activo():
    mes_activo = plan_db.get_mes_activo()
    if not mes_activo:
        raise HTTPException(status_code=400, detail="No hay mes activo.")
    
    archivo_path = os.path.join(_planillas_dir, mes_activo["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=404, detail="No se encontró el archivo Excel.")
        
    if os.name == 'nt':
        os.startfile(archivo_path)
    return {"status": "success"}

@app.get("/api/planillas/excel/abrir/{mes_id}")
def abrir_excel_por_id(mes_id: int):
    meses = plan_db.get_todos_meses()
    mes_target = next((m for m in meses if m["id"] == mes_id), None)
    
    if not mes_target:
        raise HTTPException(status_code=404, detail="Mes no encontrado")
        
    archivo_path = os.path.join(_planillas_dir, mes_target["archivo"])
    if not os.path.exists(archivo_path):
        raise HTTPException(status_code=404, detail="Archivo Excel no encontrado")
        
    if os.name == 'nt':
        os.startfile(archivo_path)
    return {"status": "success"}

# ==============================================================================
# UTILIDADES — GENERADOR DE DOCUMENTOS WORD
# ==============================================================================
import docx_generator

class DocPrestamo(BaseModel):
    emp_id: int
    monto_total: float
    pago_semanal: float

class AmoDatoItem(BaseModel):
    fecha: Optional[str] = None
    monto: Optional[float] = None
    minutos: Optional[int] = None

class DocAmonestacion(BaseModel):
    emp_id: int
    tipo: str  # "faltantes", "tardanzas", "conductas"
    datos: List[AmoDatoItem] = Field(default_factory=list)

class DocVacacionesReq(BaseModel):
    emp_id: int
    tipo: str  # 'total' or 'parcial'
    fecha_inicio: str = ""
    fecha_reingreso: str = ""
    solo_pago: bool = False
    modo_periodo: bool = False
    periodo_texto: str = ""
    dias_periodo: float = 0.0

class DocLiquidacion(BaseModel):
    emp_id: int
    vacaciones_dias: float
    vacaciones_monto: float
    aguinaldo_monto: float
    cesantia_monto: float = 0
    preaviso_monto: float = 0
    total_pagar: float
    modo_pago: str  # "Abonos" o "Total"

def _get_emp_info(emp_id: int):
    emps = plan_db.get_empleados(solo_activos=False)
    emp = next((e for e in emps if e["id"] == emp_id), None)
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    return emp.get("nombre", ""), emp.get("cedula", "")

@app.post("/api/utilidades/prestamo")
def generar_doc_prestamo(req: DocPrestamo):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    path = docx_generator.generar_prestamo(nombre, cedula, req.monto_total, req.pago_semanal, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/amonestacion")
def generar_doc_amonestacion(req: DocAmonestacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    datos_dict = [d.dict(exclude_none=True) for d in req.datos]
    path = docx_generator.generar_amonestacion(nombre, cedula, req.tipo, datos_dict, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/vacaciones")
def generar_doc_vacaciones(req: DocVacacionesReq):
    nombre, cedula = _get_emp_info(req.emp_id)
    
    # Calculate payout based on base tariff from planillas settings
    total_pagar = 0.0
    tarifa_diurna = 0.0
    horas_totales = 0.0
    if req.solo_pago:
        try:
            if req.modo_periodo:
                dias_calculados = float(req.dias_periodo)
            else:
                from datetime import datetime
                dt_inicio = datetime.strptime(req.fecha_inicio, "%Y-%m-%d")
                dt_reingreso = datetime.strptime(req.fecha_reingreso, "%Y-%m-%d")
                dias_calculados = (dt_reingreso - dt_inicio).days
            
            tarifas = plan_db.get_tarifas()
            tarifa_diurna = float(tarifas.get("tarifa_diurna", 0.0))
            horas_totales = dias_calculados * 8.0
            total_pagar = horas_totales * tarifa_diurna
        except Exception:
            pass

    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    path = docx_generator.generar_vacaciones(
        nombre, cedula, req.tipo, req.fecha_inicio, req.fecha_reingreso, 
        logo, base, solo_pago=req.solo_pago, total_pagar=total_pagar,
        modo_periodo=req.modo_periodo, periodo_texto=req.periodo_texto, dias_periodo=req.dias_periodo,
        tarifa_diurna=tarifa_diurna, horas_totales=horas_totales
    )
    return {"status": "success", "path": path}

@app.post("/api/utilidades/despido")
def generar_doc_despido(req: DocLiquidacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    path = docx_generator.generar_liquidacion("Despido", nombre, cedula, req.vacaciones_dias, req.vacaciones_monto, req.aguinaldo_monto, req.cesantia_monto, req.preaviso_monto, req.total_pagar, req.modo_pago, logo, base)
    return {"status": "success", "path": path}

@app.post("/api/utilidades/renuncia")
def generar_doc_renuncia(req: DocLiquidacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    path = docx_generator.generar_liquidacion("Renuncia", nombre, cedula, req.vacaciones_dias, req.vacaciones_monto, req.aguinaldo_monto, 0, 0, req.total_pagar, req.modo_pago, logo, base)
    return {"status": "success", "path": path}


@app.post("/api/utilidades/prestamo-carta/{prestamo_id}")
def generar_carta_prestamo(prestamo_id: int):
    prestamo = plan_db.get_prestamo(prestamo_id)
    if not prestamo:
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")
    emp_id = prestamo.get("empleado_id")
    emps = plan_db.get_empleados(solo_activos=False)
    emp = next((e for e in emps if e["id"] == emp_id), None)
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    nombre = emp.get("nombre", "")
    cedula = emp.get("cedula", "") or ""
    monto_total = float(prestamo.get("monto_total", 0))
    pago_semanal = float(prestamo.get("pago_semanal", 0))
    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    path = docx_generator.generar_prestamo(nombre, cedula, monto_total, pago_semanal, logo, base)
    return {"status": "success", "path": path}

class DocRecomendacion(BaseModel):
    emp_id: int
    puesto: str
    texto_adicional: str = ""

@app.post("/api/utilidades/recomendacion")
def generar_doc_recomendacion(req: DocRecomendacion):
    nombre, cedula = _get_emp_info(req.emp_id)
    emps = plan_db.get_empleados(solo_activos=False)
    emp = next((e for e in emps if e['id'] == req.emp_id), None)
    fecha_inicio = emp.get("fecha_inicio", "") if emp else ""
    logo = os.path.join(_planillas_dir, "logo.png")
    base = _runtime_root
    path = docx_generator.generar_recomendacion(nombre, cedula, req.puesto, fecha_inicio, req.texto_adicional, logo, base)
    return {"status": "success", "path": path}

# ==============================================================================
# INVENTARIO API ENDPOINTS
# ==============================================================================
from fastapi import UploadFile, File

INVENTARIO_BASE_DEFAULT_PATH = r"c:\Users\kenda\Downloads\Control de Inventario Pista.xlsx"


def _normalize_inventory_key(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def _to_float(value, default=0):
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _load_inventario_base_from_excel(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    articulos = []
    try:
        for sheet_name in ("S", "S2"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row_idx in range(3, ws.max_row + 1):
                nombre = ws.cell(row=row_idx, column=1).value
                if nombre is None or str(nombre).strip() == "":
                    continue
                articulos.append({
                    "codigo": str(ws.cell(row=row_idx, column=2).value or "").strip(),
                    "nombre": str(nombre).strip(),
                    "precio": _to_float(ws.cell(row=row_idx, column=3).value, 0),
                    "existencias_base": _to_float(ws.cell(row=row_idx, column=4).value, 0),
                    "hoja_origen": sheet_name,
                    "orden": len(articulos) + 1,
                })
    finally:
        wb.close()
    return articulos


def _ensure_inventario_base_seeded():
    current = plan_db.get_inventario_base()
    if current["articulos"]:
        return current
    if os.path.exists(INVENTARIO_BASE_DEFAULT_PATH):
        articulos = _load_inventario_base_from_excel(INVENTARIO_BASE_DEFAULT_PATH)
        if articulos:
            plan_db.replace_inventario_base(articulos, source_path=INVENTARIO_BASE_DEFAULT_PATH)
            return plan_db.get_inventario_base()
    return current


def _compare_inventory_against_base(upload_items, base_items):
    # Build lookup by (codigo, nombre) — both must match exactly
    base_by_key = {}
    for base in base_items:
        code = _normalize_inventory_key(base.get("codigo"))
        name = _normalize_inventory_key(base.get("nombre"))
        if code and name:
            base_by_key[(code, name)] = base

    rows = []
    matched_base_ids = set()
    resumen = {
        "total_base": len(base_items),
        "total_cargados": len(upload_items),
        "coinciden": 0,
        "con_diferencia": 0,
        "faltantes_en_carga": 0,
    }

    for item in upload_items:
        item_code = _normalize_inventory_key(item.get("codigo"))
        item_name = _normalize_inventory_key(item.get("nombre"))
        match = base_by_key.get((item_code, item_name)) if item_code and item_name else None

        if match:
            matched_base_ids.add(match["id"])
            delta = _to_float(item.get("existencias"), 0) - _to_float(match.get("existencias_base"), 0)
            resumen["coinciden" if delta == 0 else "con_diferencia"] += 1
            rows.append({
                "status": "match" if delta == 0 else "difference",
                "base_id": match["id"],
                "hoja_origen": match.get("hoja_origen"),
                "codigo": item.get("codigo") or match.get("codigo", ""),
                "nombre": item.get("nombre") or match.get("nombre", ""),
                "precio": item.get("precio", 0) or match.get("precio", 0),
                "existencias_base": _to_float(match.get("existencias_base"), 0),
                "existencias_actual": _to_float(item.get("existencias"), 0),
                "delta": delta,
            })
        # Items not in base are simply ignored — no "new" rows

    for base in base_items:
        if base["id"] in matched_base_ids:
            continue
        resumen["faltantes_en_carga"] += 1
        rows.append({
            "status": "missing",
            "base_id": base["id"],
            "hoja_origen": base.get("hoja_origen"),
            "codigo": base.get("codigo", ""),
            "nombre": base.get("nombre", ""),
            "precio": base.get("precio", 0),
            "existencias_base": _to_float(base.get("existencias_base"), 0),
            "existencias_actual": None,
            "delta": None,
        })

    order = {"difference": 0, "missing": 1, "match": 2}
    rows.sort(key=lambda r: (order.get(r["status"], 9), r.get("hoja_origen") or "Z", r.get("nombre") or ""))
    return rows, resumen


class InventarioBaseArticuloIn(BaseModel):
    id: Optional[int] = None
    codigo: str = ""
    nombre: str
    precio: float = 0
    existencias_base: float = 0
    hoja_origen: Optional[str] = None
    orden: Optional[int] = None


@app.get("/api/inventario/base")
def get_inventario_base():
    return _ensure_inventario_base_seeded()


@app.post("/api/inventario/base/import-default")
def import_inventario_base_default():
    if not os.path.exists(INVENTARIO_BASE_DEFAULT_PATH):
        raise HTTPException(status_code=404, detail=f"No se encontro el archivo base en {INVENTARIO_BASE_DEFAULT_PATH}")
    articulos = _load_inventario_base_from_excel(INVENTARIO_BASE_DEFAULT_PATH)
    if not articulos:
        raise HTTPException(status_code=400, detail="No se encontraron articulos en las hojas S y S2.")
    plan_db.replace_inventario_base(articulos, source_path=INVENTARIO_BASE_DEFAULT_PATH)
    return {"status": "success", "total_articulos": len(articulos), "source_path": INVENTARIO_BASE_DEFAULT_PATH}


@app.post("/api/inventario/base/articulo")
def save_inventario_base_articulo(payload: InventarioBaseArticuloIn):
    if not payload.nombre.strip():
        raise HTTPException(status_code=400, detail="El nombre del producto es obligatorio.")
    articulo_id = plan_db.upsert_inventario_base_articulo(payload.model_dump())
    return {"status": "success", "id": articulo_id}


@app.delete("/api/inventario/base/articulo/{articulo_id}")
def delete_inventario_base_articulo(articulo_id: int):
    plan_db.delete_inventario_base_articulo(articulo_id)
    return {"status": "success"}

@app.post("/api/inventario/upload")
async def upload_inventario(file: UploadFile = File(...)):
    """Sube un Excel de inventario, parsea artículos y guarda en DB."""
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls, .xlsm)")

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    content = await file.read()
    tmp.write(content)
    tmp.close()

    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active

        target_fields = {
            'nombre': ['nombre', 'articulo', 'producto', 'descripcion', 'item', 'descripción'],
            'precio': ['precio', 'costo', 'price', 'valor'],
            'codigo': ['codigo', 'código', 'code', 'cod', 'sku', 'ref', 'referencia'],
            'existencias': ['existencias', 'existencia', 'stock', 'cantidad', 'cant', 'inventario', 'qty'],
        }

        # Find a coherent header row inside the initial report area.
        # This avoids confusing report filters like "Inventario: Todos" with
        # the real table headers when the Excel includes a preamble.
        header_row, header_map = _detect_inventory_header_row(ws, target_fields)

        if not header_row or 'nombre' not in header_map.values():
            wb.close()
            raise HTTPException(status_code=400, detail="No se encontró encabezado con columna de nombre de artículo. Asegúrese de que el Excel tenga encabezados: Nombre, Precio, Código, Existencias.")

        # Parse articles
        articulos = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            art = {'codigo': '', 'nombre': '', 'precio': 0, 'existencias': 0}
            has_name = False
            for col_idx, field in header_map.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                if field == 'nombre':
                    if isinstance(val, str) and val.strip():
                        art['nombre'] = val.strip()
                        has_name = True
                elif field == 'codigo':
                    art['codigo'] = str(val).strip()
                elif field in ('precio', 'existencias'):
                    try:
                        art[field] = float(val)
                    except (ValueError, TypeError):
                        art[field] = 0
            if has_name:
                articulos.append(art)

        wb.close()

        if not articulos:
            raise HTTPException(status_code=400, detail="No se encontraron artículos en el Excel.")

        # Save to DB
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        carga_id = plan_db.guardar_carga_inventario(today, file.filename, articulos)

        return {
            "status": "success",
            "carga_id": carga_id,
            "total_articulos": len(articulos),
            "message": f"Se cargaron {len(articulos)} artículos exitosamente."
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar el Excel: {str(e)}")
    finally:
        os.unlink(tmp.name)

@app.get("/api/inventario/latest")
def get_inventario_latest():
    """Devuelve la carga más reciente con sus artículos."""
    data = plan_db.get_ultima_carga()
    if not data:
        return {"carga": None, "articulos": []}
    return data

@app.get("/api/inventario/diff")
def get_inventario_diff():
    """Calcula diferencias entre la ultima carga y la base maestra editable."""
    ultima = plan_db.get_ultima_carga()
    base = _ensure_inventario_base_seeded()
    if not ultima:
        return {"carga_actual": None, "base": base, "articulos": [], "resumen": {}}

    arts_out, resumen = _compare_inventory_against_base(ultima["articulos"], base["articulos"])

    return {
        "carga_actual": ultima["carga"],
        "base": base,
        "articulos": arts_out,
        "resumen": resumen
    }

@app.get("/api/inventario/history")
def get_inventario_history():
    """Historial de cargas de inventario."""
    return plan_db.get_historial_cargas(limit=30)

@app.delete("/api/inventario/{carga_id}")
def delete_inventario_carga(carga_id: int):
    """Elimina una carga de inventario."""
    plan_db.delete_carga_inventario(carga_id)
    return {"status": "success"}

# ==============================================================================
# Include Routers (for organized endpoint structure)
# ==============================================================================
app.include_router(empleados_router)
app.include_router(horarios_router)
app.include_router(planillas_router)
app.include_router(config_router)

# ==============================================================================
# Serve Frontend
# ==============================================================================
frontend_path = _frontend_dir
if os.path.exists(frontend_path):
    # Mount frontend directory for static files (React/Vue/JS/CSS)
    # Using 'html=True' lets it serve index.html for the root automatically.
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")
else:
    print(f"Warning: Frontend path {frontend_path} not found")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
