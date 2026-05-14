import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ══════════════════════════════════════════════════════════════════════════════
#  PALETA DE COLORES
# ══════════════════════════════════════════════════════════════════════════════
F_DEFAULT      = "Calibri"

C_DARK_BLUE    = "0F1923"   # Fondos título / gran total
C_TARJETA      = "1D4ED8"   # Sección PAGO POR TARJETA  (azul vibrante)
C_EFECTIVO     = "059669"   # Sección PAGO EN EFECTIVO  (verde esmeralda)
C_OCEAN        = "0369A1"   # Sub-encabezados días / horas
C_SLATE        = "243447"   # Sub-encabezados deducción
C_HDR_DARK     = "1B2838"   # Cabecera Empleado / Jornada
C_BRUTO_HDR    = "1E40AF"   # Cabecera Bruto
C_DED_HDR      = "991B1B"   # Cabeceras deducciones (Prestamo, etc.)
C_SEGURO_HDR   = "B45309"   # Cabecera Seguro
C_NETO_HDR     = "065F46"   # Cabecera NETO

C_EMP_TARJETA  = "DBEAFE"   # Celda nombre empleado — tarjeta
C_EMP_EFECT    = "D1FAE5"   # Celda nombre empleado — efectivo
C_INPUT        = "F0F9FF"   # Celdas de ingreso de horas
C_INPUT_ALT    = "E8F5E9"   # Celdas alternadas (horas mixtas/extra)
C_J_SUM        = "E0F2FE"   # Columna J totales horizontales
C_DED_CELL     = "FEF2F2"   # Celdas deducciones (M-P)
C_SEGURO_CELL  = "FFFBEB"   # Celda seguro
C_NETO_CELL    = "ECFDF5"   # Celda NETO
C_WHITE        = "FFFFFF"
C_TARIFA_BG    = "F5F3FF"   # Fila tarifas
C_TARIFA_LBL   = "4C1D95"   # Etiquetas tarifas
C_TARIFA_VAL   = "7C3AED"   # Valores tarifas
C_SUBTITLE     = "94A3B8"   # Texto secundario

# Tipos de pago / neto — fuentes de color
C_DED_FG       = "991B1B"
C_SEGURO_FG    = "92400E"
C_NETO_FG      = "065F46"
C_BONIF_HDR    = "7C3AED"   # Cabecera Bonificaciones (morado)
C_BONIF_CELL   = "F5F3FF"   # Celda bonificaciones (lavanda)
C_BONIF_FG     = "5B21B6"   # Fuente bonificaciones

# Hoja Resumen Mensual — colores de sección total
C_RED_HDR      = "B91C1C"

MONEY    = '"₡"#,##0.00'
HOURS_FMT = '0.00'
SUMMARY_MONEY_FMT = '"₡"#,##0'
SUMMARY_HOURS_FMT = '0'
CCSS_RATE = 0.1067          # Aporte CCSS empleado Costa Rica

# ══════════════════════════════════════════════════════════════════════════════
#  COMPAT MAIN.PY
# ══════════════════════════════════════════════════════════════════════════════
ROW_1      = PatternFill("solid", fgColor=C_WHITE)
ROW_2      = PatternFill("solid", fgColor="F7F9FC")
SAMPLE_EMPS = []
TARIFA_DIURNA   = 50.0
TARIFA_NOCTURNA = 75.0
TARIFA_MIXTA    = 62.5

# Aliases para compatibilidad con main.py (que referencia pl_module.f_data, etc.)
from openpyxl.styles import Side as _Side, Border as _Border
f_data = Font(name=F_DEFAULT, size=10, color="1F2937")
borde  = Border(
    top   =_Side(border_style="thin", color="CBD5E1"),
    left  =_Side(border_style="thin", color="CBD5E1"),
    right =_Side(border_style="thin", color="CBD5E1"),
    bottom=_Side(border_style="thin", color="CBD5E1"),
)

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS DE ESTILO
# ══════════════════════════════════════════════════════════════════════════════
al_c = Alignment(horizontal="center", vertical="center", wrap_text=False)
al_l = Alignment(horizontal="left",   vertical="center")
al_r = Alignment(horizontal="right",  vertical="center")

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _font(size=10, bold=False, color="1F2937", name=F_DEFAULT):
    return Font(name=name, size=size, bold=bold, color=color)

def _side(style="thin", color="CBD5E1"):
    return Side(border_style=style, color=color)

def _border(left="thin", right="thin", top="thin", bottom="thin",
            lc="CBD5E1", rc="CBD5E1", tc="CBD5E1", bc="CBD5E1"):
    return Border(
        left   = Side(border_style=left,   color=lc),
        right  = Side(border_style=right,  color=rc),
        top    = Side(border_style=top,    color=tc),
        bottom = Side(border_style=bottom, color=bc),
    )

# Borde para celdas de datos normales
B_DATA  = _border()
# Borde izquierdo grueso para columna L (separador visual horas | dinero)
B_THICK_LEFT = _border(left="medium", lc="334155")
# Borde para filas de encabezado de sección total
B_TOTAL = _border(top="medium", tc="334155", bottom="medium", bc="334155")
# Borde de total con separador izquierdo grueso
B_TOTAL_THICK_LEFT = _border(left="medium", lc="334155",
                             top="medium", tc="334155",
                             bottom="medium", bc="334155")

def sc(ws, row, col, val, font=None, fill=None, align=None, border=None, num_format=None):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    if font:       c.font       = font
    if fill:       c.fill       = fill
    if align:      c.alignment  = align
    if border:     c.border     = border
    if num_format: c.number_format = num_format

def _row_fill(ws, row, col_start, col_end, color):
    f = _fill(color)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = f

def _summary_ref_or_blank(ref):
    return f'=IF(OR({ref}="",{ref}=0),"",{ref})'

def _summary_expr_or_blank(expr):
    return f'=IF(({expr})=0,"",({expr}))'

# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════
def num_semana_anual(date_obj):
    if isinstance(date_obj, str):
        date_obj = datetime.strptime(date_obj, "%Y-%m-%d").date()
    return date_obj.isocalendar()[1]

def leer_catalogo(wb):
    if "Catalogo" not in wb.sheetnames:
        return {"tarjeta": [], "efectivo": [], "fijo": []}
    ws = wb["Catalogo"]
    res = {"tarjeta": [], "efectivo": [], "fijo": []}
    for r in range(5, ws.max_row + 1):
        nom = ws.cell(r, 1).value
        if not nom:
            continue
        tipo = str(ws.cell(r, 2).value or "").lower()
        # Coherente con tipo_map en main: "Transferencia Bancaria" no contiene "tarjeta"
        if "tarjeta" in tipo or "transferencia" in tipo:
            res["tarjeta"].append(nom)
        elif "fijo" in tipo:
            res["fijo"].append(nom)
        else:
            res["efectivo"].append(nom)
    return res

def contar_semanas(wb):
    return sum(1 for s in wb.sheetnames if s.startswith("Semana "))


def _seguro_por_empleado_map():
    """nombre (strip) -> True si aplica rebajo CCSS en planilla."""
    import database as db
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT nombre, COALESCE(aplica_seguro, 1) AS aplica_seguro FROM empleados"
    ).fetchall()
    conn.close()
    return {str(r["nombre"]).strip(): (int(r["aplica_seguro"]) != 0) for r in rows}


# ══════════════════════════════════════════════════════════════════════════════
#  HOJA SEMANAL — sección interna
# ══════════════════════════════════════════════════════════════════════════════
_DIAS_ES = ["Vie", "Sáb", "Dom", "Lun", "Mar", "Mié", "Jue"]

def _write_section(
    ws,
    start_row,
    label,
    sec_color,
    emp_color,
    emp_list,
    seguro_por_empleado=None,
    seguro_modo="porcentual",
    holiday_dates=None,
    viernes_date=None,
    horario_preview=None,
):
    """
    Escribe una sección (tarjeta / efectivo).
    Bloque por empleado: 3 filas base + 0–3 filas de extras (según horario_preview o las tres si no hay preview)
    + fila ★ Feriado solo si la semana tiene feriado en config.
    El pago de extras en el bruto se controla con la celda global $AA$1 (1=sí, 0=no).
    holiday_dates: list of ISO date strings ["YYYY-MM-DD", ...] for this week.
    horario_preview: opcional {nombre_empleado: {Vie: turno, ...}} para omitir filas de extra sin uso.
    Devuelve (next_free_row, total_row, anchor_rows_list).
    """
    import planilla_layout as pllay  # noqa: PLC0415 — evita import circular al cargar planilla

    COL_LAST = 20  # T
    # Check if any holiday falls within THIS specific week (viernes_date → viernes+6)
    has_holidays = False
    if holiday_dates and viernes_date:
        from datetime import timedelta as _td
        week_start = viernes_date
        week_end = viernes_date + _td(days=6)
        for h in holiday_dates:
            hd = h.get("date", "")
            if hd is None:
                continue
            if hasattr(hd, "isoformat"):
                hd = hd.isoformat()[:10]
            else:
                hd = str(hd).strip()[:10]
            if hd:
                try:
                    from datetime import datetime as _dt
                    h_date = _dt.strptime(hd, "%Y-%m-%d").date()
                    if week_start <= h_date <= week_end:
                        has_holidays = True
                        break
                except ValueError:
                    continue

    # ── Encabezado de sección ──────────────────────────────────────────────
    ws.row_dimensions[start_row].height = 28
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=COL_LAST)
    sc(ws, start_row, 1, f"  {label}",
       _font(12, True, C_WHITE), _fill(sec_color), al_l)
    r = start_row + 1

    # ── Sub-encabezado "Horas | Deducción" ────────────────────────────────
    ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    sc(ws, r, 1, None, fill=_fill(C_SLATE))
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    sc(ws, r, 3, "HORAS POR JORNADA",
       _font(8, True, C_WHITE), _fill(C_OCEAN), al_c)
    sc(ws, r, 11, None, fill=_fill(C_WHITE))
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=COL_LAST)
    sc(ws, r, 12, "BONIF. / DEDUCCIONES / RESUMEN",
       _font(8, True, C_WHITE), _fill(C_SLATE), al_c)
    r += 1

    # ── Cabeceras de columna con día + fecha en una sola fila ─────────────
    ws.row_dimensions[r].height = 32
    col_hdrs = [
        (1,  "Empleado",  C_HDR_DARK),
        (2,  "Jornada",   C_HDR_DARK),
        (10, "TOTAL",     C_OCEAN),
        (11, "",          C_WHITE),
        (12, "Bruto",     C_BRUTO_HDR),
        (13, "Bonific.",   C_BONIF_HDR),
        (14, "Préstamo",  C_DED_HDR),
        (15, "Combust.",  C_DED_HDR),
        (16, "Mercad.",   C_DED_HDR),
        (17, "Adelanto",  C_DED_HDR),
        (18, "Seguro",    C_SEGURO_HDR),
        (19, "Tot. Ded.", C_DED_HDR),
        (20, "NETO",      C_NETO_HDR),
    ]
    for col, txt, bg in col_hdrs:
        sc(ws, r, col, txt or None,
           _font(9, True, C_WHITE), _fill(bg), al_c,
           B_THICK_LEFT if col == 12 else B_DATA)

    # Columnas días: "Vie\n06/03"  (texto fijo del día + fórmula de fecha)
    # Como openpyxl no puede mezclar texto+fórmula en una celda, usamos
    # concatenación: ="Vie"&CHAR(10)&TEXT($C$2+0,"DD/MM")
    # Si el día es feriado, fondo dorado + ★
    for i, dia in enumerate(_DIAS_ES):
        col = 3 + i
        formula = f'="{dia}"&CHAR(10)&TEXT($C$2+{i},"DD/MM")'
        c = ws.cell(row=r, column=col)
        c.value         = formula
        c.font          = _font(9, True, C_WHITE)
        c.alignment     = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        c.border        = B_DATA
        # Marcar feriados en el header
        if holiday_dates and viernes_date:
            from datetime import timedelta as _td
            dia_date = viernes_date + _td(days=i)
            iso = dia_date.isoformat()
            matched = None
            for h in holiday_dates:
                hd = h.get("date", "")
                if hd is not None and hasattr(hd, "isoformat"):
                    hd = hd.isoformat()[:10]
                else:
                    hd = str(hd or "").strip()[:10]
                if hd == iso:
                    matched = h
                    break
            if matched:
                c.fill = _fill("D97706")  # ámbar oscuro para header
                nm = str(matched.get("name") or "").replace('"', '""').strip()
                if nm:
                    c.value = (
                        f'="★ {dia}"&CHAR(10)&TEXT($C$2+{i},"DD/MM")&CHAR(10)&"{nm}"'
                    )
                else:
                    c.value = f'="★ {dia}"&CHAR(10)&TEXT($C$2+{i},"DD/MM")'
            else:
                c.fill = _fill(C_OCEAN)
        else:
            c.fill = _fill(C_OCEAN)
    r += 1

    # ── Bloques de empleados ───────────────────────────────────────────────
    anchor_rows = []

    for emp in emp_list:
        anchor_rows.append(r)
        need = pllay.extras_set_for_employee(emp, horario_preview)
        rows_spec = [
            ("Hrs. Diurnas", C_INPUT),
            ("Hrs. Mixtas", C_INPUT_ALT),
            ("Hrs. Nocturnas", C_INPUT),
        ]
        row_by_extra = {}
        ri_next = r + 3
        for ek in pllay.EXTRA_KEYS_ORDER:
            if ek not in need:
                continue
            row_by_extra[ek] = ri_next
            bg = C_INPUT_ALT if (len(rows_spec) % 2 == 1) else C_INPUT
            rows_spec.append((pllay.EXTRA_LABELS[ek], bg))
            ri_next += 1

        n_rows = len(rows_spec)
        block_end = r + n_rows - 1
        holiday_row = (block_end + 1) if has_holidays else None
        emp_a_end = holiday_row if holiday_row is not None else block_end

        rd, rm, rn = r, r + 1, r + 2

        for idx in range(n_rows):
            ws.row_dimensions[r + idx].height = 20
        if holiday_row is not None:
            ws.row_dimensions[holiday_row].height = 20

        # Columna A — nombre (merge hasta última fila del bloque incl. feriado si aplica)
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=emp_a_end, end_column=1)
        sc(ws, r, 1, emp,
           _font(10, True), _fill(emp_color), al_c,
           _border(left="medium", lc="334155", right="thin", rc="CBD5E1",
                   top="medium", tc="334155", bottom="medium", bc="334155"))
        for ri in range(r + 1, emp_a_end + 1):
            ws.cell(row=ri, column=1).border = _border(
                left="medium", lc="334155", right="thin", rc="CBD5E1",
                top="thin", tc="CBD5E1", bottom="thin", bc="CBD5E1")

        for i, (conc, bg) in enumerate(rows_spec):
            ri = r + i
            font_c = _font(9, False)
            fill_c = _fill(bg)
            sc(ws, ri, 2, conc, font_c, fill_c, al_l, B_DATA)
            for ci in range(3, 10):
                c = ws.cell(row=ri, column=ci)
                c.value = None
                c.font = _font(10, False, "1E3A5F")
                c.fill = _fill(C_WHITE)
                c.alignment = al_c
                c.border = _border(lc="93C5FD", rc="93C5FD",
                                   tc="93C5FD", bc="93C5FD")
                c.number_format = HOURS_FMT
            ws.cell(row=ri, column=10).value = (
                f"=C{ri}+D{ri}+E{ri}+F{ri}+G{ri}+H{ri}+I{ri}"
            )
            ws.cell(row=ri, column=10).font = _font(10, True, "0F4C81")
            ws.cell(row=ri, column=10).fill = _fill(C_J_SUM)
            ws.cell(row=ri, column=10).alignment = al_c
            ws.cell(row=ri, column=10).border = B_DATA
            ws.cell(row=ri, column=10).number_format = HOURS_FMT
            ws.cell(row=ri, column=11).fill = _fill(C_WHITE)

        # ── Columnas de resumen / deducciones ─────────
        # Merge r → block_end (sin fila feriado). La fila feriado queda independiente.

        parts_br = [f"J{r}*$D$3", f"J{r+1}*$F$3", f"J{r+2}*$H$3"]
        for ek in pllay.EXTRA_KEYS_ORDER:
            if ek not in row_by_extra:
                continue
            rr = row_by_extra[ek]
            if ek == "ED":
                parts_br.append(f"IF($AA$1=1,J{rr}*$D$3*1.5,0)")
            elif ek == "EM":
                parts_br.append(f"IF($AA$1=1,J{rr}*$F$3*1.5,0)")
            else:
                parts_br.append(f"IF($AA$1=1,J{rr}*$H$3*1.5,0)")
        bruto_core = "=" + "+".join(parts_br)
        c_value = f"{bruto_core}+L{holiday_row}" if holiday_row is not None else bruto_core

        ws.merge_cells(start_row=r, start_column=12, end_row=block_end, end_column=12)
        c = ws.cell(row=r, column=12)
        c.value = c_value
        c.font          = _font(11, True, "1E3A5F")
        c.fill          = _fill(C_EMP_TARJETA if emp_color == C_EMP_TARJETA else C_EMP_EFECT)
        c.alignment     = al_c
        c.border        = _border(left="medium", lc="334155",
                                  right="thin",  rc="CBD5E1",
                                  top="medium",  tc="334155",
                                  bottom="medium", bc="334155")
        c.number_format = MONEY
        for ri in range(r+1, block_end+1):
            ws.cell(row=ri, column=12).border = _border(
                left="medium", lc="334155", right="thin", rc="CBD5E1",
                top="thin", tc="CBD5E1", bottom="thin", bc="CBD5E1")

        # M — Bonificaciones (ingresable)
        ws.merge_cells(start_row=r, start_column=13, end_row=block_end, end_column=13)
        c = ws.cell(row=r, column=13)
        c.value        = None
        c.font         = _font(10, True, C_BONIF_FG)
        c.fill         = _fill(C_BONIF_CELL)
        c.alignment    = al_c
        c.border       = B_DATA
        c.number_format = MONEY
        for ri in range(r+1, block_end+1):
            ws.cell(row=ri, column=13).border = B_DATA

        # N-Q — Deducciones ingresables (Préstamo, Combust., Mercad., Adelanto)
        for col in range(14, 18):
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=block_end, end_column=col)
            c = ws.cell(row=r, column=col)
            c.value        = None
            c.fill         = _fill(C_DED_CELL)
            c.alignment    = al_c
            c.border       = B_DATA
            c.number_format = MONEY
            for ri in range(r+1, block_end+1):
                ws.cell(row=ri, column=col).border = B_DATA

        # R — Seguro obrero (CCSS): fijo $O$3 o % solo sobre bruto L ($N$3); 0 si no aplica
        ws.merge_cells(start_row=r, start_column=18,
                       end_row=block_end, end_column=18)
        c = ws.cell(row=r, column=18)
        aps = (
            seguro_por_empleado.get(str(emp).strip(), True)
            if seguro_por_empleado
            else True
        )
        if not aps:
            c.value = 0
        elif seguro_modo == "fijo":
            c.value = f"=$O$3"
        else:
            c.value = f"=ROUND(L{r}*$N$3,2)"
        c.font         = _font(10, True, C_SEGURO_FG)
        c.fill         = _fill(C_SEGURO_CELL)
        c.alignment    = al_c
        c.border       = B_DATA
        c.number_format = MONEY
        for ri in range(r+1, block_end+1):
            ws.cell(row=ri, column=18).border = B_DATA

        # S — Total deducciones
        ws.merge_cells(start_row=r, start_column=19,
                       end_row=block_end, end_column=19)
        c = ws.cell(row=r, column=19)
        c.value        = f"=SUM(N{r}:R{r})"
        c.font         = _font(10, True, C_DED_FG)
        c.fill         = _fill(C_DED_CELL)
        c.alignment    = al_c
        c.border       = B_DATA
        c.number_format = MONEY
        for ri in range(r+1, block_end+1):
            ws.cell(row=ri, column=19).border = B_DATA

        # T — NETO  (Bruto + Bonificaciones - Tot.Ded.)
        ws.merge_cells(start_row=r, start_column=20,
                       end_row=block_end, end_column=20)
        c = ws.cell(row=r, column=20)
        c.value        = f"=L{r}+M{r}-S{r}"
        c.font         = _font(12, True, C_NETO_FG)
        c.fill         = _fill(C_NETO_CELL)
        c.alignment    = al_c
        c.border       = _border(left="thin", lc="CBD5E1",
                                 right="medium", rc="334155",
                                 top="medium",  tc="334155",
                                 bottom="medium", bc="334155")
        c.number_format = MONEY
        for ri in range(r+1, block_end+1):
            ws.cell(row=ri, column=20).border = _border(
                left="thin", lc="CBD5E1", right="medium", rc="334155",
                top="thin",  tc="CBD5E1", bottom="thin",  bc="CBD5E1")

        # ★ — FILA FERIADO: solo si la semana tiene al menos un feriado (config)
        if holiday_row is not None:
            sc(ws, holiday_row, 2, "★ Feriado", _font(9, True, "FFF59E0B"),
               _fill("FFFBEB"), al_l, B_DATA)
            for ci in range(3, 10):
                c = ws.cell(row=holiday_row, column=ci)
                c.value = None
                c.font = _font(10, False, "FFF59E0B")
                c.fill = _fill("FFFDE7")
                c.alignment = al_c
                c.border = _border(lc="FCD34D", rc="FCD34D", tc="FCD34D", bc="FCD34D")
                c.number_format = HOURS_FMT
            ws.cell(row=holiday_row, column=10).value = (
                f"=C{holiday_row}+D{holiday_row}+E{holiday_row}+F{holiday_row}+"
                f"G{holiday_row}+H{holiday_row}+I{holiday_row}"
            )
            ws.cell(row=holiday_row, column=10).font = _font(10, True, "FFF59E0B")
            ws.cell(row=holiday_row, column=10).fill = _fill("FEF3C7")
            ws.cell(row=holiday_row, column=10).alignment = al_c
            ws.cell(row=holiday_row, column=10).border = B_DATA
            ws.cell(row=holiday_row, column=10).number_format = HOURS_FMT
            ws.cell(row=holiday_row, column=11).fill = _fill(C_WHITE)
            c_l = ws.cell(row=holiday_row, column=12)
            c_l.value = pllay.feriado_monto_formula(holiday_row, rd, rm, rn, row_by_extra)
            c_l.font = _font(10, True, "FFF59E0B")
            c_l.fill = _fill("FEF3C7")
            c_l.alignment = al_c
            c_l.border = _border(left="medium", lc="334155", right="thin", rc="CBD5E1",
                                  top="thin", tc="CBD5E1", bottom="thin", bc="CBD5E1")
            c_l.number_format = MONEY
            for col in range(13, 21):
                ws.cell(row=holiday_row, column=col).value = None
                ws.cell(row=holiday_row, column=col).fill = _fill("FEF3C7")
                ws.cell(row=holiday_row, column=col).border = B_DATA

        r = (holiday_row + 1) if holiday_row is not None else (block_end + 1)

    # ── Fila TOTAL DE SECCIÓN ─────────────────────────────────────────────
    total_row = r
    ws.row_dimensions[r].height = 26
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    sc(ws, r, 1, f"TOTAL {label}",
       _font(10, True, C_WHITE), _fill(sec_color), al_r)

    def _sum_refs(col_idx):
        if not anchor_rows:
            return "0"
        return "+".join(f"{get_column_letter(col_idx)}{a}" for a in anchor_rows)

    for col in range(12, 21):
        c = ws.cell(row=r, column=col)
        c.value       = f"={_sum_refs(col)}" if anchor_rows else "=0"
        c.font        = _font(10, True, C_WHITE)
        c.fill        = _fill(sec_color)
        c.alignment   = al_c
        c.border      = B_TOTAL
        c.number_format = MONEY

    # T total = L + M - S
    ws.cell(row=r, column=20).value = f"=L{r}+M{r}-S{r}"

    r += 1
    return r, total_row, anchor_rows


# ══════════════════════════════════════════════════════════════════════════════
#  SECCIÓN SALARIO FIJO
#  Sin grilla de horas: sólo monto del período → bruto semanal, seguro CCSS y neto.
#  No aplica préstamo, combustible, mercadería ni adelantos (van al final del Excel).
# ══════════════════════════════════════════════════════════════════════════════
C_FIJO_HDR  = "6D28D9"   # Morado — encabezado sección fijo
C_FIJO_EMP  = "EDE9FE"   # Lavanda claro — celda nombre
C_FIJO_SAL  = "F5F3FF"   # Lavanda muy claro — celda salario

def _write_fijo_section(ws, start_row, fijo_list, wb_catalog, seguro_por_empleado=None, seguro_modo="porcentual"):
    """
    Sección SALARIO FIJO al final de la hoja semanal.
    Empleado | monto a pagar (período, BD) | bruto semanal (₡, calculado en Python) |
    columnas L–T alineadas al resto de la planilla: Bruto, bloque sin deducciones, seguro, neto.

    wb_catalog: respaldo de monto si falta en BD.
    Devuelve (next_row, total_row, anchor_rows).
    """
    if not fijo_list:
        return start_row, start_row, []

    COL_LAST = 20

    import database as db

    # Catálogo Excel (fallback de monto)
    salarios = {}
    if wb_catalog and "Catalogo" in wb_catalog.sheetnames:
        ws_cat = wb_catalog["Catalogo"]
        for r_cat in range(5, ws_cat.max_row + 1):
            nom = ws_cat.cell(r_cat, 1).value
            sal = ws_cat.cell(r_cat, 3).value
            if nom and isinstance(sal, (int, float)):
                salarios[nom] = sal

    conn = db.get_conn()
    rows_sf = conn.execute(
        "SELECT nombre, salario_fijo, "
        "COALESCE(NULLIF(TRIM(periodo_salario_fijo), ''), 'mensual') AS periodo "
        "FROM empleados WHERE tipo_pago='fijo'"
    ).fetchall()
    conn.close()
    fijo_by_name = {str(r["nombre"]).strip(): r for r in rows_sf}

    r = start_row

    # ── Encabezado de sección ──
    ws.row_dimensions[r].height = 28
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COL_LAST)
    sc(ws, r, 1, "  SALARIO FIJO",
       _font(12, True, C_WHITE), _fill(C_FIJO_HDR), al_l)
    r += 1

    # ── Sub-encabezado ──
    ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    sc(ws, r, 1, "Monto del período acordado y bruto semanal; sin otras deducciones — solo seguro CCSS",
       _font(8, True, C_WHITE), _fill(C_FIJO_HDR), al_c)
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=COL_LAST)
    sc(ws, r, 12, "Resumen planilla (misma estructura que transferencia / efectivo)",
       _font(8, True, C_WHITE), _fill(C_SLATE), al_c)
    r += 1

    # ── Cabeceras de columna ──
    ws.row_dimensions[r].height = 26
    hdr_row = r
    hdrs_fijo = [
        (1,  "Empleado",        C_HDR_DARK),
        (2,  "Monto a pagar",   C_FIJO_HDR),
        (3,  "Bruto semanal",   C_FIJO_HDR),
        (12, "Bruto",           C_BRUTO_HDR),
        (18, "Seguro",          C_SEGURO_HDR),
        (19, "Tot. ded.",       C_DED_HDR),
        (20, "NETO",            C_NETO_HDR),
    ]
    for col, txt, bg in hdrs_fijo:
        sc(ws, hdr_row, col, txt,
           _font(9, True, C_WHITE), _fill(bg), al_c,
           B_THICK_LEFT if col == 12 else B_DATA)
    ws.merge_cells(start_row=hdr_row, start_column=4, end_row=hdr_row, end_column=11)
    sc(ws, hdr_row, 4, "", _font(9, True, C_WHITE), _fill(C_FIJO_HDR), al_c, B_DATA)
    ws.merge_cells(start_row=hdr_row, start_column=13, end_row=hdr_row, end_column=17)
    sc(
        ws, hdr_row, 13,
        "Bonif. y otras deducciones\n(no aplica)",
        _font(8, True, C_WHITE), _fill(C_DED_HDR), al_c, B_DATA,
    )
    r += 1

    anchor_rows = []
    for emp in fijo_list:
        anchor_rows.append(r)
        ws.row_dimensions[r].height = 28

        rowdb = fijo_by_name.get(str(emp).strip()) or fijo_by_name.get(emp)
        if rowdb and rowdb["salario_fijo"] is not None:
            try:
                sal_periodo = float(rowdb["salario_fijo"])
            except (TypeError, ValueError):
                sal_periodo = float(salarios.get(emp, 0) or 0)
            pk = str(rowdb["periodo"] or "mensual").strip() or "mensual"
        else:
            sal_periodo = float(salarios.get(emp, 0) or 0)
            pk = "mensual"

        bruto_sem = db.salario_fijo_a_bruto_semanal(sal_periodo, pk)

        # A — Nombre
        sc(ws, r, 1, emp, _font(10, True), _fill(C_FIJO_EMP), al_l,
           _border(left="medium", lc="6D28D9", right="thin", rc="CBD5E1",
                   top="medium", tc="6D28D9", bottom="medium", bc="6D28D9"))

        # B — Monto a pagar del período (editable; sincronizar con RR.HH. si cambia el período)
        sc(ws, r, 2, sal_periodo, _font(10, False, "4C1D95"), _fill(C_FIJO_SAL),
           al_c, B_DATA, MONEY)

        # C — Bruto semanal (valor al generar la hoja; rég. semanal en BD)
        ws.cell(row=r, column=3).value        = bruto_sem
        ws.cell(row=r, column=3).font         = _font(10, True, "4C1D95")
        ws.cell(row=r, column=3).fill         = _fill(C_FIJO_SAL)
        ws.cell(row=r, column=3).alignment    = al_c
        ws.cell(row=r, column=3).border       = B_DATA
        ws.cell(row=r, column=3).number_format = MONEY

        # D-K — hueco alineado con grilla de horas (relleno homogéneo)
        for col in range(4, 12):
            ws.cell(row=r, column=col).fill   = _fill(C_FIJO_SAL)
            ws.cell(row=r, column=col).border = B_DATA

        # L — Bruto en columnario de planilla (= bruto semanal)
        c = ws.cell(row=r, column=12)
        c.value        = f"=C{r}"
        c.font         = _font(11, True, "1E3A5F")
        c.fill         = _fill(C_EMP_TARJETA)
        c.alignment    = al_c
        c.border       = _border(left="medium", lc="334155",
                                 right="thin",  rc="CBD5E1",
                                 top="medium",  tc="334155",
                                 bottom="medium", bc="334155")
        c.number_format = MONEY

        # M-Q — bloque sin deducciones (0 ₡ para totales y resúmenes; estilo neutro)
        ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=17)
        c_blk = ws.cell(row=r, column=13)
        c_blk.value        = 0
        c_blk.font         = _font(10, False, "64748B")
        c_blk.fill         = _fill("E2E8F0")
        c_blk.alignment    = al_c
        c_blk.border       = B_DATA
        c_blk.number_format = MONEY

        # R — Seguro: % solo sobre bruto L; fijo $O$3; 0 si no aplica
        aps = (
            seguro_por_empleado.get(str(emp).strip(), True)
            if seguro_por_empleado
            else True
        )
        c3 = ws.cell(row=r, column=18)
        if not aps:
            c3.value = 0
        elif seguro_modo == "fijo":
            c3.value = f"=$O$3"
        else:
            c3.value = f"=ROUND(L{r}*$N$3,2)"
        c3.font         = _font(10, True, C_SEGURO_FG)
        c3.fill         = _fill(C_SEGURO_CELL)
        c3.alignment    = al_c
        c3.border       = B_DATA
        c3.number_format = MONEY

        # S — Total deducciones = solo seguro en esta sección
        c4 = ws.cell(row=r, column=19)
        c4.value        = f"=R{r}"
        c4.font         = _font(10, True, C_DED_FG)
        c4.fill         = _fill(C_DED_CELL)
        c4.alignment    = al_c
        c4.border       = B_DATA
        c4.number_format = MONEY

        # T — NETO (sin bonif.; bloque M-Q es texto — no suma a L)
        c5 = ws.cell(row=r, column=20)
        c5.value        = f"=L{r}-S{r}"
        c5.font         = _font(12, True, C_NETO_FG)
        c5.fill         = _fill(C_NETO_CELL)
        c5.alignment    = al_c
        c5.border       = _border(left="thin", lc="CBD5E1",
                                  right="medium", rc="334155",
                                  top="medium",   tc="334155",
                                  bottom="medium", bc="334155")
        c5.number_format = MONEY

        r += 1

    # ── Fila TOTAL SALARIO FIJO ──
    total_row = r
    ws.row_dimensions[r].height = 24
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    sc(ws, r, 1, "TOTAL SALARIO FIJO",
       _font(10, True, C_WHITE), _fill(C_FIJO_HDR), al_r)

    def _sum_col(col_idx):
        if not anchor_rows:
            return "=0"
        cl = get_column_letter(col_idx)
        return "=" + "+".join(f"{cl}{a}" for a in anchor_rows)

    ws.cell(row=r, column=12).value = _sum_col(12)
    ws.cell(row=r, column=12).font = _font(10, True, C_WHITE)
    ws.cell(row=r, column=12).fill = _fill(C_FIJO_HDR)
    ws.cell(row=r, column=12).alignment = al_c
    ws.cell(row=r, column=12).border = B_TOTAL
    ws.cell(row=r, column=12).number_format = MONEY

    ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=17)
    c13t = ws.cell(row=r, column=13)
    c13t.value = 0
    c13t.font = _font(9, True, C_WHITE)
    c13t.fill = _fill(C_FIJO_HDR)
    c13t.alignment = al_c
    c13t.border = B_TOTAL
    c13t.number_format = MONEY

    for col in (18, 19):
        c = ws.cell(row=r, column=col)
        c.value = _sum_col(col)
        c.font = _font(10, True, C_WHITE)
        c.fill = _fill(C_FIJO_HDR)
        c.alignment = al_c
        c.border = B_TOTAL
        c.number_format = MONEY

    c20 = ws.cell(row=r, column=20)
    c20.value = f"=L{r}-S{r}"
    c20.font = _font(10, True, C_WHITE)
    c20.fill = _fill(C_FIJO_HDR)
    c20.alignment = al_c
    c20.border = B_TOTAL
    c20.number_format = MONEY

    r += 1
    return r, total_row, anchor_rows



def crear_hoja_semanal(
    wb,
    num,
    viernes_date,
    empleados,
    seguro=0,
    tarifas=None,
    holiday_dates=None,
    horario_preview=None,
):
    if isinstance(viernes_date, str):
        viernes_date = datetime.strptime(viernes_date, "%Y-%m-%d").date()

    tarifas = tarifas or {}
    holiday_dates = holiday_dates or []
    seguro_modo = str(tarifas.get("seguro_modo") or "porcentual").strip().lower()
    if seguro_modo not in ("porcentual", "fijo"):
        seguro_modo = "porcentual"
    sval = float(tarifas.get("seguro_valor", CCSS_RATE))
    if seguro_modo == "porcentual" and sval > 1.0:
        sval = sval / 100.0

    sem_num    = num_semana_anual(viernes_date)
    sheet_name = f"Semana {sem_num}"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    # Color de pestaña
    ws.sheet_properties.tabColor = "1D4ED8"

    COL_LAST = 20
    col_widths = {
        "A": 22, "B": 16, "C": 10, "D": 10, "E": 10,
        "F": 10, "G": 10, "H": 10, "I": 10, "J": 11,
        "K": 1,  "L": 15, "M": 12, "N": 12, "O": 11,
        "P": 11, "Q": 11, "R": 13, "S": 13, "T": 14,
        "AA": 2,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["AA"].hidden = True
    _pex = tarifas.get("pagar_horas_extra")
    if _pex is None:
        _pex = 1
    try:
        ws["AA1"] = 1 if int(_pex) != 0 else 0
    except (TypeError, ValueError):
        ws["AA1"] = 1

    # ── Fila 1: Título ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:T1")
    sc(ws, 1, 1, "PLANILLA DE PAGO SEMANAL",
       _font(16, True, C_WHITE), _fill(C_DARK_BLUE), al_c)

    # ── Fila 2: Período ──────────────────────────────────────────────────
    ws.row_dimensions[2].height = 26
    _row_fill(ws, 2, 1, COL_LAST, C_DARK_BLUE)
    sc(ws, 2, 1, "PERÍODO:", _font(10, True, C_WHITE), _fill(C_DARK_BLUE), al_r)
    sc(ws, 2, 3, viernes_date, _font(11, False, "BFDBFE"),
       _fill(C_DARK_BLUE), al_c, num_format="DD/MM/YYYY")
    sc(ws, 2, 4, "→", _font(11, False, C_SUBTITLE), _fill(C_DARK_BLUE), al_c)
    sc(ws, 2, 5, "=$C$2+6", _font(11, False, "BFDBFE"),
       _fill(C_DARK_BLUE), al_c, num_format="DD/MM/YYYY")
    sc(ws, 2, 7, f"Semana #{sem_num}",
       _font(11, False, C_SUBTITLE), _fill(C_DARK_BLUE), al_c)

    # ── Fila 3: Tarifas + Tasa CCSS ──────────────────────────────────────
    ws.row_dimensions[3].height = 24
    _row_fill(ws, 3, 1, COL_LAST, C_TARIFA_BG)
    ws.merge_cells("A3:B3")
    sc(ws, 3, 1, "TARIFAS (₡/hora):",
       _font(9, True, C_TARIFA_VAL), _fill(C_TARIFA_BG), al_r)
    tarifa_items = [
        (3, "Diurna:",   4,  TARIFA_DIURNA),
        (5, "Mixta:",    6,  TARIFA_MIXTA),
        (7, "Nocturna:", 8,  TARIFA_NOCTURNA),
    ]
    for lc, lbl, vc, val in tarifa_items:
        sc(ws, 3, lc, lbl, _font(8, False, C_TARIFA_LBL), _fill(C_TARIFA_BG), al_r)
        sc(ws, 3, vc, val, _font(10, True, C_TARIFA_VAL), _fill(C_WHITE), al_c)

    # N3 = tasa % aporte obrero (solo sobre bruto L); O3 = monto fijo semanal (solo uno aplica)
    if seguro_modo == "porcentual":
        sc(ws, 3, 13, "Seguro obrero (% CCSS):", _font(8, False, C_TARIFA_LBL), _fill(C_TARIFA_BG), al_r)
        sc(ws, 3, 14, sval, _font(10, True, C_TARIFA_VAL), _fill(C_WHITE), al_c, num_format="0.00%")
        sc(ws, 3, 15, 0, _font(10, True, C_TARIFA_VAL), _fill(C_WHITE), al_c, num_format=MONEY)
    else:
        sc(ws, 3, 13, "Seg. % (no usado):", _font(8, False, C_TARIFA_LBL), _fill(C_TARIFA_BG), al_r)
        sc(ws, 3, 14, 0, _font(10, True, C_TARIFA_VAL), _fill(C_WHITE), al_c, num_format="0.00%")
        sc(ws, 3, 15, sval, _font(10, True, C_TARIFA_VAL), _fill(C_WHITE), al_c, num_format=MONEY)

    # ── Fila 4: espaciador ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 5
    _row_fill(ws, 4, 1, COL_LAST, C_DARK_BLUE)

    # ── Secciones de pago ────────────────────────────────────────────────
    tarjeta_emps  = empleados.get("tarjeta",  [])
    efectivo_emps = empleados.get("efectivo", [])
    seguro_map = _seguro_por_empleado_map()

    r = 5
    r, tarjeta_total, _ = _write_section(
        ws,
        r,
        "TRANSFERENCIA BANCARIA",
        C_TARJETA,
        C_EMP_TARJETA,
        tarjeta_emps,
        seguro_map,
        seguro_modo,
        holiday_dates,
        viernes_date,
        horario_preview=horario_preview,
    )

    # separador visual entre secciones
    ws.row_dimensions[r].height = 4
    _row_fill(ws, r, 1, COL_LAST, "E2E8F0")
    r += 1

    r, efectivo_total, _ = _write_section(
        ws,
        r,
        "PAGO EN EFECTIVO",
        C_EFECTIVO,
        C_EMP_EFECT,
        efectivo_emps,
        seguro_map,
        seguro_modo,
        holiday_dates,
        viernes_date,
        horario_preview=horario_preview,
    )

    # ── Sección Salario Fijo (sólo si hay empleados fijos) ────────────────
    fijo_emps = empleados.get("fijo", [])
    fijo_total = None
    if fijo_emps:
        ws.row_dimensions[r].height = 4
        _row_fill(ws, r, 1, COL_LAST, "EDE9FE")
        r += 1
        r, fijo_total, _ = _write_fijo_section(ws, r, fijo_emps, wb, seguro_map, seguro_modo)

    # ── Gran Total ───────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 4
    _row_fill(ws, r, 1, COL_LAST, C_DARK_BLUE)

    gt = r + 1
    ws.row_dimensions[gt].height = 30
    ws.merge_cells(start_row=gt, start_column=1, end_row=gt, end_column=11)
    sc(ws, gt, 1, "GRAN TOTAL DE LA SEMANA",
       _font(12, True, C_WHITE), _fill(C_DARK_BLUE), al_r)
    gran_cols = {
        12: ("BFDBFE", True),
        13: ("E9D5FF", False),
        14: (C_WHITE, False), 15: (C_WHITE, False),
        16: (C_WHITE, False), 17: (C_WHITE, False),
        18: ("FEF3C7", False),
        19: ("FEE2E2", False),
        20: ("D1FAE5", True),
    }
    for col, (fc, bold) in gran_cols.items():
        cl = get_column_letter(col)
        # Suma las tres secciones si existe fijo, si no sólo las dos primeras
        if fijo_total:
            formula = f"={cl}{tarjeta_total}+{cl}{efectivo_total}+{cl}{fijo_total}"
        else:
            formula = f"={cl}{tarjeta_total}+{cl}{efectivo_total}"
        c = ws.cell(row=gt, column=col)
        c.value       = formula
        c.font        = _font(11 if col in (12, 20) else 10, bold, fc)
        c.fill        = _fill(C_DARK_BLUE)
        c.alignment   = al_c
        c.number_format = MONEY
    ws.cell(row=gt, column=20).value = f"=L{gt}+M{gt}-S{gt}"

    ws.row_dimensions[gt+1].height = 4
    _row_fill(ws, gt+1, 1, COL_LAST, C_DARK_BLUE)

    # ── Freeze panes: fija columnas A-B y filas de título/tarifas ────────
    # Con C5 fijamos: ver siempre columnas Empleado+Jornada al deslizar ↔
    # y el título/tarifas al deslizar ↕
    ws.freeze_panes = "C5"

    # ── Sub-hojas ─────────────────────────────────────────────────────────
    crear_resumen_semanal(wb, sheet_name, sem_num, viernes_date)
    crear_resumen_mensual(wb)
    crear_dashboard(wb)

    return sheet_name, gt + 2, {}


# ══════════════════════════════════════════════════════════════════════════════
#  RESUMEN SEMANAL
# ══════════════════════════════════════════════════════════════════════════════
def crear_resumen_semanal(wb, nombre_hoja_sem, sem_num, viernes_date):
    if isinstance(viernes_date, str):
        viernes_date = datetime.strptime(viernes_date, "%Y-%m-%d").date()

    import planilla_layout as pllay  # noqa: PLC0415

    sheet_name = f"Res. Sem. {sem_num}"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "0369A1"

    col_widths = {
        "A": 24, "B": 13, "C": 13, "D": 14, "E": 16, "F": 13,
        "G": 14, "H": 15, "I": 12, "J": 13, "K": 13, "L": 15,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:L1")
    sc(ws, 1, 1, f"RESUMEN SEMANAL — Semana {sem_num}",
       _font(14, True, C_WHITE), _fill(C_DARK_BLUE), al_c)

    fin = viernes_date + timedelta(days=6)
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:L2")
    sc(ws, 2, 1,
       f"{viernes_date.strftime('%d/%m/%Y')}  →  {fin.strftime('%d/%m/%Y')}",
       _font(10, False, C_SUBTITLE), _fill(C_DARK_BLUE), al_c)

    ws.row_dimensions[3].height = 5

    ws.row_dimensions[4].height = 22
    hdrs = [
        (1, "Empleado",      C_HDR_DARK),
        (2, "Hrs Diurnas",   C_OCEAN),
        (3, "Hrs Mixtas",    C_OCEAN),
        (4, "Hrs Nocturnas", C_OCEAN),
        (5, "Monto Extra (₡)", C_BRUTO_HDR),
        (6, "Hrs Extra",     C_OCEAN),
        (7, "Total Hrs",     C_OCEAN),
        (8, "Bruto (₡)",     C_BRUTO_HDR),
        (9, "Bonific. (₡)",  C_BONIF_HDR),
        (10, "Rebajos (₡)",  C_RED_HDR),
        (11, "Seguro (₡)",   C_SEGURO_HDR),
        (12, "NETO (₡)",     C_NETO_HDR),
    ]
    hour_cols = {2, 3, 4, 6, 7}
    money_cols = {5, 8, 9, 10, 11, 12}
    for col, txt, bg in hdrs:
        sc(ws, 4, col, txt, _font(9, True, C_WHITE), _fill(bg), al_c,
           B_THICK_LEFT if col == 8 else B_DATA)

    try:
        hs = wb[nombre_hoja_sem]
    except Exception:
        return

    def _write_sem_section(start_r, sec_label, sec_color):
        ws.row_dimensions[start_r].height = 24
        ws.merge_cells(start_row=start_r, start_column=1,
                       end_row=start_r, end_column=12)
        sc(ws, start_r, 1, f"  {sec_label}",
           _font(11, True, C_WHITE), _fill(sec_color), al_l)
        r = start_r + 1

        emp_rows = []
        in_section = False
        sn = nombre_hoja_sem
        for hr in range(1, hs.max_row + 1):
            v = hs.cell(row=hr, column=1).value
            # Encabezado de sección: texto en A incluye el nombre de la sección
            # (p. ej. "TRANSFERENCIA BANCARIA" no contiene la palabra "PAGO").
            if isinstance(v, str) and sec_label in v:
                in_section = True
                continue
            if in_section and isinstance(v, str) and v.startswith("TOTAL") and "GRAN" not in v:
                break
            if in_section and v and hs.cell(row=hr, column=2).value == "Hrs. Diurnas":
                emp_name = v
                scn = pllay.scan_jornada_block_rows(hs, hr)
                rd, rm, rn = scn["rd"], scn["rm"], scn["rn"]
                parts_m = []
                for key, tar in (("ed", "$D$3"), ("em", "$F$3"), ("en", "$H$3")):
                    rw = scn.get(key)
                    if rw:
                        parts_m.append(f"'{sn}'!J{rw}*'{sn}'!{tar}*1.5")
                if parts_m:
                    col5_inner = "+".join(parts_m)
                    col5_expr = f"IF('{sn}'!$AA$1=1,{col5_inner},0)"
                else:
                    col5_expr = "0"
                parts_h = []
                for key in ("ed", "em", "en"):
                    rw = scn.get(key)
                    if rw:
                        parts_h.append(f"'{sn}'!J{rw}")
                col6_expr = "+".join(parts_h) if parts_h else "0"
                sum_parts = [
                    f"'{sn}'!J{rd}",
                    f"'{sn}'!J{rm}",
                    f"'{sn}'!J{rn}",
                ]
                for key in ("ed", "em", "en"):
                    if scn.get(key):
                        sum_parts.append(f"'{sn}'!J{scn[key]}")
                if scn.get("fer"):
                    sum_parts.append(f"'{sn}'!J{scn['fer']}")
                sum_j_inner = ",".join(sum_parts)
                alt = (len(emp_rows) % 2 == 0)
                bg  = "F7F9FC" if alt else C_WHITE
                sc(ws, r, 1, emp_name, _font(10), _fill(bg), al_l, B_DATA)
                ws.cell(row=r, column=2).value = _summary_ref_or_blank(f"'{sn}'!J{rd}")
                ws.cell(row=r, column=3).value = _summary_ref_or_blank(f"'{sn}'!J{rm}")
                ws.cell(row=r, column=4).value = _summary_ref_or_blank(f"'{sn}'!J{rn}")
                ws.cell(row=r, column=5).value = _summary_expr_or_blank(col5_expr)
                ws.cell(row=r, column=6).value = _summary_expr_or_blank(col6_expr)
                ws.cell(row=r, column=7).value = _summary_expr_or_blank(f"SUM({sum_j_inner})")
                ws.cell(row=r, column=8).value = _summary_ref_or_blank(f"'{sn}'!L{rd}")
                ws.cell(row=r, column=9).value = _summary_ref_or_blank(f"'{sn}'!M{rd}")
                ws.cell(row=r, column=10).value = _summary_expr_or_blank(
                    f"'{sn}'!S{rd}-'{sn}'!R{rd}")
                ws.cell(row=r, column=11).value = _summary_ref_or_blank(f"'{sn}'!R{rd}")
                ws.cell(row=r, column=12).value = _summary_ref_or_blank(f"'{sn}'!T{rd}")
                for ci in range(1, 13):
                    ws.cell(row=r, column=ci).fill      = _fill(bg)
                    ws.cell(row=r, column=ci).alignment = al_c if ci > 1 else al_l
                    ws.cell(row=r, column=ci).border    = (
                        B_THICK_LEFT if ci == 8 else B_DATA)
                    if ci in hour_cols:
                        ws.cell(row=r, column=ci).number_format = SUMMARY_HOURS_FMT
                    if ci in money_cols:
                        ws.cell(row=r, column=ci).number_format = SUMMARY_MONEY_FMT
                emp_rows.append(r)
                r += 1

        # Fila total de sección
        tot = r
        ws.row_dimensions[r].height = 22
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
        sc(ws, r, 1, f"TOTAL {sec_label}",
           _font(10, True, C_WHITE), _fill(sec_color), al_r, B_TOTAL)
        if emp_rows:
            fr, lr = emp_rows[0], emp_rows[-1]
            for ci in range(2, 13):
                cl = get_column_letter(ci)
                ws.cell(row=r, column=ci).value       = _summary_expr_or_blank(
                    f"SUM({cl}{fr}:{cl}{lr})")
                ws.cell(row=r, column=ci).font        = _font(10, True, C_WHITE)
                ws.cell(row=r, column=ci).fill        = _fill(sec_color)
                ws.cell(row=r, column=ci).alignment   = al_c
                ws.cell(row=r, column=ci).border      = (
                    B_TOTAL_THICK_LEFT if ci == 8 else B_TOTAL)
                if ci in hour_cols:
                    ws.cell(row=r, column=ci).number_format = SUMMARY_HOURS_FMT
                if ci in money_cols:
                    ws.cell(row=r, column=ci).number_format = SUMMARY_MONEY_FMT
        return r + 1, tot

    r5 = 5
    r5, tot_tarjeta = _write_sem_section(r5, "TRANSFERENCIA BANCARIA", C_TARJETA)
    r5 += 1
    r5, tot_efectivo = _write_sem_section(r5, "PAGO EN EFECTIVO", C_EFECTIVO)

    # ── Sección Salario Fijo en el Resumen Semanal ──
    # Localizar filas de datos: (1) título "…SALARIO FIJO" en A, primera data en +3;
    # (2) respaldo por cabecera de grilla col B "Monto a pagar" (merge del título puede
    #     dejar A vacío en algunas lecturas). Cada empleado fijo tiene L = "=C{fila}".
    tot_fijo = None

    def _primera_fila_datos_salario_fijo(hoja):
        for rr in range(1, hoja.max_row + 1):
            v = hoja.cell(row=rr, column=1).value
            if isinstance(v, str):
                vs = v.strip().upper()
                if "SALARIO" in vs and "FIJO" in vs and not vs.startswith("TOTAL"):
                    return rr + 3
        for rr in range(1, hoja.max_row + 1):
            if str(hoja.cell(row=rr, column=2).value or "").strip() == "Monto a pagar":
                return rr + 1
        return None

    if hs:
        start_fr = _primera_fila_datos_salario_fijo(hs)
        fijo_filas_hoja = []
        if start_fr is not None:
            fr_scan = start_fr
            while fr_scan <= hs.max_row:
                v2 = hs.cell(row=fr_scan, column=1).value
                t2 = str(v2 or "").strip()
                if t2.upper().startswith("TOTAL SALARIO FIJO"):
                    break
                lval = hs.cell(row=fr_scan, column=12).value
                lnorm = str(lval or "").replace(" ", "").upper()
                es_fila_fijo = isinstance(lval, str) and f"=C{fr_scan}".upper() in lnorm
                if v2 and es_fila_fijo:
                    fijo_filas_hoja.append(fr_scan)
                fr_scan += 1

        if fijo_filas_hoja:
            emp_fijo_rows = []
            r5 += 1
            ws.row_dimensions[r5].height = 22
            ws.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=12)
            sc(ws, r5, 1, "  SALARIO FIJO",
               _font(11, True, C_WHITE), _fill(C_FIJO_HDR), al_l)
            r5 += 1
            sn = nombre_hoja_sem
            for fr2 in fijo_filas_hoja:
                v2 = hs.cell(row=fr2, column=1).value
                alt = (len(emp_fijo_rows) % 2 == 0)
                bg = "F5F3FF" if alt else "EDE9FE"
                sc(ws, r5, 1, v2, _font(10), _fill(bg), al_l, B_DATA)
                for ci in range(2, 8):
                    ws.cell(row=r5, column=ci).value = None
                    ws.cell(row=r5, column=ci).fill = _fill(bg)
                    ws.cell(row=r5, column=ci).alignment = al_c
                    ws.cell(row=r5, column=ci).border = B_DATA
                    if ci in hour_cols:
                        ws.cell(row=r5, column=ci).number_format = SUMMARY_HOURS_FMT
                    if ci in money_cols:
                        ws.cell(row=r5, column=ci).number_format = SUMMARY_MONEY_FMT
                ws.cell(row=r5, column=8).value = _summary_ref_or_blank(f"'{sn}'!L{fr2}")
                ws.cell(row=r5, column=9).value = _summary_ref_or_blank(f"'{sn}'!M{fr2}")
                ws.cell(row=r5, column=10).value = _summary_expr_or_blank(
                    f"'{sn}'!S{fr2}-'{sn}'!R{fr2}")
                ws.cell(row=r5, column=11).value = _summary_ref_or_blank(f"'{sn}'!R{fr2}")
                ws.cell(row=r5, column=12).value = _summary_ref_or_blank(f"'{sn}'!T{fr2}")
                for ci in range(8, 13):
                    ws.cell(row=r5, column=ci).fill = _fill(bg)
                    ws.cell(row=r5, column=ci).alignment = al_c
                    ws.cell(row=r5, column=ci).border = (
                        B_THICK_LEFT if ci == 8 else B_DATA)
                    ws.cell(row=r5, column=ci).number_format = SUMMARY_MONEY_FMT
                emp_fijo_rows.append(r5)
                r5 += 1

            tot_fijo = r5
            ws.row_dimensions[r5].height = 22
            ws.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=1)
            sc(ws, r5, 1, "TOTAL SALARIO FIJO",
               _font(10, True, C_WHITE), _fill(C_FIJO_HDR), al_r, B_TOTAL)
            frf, lrf = emp_fijo_rows[0], emp_fijo_rows[-1]
            for ci in range(2, 13):
                cl = get_column_letter(ci)
                ws.cell(row=r5, column=ci).value = _summary_expr_or_blank(
                    f"SUM({cl}{frf}:{cl}{lrf})")
                ws.cell(row=r5, column=ci).font = _font(10, True, C_WHITE)
                ws.cell(row=r5, column=ci).fill = _fill(C_FIJO_HDR)
                ws.cell(row=r5, column=ci).alignment = al_c
                ws.cell(row=r5, column=ci).border = (
                    B_TOTAL_THICK_LEFT if ci == 8 else B_TOTAL)
                if ci in hour_cols:
                    ws.cell(row=r5, column=ci).number_format = SUMMARY_HOURS_FMT
                if ci in money_cols:
                    ws.cell(row=r5, column=ci).number_format = SUMMARY_MONEY_FMT
            r5 += 1

    r5 += 1
    # Gran total semanal
    gt = r5
    ws.row_dimensions[gt].height = 26
    ws.merge_cells(start_row=gt, start_column=1, end_row=gt, end_column=1)
    sc(ws, gt, 1, "GRAN TOTAL SEMANAL",
       _font(11, True, C_WHITE), _fill(C_DARK_BLUE), al_r, B_TOTAL)
    for ci in range(2, 13):
        cl = get_column_letter(ci)
        refs = [f"{cl}{tot_tarjeta}", f"{cl}{tot_efectivo}"]
        if tot_fijo:
            refs.append(f"{cl}{tot_fijo}")
        ws.cell(row=gt, column=ci).value       = _summary_expr_or_blank(
            f"SUM({','.join(refs)})")
        ws.cell(row=gt, column=ci).font        = _font(10, True, C_WHITE)
        ws.cell(row=gt, column=ci).fill        = _fill(C_DARK_BLUE)
        ws.cell(row=gt, column=ci).alignment   = al_c
        ws.cell(row=gt, column=ci).border      = (
            B_TOTAL_THICK_LEFT if ci == 8 else B_TOTAL)
        if ci in hour_cols:
            ws.cell(row=gt, column=ci).number_format = SUMMARY_HOURS_FMT
        if ci in money_cols:
            ws.cell(row=gt, column=ci).number_format = SUMMARY_MONEY_FMT

    ws.freeze_panes = "B5"


# ══════════════════════════════════════════════════════════════════════════════
#  RESUMEN MENSUAL
# ══════════════════════════════════════════════════════════════════════════════
def crear_resumen_mensual(wb):
    sheet_name = "Resumen Mensual"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "374151"

    col_widths = {
        "A": 24, "B": 18, "C": 12, "D": 12, "E": 12, "F": 16,
        "G": 12, "H": 12, "I": 16, "J": 12, "K": 15, "L": 15, "M": 16,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:M1")
    sc(ws, 1, 1, "RESUMEN MENSUAL DE PLANILLA",
       _font(14, True, C_WHITE), _fill(C_DARK_BLUE), al_c)

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:M2")
    sc(ws, 2, 1, "Totales acumulados del mes por empleado",
       _font(9, False, C_SUBTITLE), _fill(C_DARK_BLUE), al_c)

    ws.row_dimensions[3].height = 5

    ws.row_dimensions[4].height = 22
    hdrs = [
        (1,  "Empleado",       C_HDR_DARK),
        (2,  "Semana",         C_HDR_DARK),
        (3,  "Hrs Diurnas",    C_OCEAN),
        (4,  "Hrs Mixtas",     C_OCEAN),
        (5,  "Hrs Nocturnas",  C_OCEAN),
        (6,  "Monto Extra (₡)", C_BRUTO_HDR),
        (7,  "Hrs Extra",      C_OCEAN),
        (8,  "Hrs Totales",    C_OCEAN),
        (9,  "Salario Bruto",  C_BRUTO_HDR),
        (10, "Bonific.",       C_BONIF_HDR),
        (11, "Rebajos",        C_RED_HDR),
        (12, "Seguro",         C_SEGURO_HDR),
        (13, "Neto a Pagar",   C_NETO_HDR),
    ]
    hour_cols = {3, 4, 5, 7, 8}
    money_cols = {6, 9, 10, 11, 12, 13}
    for col, txt, bg in hdrs:
        sc(ws, 4, col, txt, _font(9, True, C_WHITE), _fill(bg), al_c,
           B_THICK_LEFT if col == 9 else B_DATA)

    cur_r = 5
    empleados = leer_catalogo(wb)
    sections = [
        ("TRANSFERENCIA BANCARIA", empleados.get("tarjeta",  []), C_TARJETA,  C_EMP_TARJETA),
        ("PAGO EN EFECTIVO", empleados.get("efectivo", []), C_EFECTIVO, C_EMP_EFECT),
        ("SALARIO FIJO",     empleados.get("fijo",     []), C_FIJO_HDR, C_FIJO_EMP),
    ]
    sec_total_rows = []

    for sec_label, emp_list, sec_color, emp_bg in sections:
        ws.row_dimensions[cur_r].height = 24
        ws.merge_cells(start_row=cur_r, start_column=1,
                       end_row=cur_r, end_column=13)
        sc(ws, cur_r, 1, f"  {sec_label}",
           _font(11, True, C_WHITE), _fill(sec_color), al_l)
        cur_r += 1
        sec_emp_total_rows = []

        for emp in emp_list:
            emp_data_rows = []
            for sname in wb.sheetnames:
                if not sname.startswith("Res. Sem. "):
                    continue
                sws = wb[sname]
                for sr in range(4, sws.max_row + 1):
                    if sws.cell(row=sr, column=1).value == emp:
                        alt = len(emp_data_rows) % 2 == 0
                        bg  = "F7F9FC" if alt else C_WHITE
                        sc(ws, cur_r, 1, emp, _font(10), _fill(bg), al_l, B_DATA)
                        sc(ws, cur_r, 2,
                           sname.replace("Res. Sem. ", "Semana "),
                           _font(10), _fill(bg), al_c, B_DATA)
                        # Resumen semanal cols: 2=HrsD,3=HrsM,4=HrsN,5=MntExt,
                        # 6=HrsExt,7=TotHrs,8=Bruto,9=Bonif,10=Rebajos,11=Seguro,12=Neto
                        col_map = {
                            3: 2, 4: 3, 5: 4, 6: 5, 7: 6,
                            8: 7, 9: 8, 10: 9, 11: 10, 12: 11, 13: 12,
                        }
                        for dst, src in col_map.items():
                            c = ws.cell(row=cur_r, column=dst)
                            ref = f"'{sname}'!{get_column_letter(src)}{sr}"
                            c.value     = _summary_ref_or_blank(ref)
                            c.fill      = _fill(bg)
                            c.alignment = al_c
                            c.border    = B_THICK_LEFT if dst == 9 else B_DATA
                            if dst in hour_cols:
                                c.number_format = SUMMARY_HOURS_FMT
                            if dst in money_cols:
                                c.number_format = SUMMARY_MONEY_FMT
                        emp_data_rows.append(cur_r)
                        cur_r += 1

            if emp_data_rows:
                ws.row_dimensions[cur_r].height = 20
                ws.merge_cells(start_row=cur_r, start_column=1,
                               end_row=cur_r, end_column=2)
                sc(ws, cur_r, 1, "TOTAL MENSUAL",
                   _font(10, True), _fill(emp_bg), al_r, B_DATA)
                fr, lr = emp_data_rows[0], emp_data_rows[-1]
                for ci in range(3, 14):
                    cl = get_column_letter(ci)
                    fg = (C_SEGURO_FG if ci == 12 else
                          C_NETO_FG   if ci == 13 else
                          C_BONIF_FG  if ci == 10 else "1F2937")
                    bg2 = (C_SEGURO_CELL if ci == 12 else
                           C_NETO_CELL   if ci == 13 else
                           C_BONIF_CELL  if ci == 10 else emp_bg)
                    c = ws.cell(row=cur_r, column=ci)
                    c.value       = _summary_expr_or_blank(f"SUM({cl}{fr}:{cl}{lr})")
                    c.font        = _font(10, True, fg)
                    c.fill        = _fill(bg2)
                    c.alignment   = al_c
                    c.border      = B_THICK_LEFT if ci == 9 else B_DATA
                    if ci in hour_cols:
                        c.number_format = SUMMARY_HOURS_FMT
                    if ci in money_cols:
                        c.number_format = SUMMARY_MONEY_FMT
                sec_emp_total_rows.append(cur_r)
                cur_r += 1

        sec_tot = cur_r
        ws.row_dimensions[cur_r].height = 22
        ws.merge_cells(start_row=cur_r, start_column=1,
                       end_row=cur_r, end_column=2)
        sc(ws, cur_r, 1, f"TOTAL {sec_label}",
           _font(10, True, C_WHITE), _fill(sec_color), al_r, B_TOTAL)
        if sec_emp_total_rows:
            for ci in range(3, 14):
                cl = get_column_letter(ci)
                expr = f"SUM({','.join(f'{cl}{row}' for row in sec_emp_total_rows)})"
                c = ws.cell(row=cur_r, column=ci)
                c.value       = _summary_expr_or_blank(expr)
                c.font        = _font(10, True, C_WHITE)
                c.fill        = _fill(sec_color)
                c.alignment   = al_c
                c.border      = B_TOTAL_THICK_LEFT if ci == 9 else B_TOTAL
                if ci in hour_cols:
                    c.number_format = SUMMARY_HOURS_FMT
                if ci in money_cols:
                    c.number_format = SUMMARY_MONEY_FMT
        sec_total_rows.append(cur_r)
        cur_r += 1

    cur_r += 1
    ws.row_dimensions[cur_r].height = 28
    ws.merge_cells(start_row=cur_r, start_column=1,
                   end_row=cur_r, end_column=2)
    sc(ws, cur_r, 1, "GRAN TOTAL MENSUAL",
       _font(11, True, C_WHITE), _fill(C_DARK_BLUE), al_r, B_TOTAL)
    for ci in range(3, 14):
        cl = get_column_letter(ci)
        refs = ",".join(f"{cl}{tr}" for tr in sec_total_rows)
        c = ws.cell(row=cur_r, column=ci)
        c.value       = _summary_expr_or_blank(f"SUM({refs})")
        c.font        = _font(10, True, C_WHITE)
        c.fill        = _fill(C_DARK_BLUE)
        c.alignment   = al_c
        c.border      = B_TOTAL_THICK_LEFT if ci == 9 else B_TOTAL
        if ci in hour_cols:
            c.number_format = SUMMARY_HOURS_FMT
        if ci in money_cols:
            c.number_format = SUMMARY_MONEY_FMT

    ws.freeze_panes = "C5"


# ══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
def crear_dashboard(wb):
    sheet_name = "Dashboard"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "065F46"

    col_widths = {"A": 20, "B": 18, "C": 14, "D": 14,
                  "E": 14, "F": 24, "G": 18, "H": 14, "I": 14, "J": 14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:J1")
    sc(ws, 1, 1, "DASHBOARD — RESUMEN GRÁFICO DEL MES",
       _font(14, True, C_WHITE), _fill(C_DARK_BLUE), al_c)

    num_sem = contar_semanas(wb)
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:J2")
    sc(ws, 2, 1, f"Semanas procesadas: {num_sem}",
       _font(10, False, C_SUBTITLE), _fill(C_DARK_BLUE), al_c)

    ws.row_dimensions[3].height = 6

    # Panel "Datos por semana"
    ws.row_dimensions[4].height = 22
    ws.merge_cells("A4:D4")
    sc(ws, 4, 1, "DATOS POR SEMANA",
       _font(10, True, C_WHITE), _fill(C_SLATE), al_c)

    ws.row_dimensions[5].height = 20
    for col, txt, bg in [
        (1, "Semana", C_HDR_DARK),
        (2, "Bruto",  C_BRUTO_HDR),
        (3, "Neto",   C_NETO_HDR),
    ]:
        sc(ws, 5, col, txt, _font(9, True, C_WHITE), _fill(bg), al_c, B_DATA)

    row_idx = 6
    for sname in wb.sheetnames:
        if not sname.startswith("Res. Sem. "):
            continue
        sem_ws = wb[sname]
        gt_row = None
        for r in range(1, sem_ws.max_row + 1):
            if isinstance(sem_ws.cell(row=r, column=1).value, str) \
               and "GRAN TOTAL" in sem_ws.cell(row=r, column=1).value:
                gt_row = r
                break
        if gt_row is None:
            continue
        ws.row_dimensions[row_idx].height = 19
        alt = (row_idx % 2 == 0)
        sc(ws, row_idx, 1,
           sname.replace("Res. Sem. ", "Sem. "),
           _font(10), _fill("F7F9FC" if alt else C_WHITE), al_c, B_DATA)
        for ci, src_col, bg_cell, fc in [
            (2, "H", C_EMP_TARJETA, "1E3A5F"),
            (3, "L", C_NETO_CELL,   C_NETO_FG),
        ]:
            c = ws.cell(row=row_idx, column=ci)
            c.value        = f"='{sname}'!{src_col}{gt_row}"
            c.font         = _font(10, True, fc)
            c.fill         = _fill(bg_cell)
            c.alignment    = al_c
            c.border       = B_DATA
            c.number_format = MONEY
        row_idx += 1

    # Panel "Totales del mes"
    ws.row_dimensions[14].height = 22
    ws.merge_cells("A14:D14")
    sc(ws, 14, 1, "TOTALES DEL MES",
       _font(10, True, C_WHITE), _fill(C_SLATE), al_c)
    ws.merge_cells("F14:H14")
    sc(ws, 14, 6, "DISTRIBUCIÓN POR MÉTODO",
       _font(10, True, C_WHITE), _fill(C_SLATE), al_c)

    rm = wb.get("Resumen Mensual") if hasattr(wb, "get") else (
        wb["Resumen Mensual"] if "Resumen Mensual" in wb.sheetnames else None)
    gt_rm = None
    if rm:
        for r in range(1, rm.max_row + 1):
            v = rm.cell(row=r, column=1).value
            if isinstance(v, str) and "GRAN TOTAL" in v:
                gt_rm = r
                break

    totals_panel = [
        (15, "Total Bruto",  "F7F9FC",    9,  "1E3A5F", C_EMP_TARJETA),
        (16, "Tot. Bonific.","F5F3FF",     10, C_BONIF_FG, C_BONIF_CELL),
        (17, "Tot. Rebajos", C_WHITE,      11, C_DED_FG,  C_DED_CELL),
        (18, "Tot. Seguros", "F7F9FC",     12, C_SEGURO_FG, C_SEGURO_CELL),
        (19, "TOTAL NETO",   C_NETO_HDR,   13, C_WHITE,   C_NETO_CELL),
    ]
    for row_n, lbl, bg_lbl, src_ci, fc_val, bg_val in totals_panel:
        ws.row_dimensions[row_n].height = 22
        is_neto = lbl == "TOTAL NETO"
        sc(ws, row_n, 1, lbl,
           _font(10, is_neto, C_WHITE if is_neto else "1F2937"),
           _fill(bg_lbl if not is_neto else C_NETO_HDR), al_c, B_DATA)
        c = ws.cell(row=row_n, column=2)
        if gt_rm:
            c.value = f"='Resumen Mensual'!{get_column_letter(src_ci)}{gt_rm}"
        c.font         = _font(10, True, fc_val)
        c.fill         = _fill(bg_val)
        c.alignment    = al_c
        c.border       = B_DATA
        c.number_format = MONEY

    sc(ws, 15, 6, "Método",          _font(9, True, C_WHITE), _fill(C_HDR_DARK), al_c, B_DATA)
    sc(ws, 15, 7, "Total Neto",      _font(9, True, C_WHITE), _fill(C_NETO_HDR), al_c, B_DATA)
    sc(ws, 17, 6, "TRANSFERENCIA BANCARIA",_font(9),               _fill(C_EMP_TARJETA), al_c, B_DATA)
    sc(ws, 18, 6, "PAGO EN EFECTIVO",_font(9),               _fill(C_EMP_EFECT),   al_c, B_DATA)

    if gt_rm:
        # Tarjeta total neto
        c17 = ws.cell(row=17, column=7)
        c17.value       = f"='Resumen Mensual'!M{gt_rm}"
        c17.fill        = _fill(C_EMP_TARJETA)
        c17.alignment   = al_c
        c17.number_format = MONEY
        # Efectivo total neto
        c18 = ws.cell(row=18, column=7)
        c18.value       = f"='Resumen Mensual'!M{gt_rm}"
        c18.fill        = _fill(C_EMP_EFECT)
        c18.alignment   = al_c
        c18.number_format = MONEY


# ══════════════════════════════════════════════════════════════════════════════
#  CATÁLOGO
# ══════════════════════════════════════════════════════════════════════════════
def crear_catalogo(wb):
    sheet_name = "Catalogo"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name, 0)
    ws.sheet_properties.tabColor = "1D4ED8"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:D1")
    sc(ws, 1, 1, "CATÁLOGO DE EMPLEADOS",
       _font(14, True, C_WHITE), _fill(C_DARK_BLUE), al_c)

    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:D2")
    sc(ws, 2, 1,
       "Edite esta lista y luego use: python planilla.py agregar --viernes YYYY-MM-DD",
       _font(9, False, C_SUBTITLE), _fill(C_DARK_BLUE), al_c)

    ws.row_dimensions[4].height = 22
    hdrs = ["Nombre Completo", "Tipo de Pago", "Salario Fijo (₡)"]
    for i, h in enumerate(hdrs):
        sc(ws, 4, i+1, h, _font(9, True, C_WHITE), _fill(C_HDR_DARK), al_c, B_DATA)
