import os
import openpyxl
from datetime import date, datetime, timedelta
from PIL import Image, ImageDraw, ImageFont

try:
    from openpyxl.utils.datetime import from_excel as _xl_from_excel
except ImportError:
    _xl_from_excel = None

# ── Layout Constants ─────────────────────────────────────────────────────────
WIDTH      = 800
ROW_H      = 36        # standard row height
BIG_ROW_H  = 56        # salario neto / total a depositar

BG_COLOR   = "#FFFFFF"
TEXT_COLOR = "#000000"
BLUE_COLOR = "#1a5db4"
GRAY_DARK  = "#d9d9d9"
GRAY_LIGHT = "#f4f6f8"
YELLOW_BG  = "#FFEB3B"   # amarillo legible para sección CCSS
RED_COLOR  = "#C62828"   # rojo más intenso para montos/horas/valor hora

# Column X positions for the 4-column table: Concepto | Horas | Valor/Hora | Monto
# Each number is the X of the RIGHT edge (= left edge of next column)
COL_HRS_X   = 390   # end of Concepto / start of Horas
COL_VAL_X   = 510   # end of Horas   / start of Valor/hora
COL_MNT_X   = 680   # end of Valor   / start of Monto
# Monto goes from COL_MNT_X to WIDTH


def cargar_fuente(size, bold=False):
    try:
        if os.name == 'nt':
            fname = "arialbd.ttf" if bold else "arial.ttf"
            return ImageFont.truetype(fname, size)
    except Exception:
        pass
    return ImageFont.load_default()


def fmt_money(val):
    """Formats a value as Costa Rican colones. Uses abs() to avoid negatives."""
    try:
        f = abs(float(val))
    except (TypeError, ValueError):
        return ""
    if f == 0:
        return ""
    return f"C {f:,.2f}"  # Use 'C' prefix (plain) to avoid glyph issues


def _cell_to_date(v):
    """Convierte celda Excel a date (inicio de semana planilla = viernes)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and _xl_from_excel is not None:
        try:
            dt = _xl_from_excel(v)
            return dt.date() if hasattr(dt, "date") else dt
        except Exception:
            pass
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _fmt_fecha_d(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v) if v else ""


def periodo_y_fecha_pago_desde_hoja(ws):
    """
    Semana de planilla: viernes (inicio) → jueves (fin), 7 días.
    Fecha de pago: el día siguiente al último día del periodo (p. ej. viernes si el periodo cierra en jueves).
    Devuelve (texto_periodo, fecha_pago_str) o (None, None) si no hay fechas.
    """
    c2 = _cell_to_date(ws.cell(2, 3).value)
    e2 = _cell_to_date(ws.cell(2, 5).value)

    # C2 = viernes inicio; fin de semana = jueves (inicio + 6 días)
    if c2:
        inicio = c2
        fin = c2 + timedelta(days=6)
    elif e2:
        fin = e2
        inicio = e2 - timedelta(days=6)
    else:
        return None, None

    fecha_pago = fin + timedelta(days=1)
    t_per = f"Del {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}"
    t_fp = fecha_pago.strftime("%d/%m/%Y")
    return t_per, t_fp


def _tiene_hrs_extra(datos):
    try:
        return float(datos.get("HrsExtra") or 0) != 0.0
    except (TypeError, ValueError):
        return False


def _empleado_aplica_seguro(datos):
    v = datos.get("aplica_seguro")
    if v is None:
        return True
    try:
        return int(v) != 0
    except (TypeError, ValueError):
        return bool(v)


def _ccss_efectivo_boleta(datos):
    """Monto CCSS a mostrar/descontar en boleta (0 si no aplica seguro al empleado)."""
    if not _empleado_aplica_seguro(datos):
        return 0.0
    try:
        return float(datos.get("ReduccionCCSS") or 0)
    except (TypeError, ValueError):
        return 0.0


def _enrich_seguro_si_formula_vacia(emp_data, tarifas_cfg):
    """
    Con data_only=True las fórmulas pueden venir en 0. Recalcula según tarifas en BD:
    porcentual sobre bruto (sin bonificaciones) o monto fijo por semana.
    """
    if not _empleado_aplica_seguro(emp_data):
        return
    try:
        seg = float(emp_data.get("ReduccionCCSS") or 0)
        bruto_l = float(emp_data.get("TotalSalario") or 0)
        bon = float(emp_data.get("Bonificacion") or 0)
    except (TypeError, ValueError):
        return
        
    modo = str(tarifas_cfg.get("seguro_modo") or "porcentual").strip().lower()
    if modo not in ("porcentual", "fijo"):
        modo = "porcentual"
    emp_data["seguro_modo"] = modo

    if seg > 0:
        return
        
    base_para_seguro = bruto_l
    if base_para_seguro <= 0:
        return
        
    val = float(tarifas_cfg.get("seguro_valor", 0.1067))
    if modo == "porcentual" and val > 1.0:
        val = val / 100.0
        
    if modo == "fijo":
        emp_data["ReduccionCCSS"] = round(val, 2)
    else:
        emp_data["ReduccionCCSS"] = round(base_para_seguro * val, 2)
    try:
        prest = float(emp_data.get("Prestamos") or 0)
        comb = float(emp_data.get("Combustible") or 0)
        merc = float(emp_data.get("Mercaderia") or 0)
        adel = float(emp_data.get("Adelantos") or 0)
        seg_n = float(emp_data["ReduccionCCSS"])
        tot_ded = prest + comb + merc + adel + seg_n
        emp_data["SalarioSemanal"] = round(bruto_l + bon - tot_ded, 2)
    except (TypeError, ValueError):
        pass


def fmt_hrs(val):
    """Formats hours without trailing decimals, or '' if zero."""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return ""
    if f == 0:
        return ""
    # Remove trailing .0
    if f == int(f):
        return str(int(f))
    return f"{f:.1f}".rstrip('0').rstrip('.')


# ── Low-level drawing helpers ─────────────────────────────────────────────────

def _dividers(d, y, h):
    """Draw the THREE inner vertical dividers for the 4-column table."""
    d.line([(COL_HRS_X, y), (COL_HRS_X, y + h)], fill=TEXT_COLOR, width=1)
    d.line([(COL_VAL_X, y), (COL_VAL_X, y + h)], fill=TEXT_COLOR, width=1)
    d.line([(COL_MNT_X, y), (COL_MNT_X, y + h)], fill=TEXT_COLOR, width=1)


def _row_box(d, y, h, fill=BG_COLOR):
    """Draw a full-width row rectangle with a thin black border."""
    d.rectangle([(0, y), (WIDTH - 1, y + h)], fill=fill, outline=TEXT_COLOR, width=1)


def draw_table_header(d, y, f_bold):
    """Draw the gray 4-column header row."""
    _row_box(d, y, ROW_H, fill=GRAY_DARK)
    _dividers(d, y, ROW_H)
    cy = y + ROW_H // 2
    # Center each label inside its column
    d.text((COL_HRS_X // 2,                     cy), "Concepto",   font=f_bold, fill=TEXT_COLOR, anchor="mm")
    d.text(((COL_HRS_X + COL_VAL_X) // 2,       cy), "Horas",      font=f_bold, fill=TEXT_COLOR, anchor="mm")
    d.text(((COL_VAL_X + COL_MNT_X) // 2,       cy), "Valor/hora", font=f_bold, fill=TEXT_COLOR, anchor="mm")
    d.text(((COL_MNT_X + WIDTH)      // 2,       cy), "Monto",      font=f_bold, fill=TEXT_COLOR, anchor="mm")
    return y + ROW_H


def draw_hora_row(d, y, concepto, hrs, valor_hora, monto, f_norm, f_bold, f_values=None):
    """Always draw a row; show '-' if hrs is zero (per user request).
    Horas, valor/hora y monto en rojo y negrita (más llamativo) cuando aplica."""
    f_values = f_values or f_bold
    _row_box(d, y, ROW_H)
    _dividers(d, y, ROW_H)
    cy = y + ROW_H // 2
    hrs_txt   = fmt_hrs(hrs)       or "-"
    val_txt   = fmt_money(valor_hora) or "-"
    monto_txt = fmt_money(monto)   or "-"

    def _strike(txt):
        return txt != "-"

    hrs_color = RED_COLOR if _strike(hrs_txt) else TEXT_COLOR
    val_color = RED_COLOR if _strike(val_txt) else TEXT_COLOR
    mnt_color = RED_COLOR if _strike(monto_txt) else TEXT_COLOR
    fh = f_values if _strike(hrs_txt) else f_norm
    fv = f_values if _strike(val_txt) else f_norm
    fm = f_values if _strike(monto_txt) else f_norm

    # Each value drawn AFTER boxes, centered or aligned within its column
    d.text((COL_HRS_X - 8, cy),                   concepto,  font=f_bold, fill=TEXT_COLOR, anchor="rm")  # right-align in Concepto col
    d.text(((COL_HRS_X + COL_VAL_X) // 2,  cy),   hrs_txt,   font=fh, fill=hrs_color, anchor="mm")  # centered in Horas col
    d.text(((COL_VAL_X + COL_MNT_X) // 2,  cy),   val_txt,   font=fv, fill=val_color, anchor="mm")  # centered in Valor/hora col
    d.text((WIDTH - 8,                      cy),   monto_txt, font=fm, fill=mnt_color, anchor="rm")  # right-align in Monto col
    return y + ROW_H


def draw_summary_row(d, y, label, monto, f_bold, f_big=None, fill=BG_COLOR, label_color=TEXT_COLOR, monto_color=TEXT_COLOR, h=None):
    """Draw a summary row with label right-aligned before amount, amount right-aligned."""
    row_h = h or ROW_H
    _row_box(d, y, row_h, fill=fill)
    # Single vertical divider before Monto column
    d.line([(COL_MNT_X, y), (COL_MNT_X, y + row_h)], fill=TEXT_COLOR, width=1)
    cy = y + row_h // 2
    lf = f_big if f_big else f_bold
    mf = f_big if f_big else f_bold
    d.text((COL_MNT_X - 8, cy), label,                 font=lf, fill=label_color, anchor="rm")
    d.text((WIDTH - 8,     cy), fmt_money(monto) or "-", font=mf, fill=monto_color, anchor="rm")
    return y + row_h


def draw_section_title(d, y, text, f_sub, fill=GRAY_DARK, h=None):
    row_h = h or ROW_H
    _row_box(d, y, row_h, fill=fill)
    cy = y + row_h // 2
    d.text((WIDTH // 2, cy), text, font=f_sub, fill=TEXT_COLOR, anchor="mm")
    return y + row_h


# ── Main boleta renderer ──────────────────────────────────────────────────────

def generar_boleta_jpeg(
    datos,
    ruta_salida,
    sem_nombre,
    periodo_linea,
    fecha_pago_str,
    logo_path,
    tipo_bono="otros",
    valor_sticker=150.0,
):
    HEIGHT = 1100
    img = Image.new('RGB', (WIDTH, HEIGHT), color=BG_COLOR)
    d   = ImageDraw.Draw(img)

    f_norm  = cargar_fuente(14)
    f_bold  = cargar_fuente(14, bold=True)
    f_values = cargar_fuente(16, bold=True)  # horas / valor hora / monto — más visible
    f_note  = cargar_fuente(11)  # notas aclaratorias
    f_sub   = cargar_fuente(16, bold=True)
    f_big   = cargar_fuente(20, bold=True)
    f_huge  = cargar_fuente(26, bold=True)
    f_title = cargar_fuente(22, bold=True)

    # ── 1. Blue header bar ────────────────────────────────────────────────────
    HEADER_H = 95
    d.rectangle([(0, 0), (WIDTH, HEADER_H)], fill=BLUE_COLOR)

    # Wide landscape logo - SLM logo is 736x201, so we give it ~230x63 in header
    LOGO_W, LOGO_H_px = 230, 63
    LOGO_X, LOGO_Y = 8, (HEADER_H - LOGO_H_px) // 2
    if os.path.exists(logo_path):
        try:
            logo_img = Image.open(logo_path).convert("RGBA")
            wbg = Image.new("RGBA", logo_img.size, "WHITE")
            wbg.paste(logo_img, (0, 0), logo_img)
            logo_rgb = wbg.convert("RGB")
            logo_rgb.thumbnail((LOGO_W, LOGO_H_px))
            ox = LOGO_X + (LOGO_W - logo_rgb.width)  // 2
            oy = LOGO_Y + (LOGO_H_px - logo_rgb.height) // 2
            img.paste(logo_rgb, (ox, oy))
        except Exception as e:
            print("Logo error:", e)

    y = HEADER_H

    # ── 2. Info row: ROCO S.A. | SEMANA X ────────────────────────────────────
    INFO_H = 75
    d.rectangle([(0, y), (WIDTH, y + INFO_H)], fill=BG_COLOR, outline=TEXT_COLOR, width=2)
    # Vertical divider
    DIV_X = 560
    d.line([(DIV_X, y), (DIV_X, y + INFO_H)], fill=TEXT_COLOR, width=2)
    # Left: company legal name (phone already in header blue bar)
    d.text((10, y + 16), "ROCO S.A.",                      font=f_sub,  fill=TEXT_COLOR)
    d.text((10, y + 44), "CÉD. JURÍDICA: 3-101-130178", font=f_norm, fill=TEXT_COLOR)
    # Right: week label
    d.text(((DIV_X + WIDTH) // 2, y + INFO_H // 2),
           sem_nombre.upper(), font=f_big, fill=TEXT_COLOR, anchor="mm")
    y += INFO_H

    # ── 3. "COMPROBANTE DE PAGO" title ────────────────────────────────────────
    y = draw_section_title(d, y, "COMPROBANTE DE PAGO", f_sub, fill=GRAY_LIGHT)

    # ── 4. Period row ─────────────────────────────────────────────────────────
    PER_H = 50
    d.rectangle([(0, y), (WIDTH, y + PER_H)], fill=BG_COLOR, outline=TEXT_COLOR, width=1)
    d.line([(WIDTH // 2, y), (WIDTH // 2, y + PER_H)], fill=TEXT_COLOR, width=1)

    cx_l = WIDTH // 4
    cx_r = 3 * WIDTH // 4

    d.text((cx_l, y + 13), "Per\u00edodo de pago", font=f_norm, fill=TEXT_COLOR, anchor="mm")
    d.text((cx_l, y + 35), periodo_linea, font=f_norm, fill=TEXT_COLOR, anchor="mm")
    d.text((cx_r, y + 13), "Fecha de pago", font=f_norm, fill=TEXT_COLOR, anchor="mm")
    d.text((cx_r, y + 35), fecha_pago_str or "\u2014", font=f_norm, fill=TEXT_COLOR, anchor="mm")
    y += PER_H

    # ── 5. Collaborator rows ──────────────────────────────────────────────────
    _row_box(d, y, ROW_H, fill=GRAY_DARK)
    d.text((8, y + ROW_H // 2), f"COLABORADOR: {datos['nombre']}", font=f_sub, fill=TEXT_COLOR, anchor="lm")
    y += ROW_H

    _row_box(d, y, ROW_H, fill=GRAY_LIGHT)
    cedula_txt = datos.get("cedula") or "-"
    d.text((8, y + ROW_H // 2), f"C\u00c9DULA: {cedula_txt}", font=f_bold, fill=TEXT_COLOR, anchor="lm")
    y += ROW_H

    # ── 6. Hours table ────────────────────────────────────────────────────────
    y = draw_table_header(d, y, f_bold)
    y = draw_hora_row(d, y, "Hrs laboradas Diurnas",   datos["HrsDiurnas"],   datos["ValorHoraDiurna"],   datos["MontoDiurna"],   f_norm, f_bold, f_values)
    y = draw_hora_row(d, y, "Hrs laboradas Mixtas",    datos["HrsMixta"],     datos["ValorHoraMixta"],     datos["MontoMixta"],    f_norm, f_bold, f_values)
    y = draw_hora_row(d, y, "Hrs laboradas Nocturnas", datos["HrsNocturnas"], datos["ValorHoraNocturna"], datos["MontoNocturna"],  f_norm, f_bold, f_values)
    if _tiene_hrs_extra(datos):
        y = draw_hora_row(d, y, "Hrs Extra", datos["HrsExtra"], datos["ValorHoraExtra"], datos["MontoExtra"], f_norm, f_bold, f_values)
    
    bonif = datos.get("Bonificacion", 0)
    if bonif > 0:
        if tipo_bono == "stickers":
            cant = int(bonif / valor_sticker) if valor_sticker > 0 else 0
            y = draw_hora_row(d, y, "Stickers", cant, valor_sticker, bonif, f_norm, f_bold, f_values)
        else:
            y = draw_hora_row(d, y, "Bonificación", "", "", bonif, f_norm, f_bold, f_values)

    # El Salario Bruto en el boleta será la suma de horas + bonificaciones
    datos["TotalSalarioBoleta"] = datos["TotalSalario"] + bonif

    # ── 7. Salario Bruto ──────────────────────────────────────────────────────
    y = draw_summary_row(d, y, "SALARIO BRUTO", datos["TotalSalarioBoleta"], f_bold, fill=BG_COLOR)

    # ── 8. Rebajo CCSS — solo si el empleado aplica seguro y hay monto ────────
    ccss_monto = _ccss_efectivo_boleta(datos)
    if ccss_monto > 0:
        modo_seguro = datos.get("seguro_modo", "porcentual")
        pct_txt = ""
        
        if modo_seguro == "porcentual":
            base_seguro = float(datos.get("TotalSalario") or 0)
            if base_seguro > 0:
                pct = round(100.0 * ccss_monto / base_seguro, 2)
                pct_txt = f" ({pct:g}%)"

        _row_box(d, y, ROW_H, fill=YELLOW_BG)
        d.line([(COL_VAL_X, y), (COL_VAL_X, y + ROW_H)], fill=TEXT_COLOR, width=1)
        cy = y + ROW_H // 2
        d.text((COL_VAL_X - 8, cy), f"Monto rebajado por CCSS{pct_txt}",
               font=f_bold, fill=TEXT_COLOR, anchor="rm")
        d.text((WIDTH - 8, cy), fmt_money(ccss_monto), font=f_values, fill=RED_COLOR, anchor="rm")
        y += ROW_H

    # ── 9. Salario Neto (grand, tall row) ─────────────────────────────────────
    neto_intermedio = datos["TotalSalarioBoleta"] - ccss_monto
    _row_box(d, y, BIG_ROW_H, fill=BG_COLOR)
    d.line([(COL_VAL_X, y), (COL_VAL_X, y + BIG_ROW_H)], fill=TEXT_COLOR, width=1)
    cy = y + BIG_ROW_H // 2
    d.text((COL_VAL_X - 8, cy), "SALARIO NETO",             font=f_big,  fill=TEXT_COLOR, anchor="rm")
    d.text((WIDTH - 8,     cy), fmt_money(neto_intermedio),  font=f_huge, fill=TEXT_COLOR, anchor="rm")
    y += BIG_ROW_H

    # ── 10. Deducciones (if any) ──────────────────────────────────────────────
    ded_items = [
        ("Mercader\u00eda de Cr\u00e9dito", datos["Mercaderia"]),
        ("Combustible de Cr\u00e9dito",    datos["Combustible"]),
        ("Adelantos",                       datos["Adelantos"]),
        ("Pr\u00e9stamos",                datos["Prestamos"]),
    ]
    ded_items = [(lbl, val) for lbl, val in ded_items if val > 0]

    if ded_items:
        y = draw_section_title(d, y, "DEDUCCIONES (OTRAS)", f_bold, fill=GRAY_DARK)
        for lbl, val in ded_items:
            _row_box(d, y, ROW_H)
            d.line([(COL_VAL_X, y), (COL_VAL_X, y + ROW_H)], fill=TEXT_COLOR, width=1)
            cy = y + ROW_H // 2
            d.text((COL_VAL_X - 8, cy), lbl, font=f_bold, fill=TEXT_COLOR, anchor="rm")
            d.text((WIDTH - 8, cy), fmt_money(val), font=f_bold, fill=RED_COLOR, anchor="rm")
            y += ROW_H

    # ── 11. Total a Depositar ─────────────────────────────────────────────────
    _row_box(d, y, BIG_ROW_H, fill=GRAY_LIGHT)
    d.line([(COL_VAL_X, y), (COL_VAL_X, y + BIG_ROW_H)], fill=TEXT_COLOR, width=1)
    cy = y + BIG_ROW_H // 2
    d.text((COL_VAL_X - 8, cy), "TOTAL A DEPOSITAR",           font=f_big,  fill=TEXT_COLOR, anchor="rm")
    d.text((WIDTH - 8,     cy), fmt_money(datos["SalarioSemanal"]), font=f_huge, fill=TEXT_COLOR, anchor="rm")
    y += BIG_ROW_H

    # ── 12. Forma de Pago ─────────────────────────────────────────────────────
    _row_box(d, y, ROW_H, fill=BG_COLOR)
    d.text((WIDTH // 2, y + ROW_H // 2),
           f"Forma de pago: {datos['forma_pago'].upper()}", font=f_bold, fill=TEXT_COLOR, anchor="mm")
    y += ROW_H

    # ── 13. Signatures ────────────────────────────────────────────────────────
    y += 30
    d.text((30,        y), "FIRMA: _______________________", font=f_bold, fill=TEXT_COLOR)
    d.text((WIDTH - 30, y), "FECHA: _______________________", font=f_bold, fill=TEXT_COLOR, anchor="rt")
    y += 40

    # Crop and border
    final_h = y + 10
    img = img.crop((0, 0, WIDTH, final_h))
    ImageDraw.Draw(img).rectangle([(0, 0), (WIDTH - 1, final_h - 1)], outline=TEXT_COLOR, width=2)
    img.save(ruta_salida, "JPEG", quality=95)


# ── Excel data extraction ─────────────────────────────────────────────────────

def extract_employee_data(ws, row_idx, headers):
    def get_val(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    def get_horas(r):
        tot = get_val(r, 10)
        # Fallback to sum of days C(3) through I(9) if formula un-evaluated
        if tot == 0.0:
            tot = sum(get_val(r, c) for c in range(3, 10))
        return tot

    h_diurnas = get_horas(row_idx)
    h_mixtas  = get_horas(row_idx + 1)
    h_noct    = get_horas(row_idx + 2)
    h_extra   = get_horas(row_idx + 3)

    t_diurna = float(headers.get("Tarifa Diurna", 0))
    t_noct   = float(headers.get("Tarifa Noct", 0))
    t_mixta  = float(headers.get("Tarifa Mixta", 0))
    t_extra  = float(headers.get("Tarifa Extra", 0))

    m_diurna = h_diurnas * t_diurna
    m_noct   = h_noct * t_noct
    m_mixta  = h_mixtas * t_mixta
    m_extra  = h_extra * t_extra

    bruto = get_val(row_idx, 12)
    # If un-evaluated formula
    if bruto == 0.0:
        bruto = m_diurna + m_noct + m_mixta + m_extra

    bonif   = get_val(row_idx, 13)
    prest   = get_val(row_idx, 14)
    combust = get_val(row_idx, 15)
    mercad  = get_val(row_idx, 16)
    adelant = get_val(row_idx, 17)
    seguro  = get_val(row_idx, 18)
    
    tot_ded = get_val(row_idx, 19)
    if tot_ded == 0.0:
        tot_ded = prest + combust + mercad + adelant + seguro
        
    neto = get_val(row_idx, 20)
    if neto == 0.0:
        neto = (bruto + bonif) - tot_ded

    return {
        "HrsDiurnas":      h_diurnas,
        "ValorHoraDiurna": t_diurna,
        "MontoDiurna":     m_diurna,
        "HrsNocturnas":    h_noct,
        "ValorHoraNocturna": t_noct,
        "MontoNocturna":   m_noct,
        "HrsMixta":        h_mixtas,
        "ValorHoraMixta":  t_mixta,
        "MontoMixta":      m_mixta,
        "HrsExtra":        h_extra,
        "ValorHoraExtra":  t_extra,
        "MontoExtra":      m_extra,
        "TotalSalario":    bruto,
        "ReduccionCCSS":   seguro,
        "Mercaderia":      mercad,
        "Combustible":     combust,
        "Adelantos":       adelant,
        "Prestamos":       prest,
        "Bonificacion":    bonif,
        "SalarioSemanal":  neto,
    }


# ── Public entry point ────────────────────────────────────────────────────────

def generar_boletas_semana(archivo_excel, nombre_semana, logo_path, tipo_bono="otros", valor_sticker=150.0):
    """
    Lee la semana especificada del archivo Excel y genera boletas JPG
    en la carpeta BOLETAS - <nombre_semana> junto al archivo.
    Returns (ok: bool, message: str, out_dir: str).
    """
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    except Exception as e:
        return False, f"No se pudo cargar el Excel: {e}", ""

    if nombre_semana not in wb.sheetnames:
        return False, f"Hoja '{nombre_semana}' no encontrada.", ""

    ws = wb[nombre_semana]

    periodo_line, fecha_pago_str = periodo_y_fecha_pago_desde_hoja(ws)
    if not periodo_line:
        fi = _cell_to_date(ws.cell(2, 3).value)
        ff = _cell_to_date(ws.cell(2, 5).value)
        if fi and ff:
            periodo_line = f"Del {fi.strftime('%d/%m/%Y')} al {ff.strftime('%d/%m/%Y')}"
            fecha_pago_str = (ff + timedelta(days=1)).strftime("%d/%m/%Y")
        elif fi:
            ff = fi + timedelta(days=6)
            periodo_line = f"Del {fi.strftime('%d/%m/%Y')} al {ff.strftime('%d/%m/%Y')}"
            fecha_pago_str = (ff + timedelta(days=1)).strftime("%d/%m/%Y")
        else:
            p1 = _fmt_fecha_d(ws.cell(2, 3).value)
            p2 = _fmt_fecha_d(ws.cell(2, 5).value)
            periodo_line = f"Del {p1} al {p2}" if p1 and p2 else (p1 or p2 or "Per\u00edodo")
            fecha_pago_str = ""

    def get_tarifa(r, c):
        v = ws.cell(r, c).value
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    t_diurna = get_tarifa(3, 4)
    t_mixta  = get_tarifa(3, 6)
    t_noct   = get_tarifa(3, 8)
    
    t_extra_raw = ws.cell(3, 10).value
    try:
        t_extra = float(t_extra_raw)
    except (ValueError, TypeError):
        # Fallback for =$D$3*1.5
        t_extra = t_diurna * 1.5

    tarifas = {
        "Tarifa Diurna": t_diurna,
        "Tarifa Mixta":  t_mixta,
        "Tarifa Noct":   t_noct,
        "Tarifa Extra":  t_extra,
    }

    # Tarifas (modo seguro) y empleados desde BD
    tarifas_cfg = {}
    cedulas = {}
    seguro_por_nombre = {}
    try:
        import database as db
        tarifas_cfg = db.get_tarifas()
        for emp in db.get_empleados(solo_activos=False):
            key = emp["nombre"].strip().upper()
            if emp.get("cedula"):
                cedulas[key] = emp["cedula"]
            seguro_por_nombre[key] = emp.get("aplica_seguro", 1)
    except Exception:
        pass

    out_dir = os.path.join(os.path.dirname(os.path.abspath(archivo_excel)),
                           f"BOLETAS - {nombre_semana.upper()}")
    os.makedirs(out_dir, exist_ok=True)

    metodo_actual = "Transferencia Bancaria"
    last_val      = None
    count         = 0

    for row in range(5, ws.max_row + 1):
        val = ws.cell(row, 1).value

        # Detect payment method section headers
        if val and isinstance(val, str) and "PAGO" in val.upper():
            if "TARJETA" in val.upper() or "TRANSFERENCIA" in val.upper(): metodo_actual = "Transferencia Bancaria"
            elif "EFECTIVO" in val.upper(): metodo_actual = "Efectivo"
            elif "FIJO"     in val.upper(): metodo_actual = "Salario Fijo"
            continue

        if val and isinstance(val, str) and "TOTAL" not in val.upper() and val != last_val:
            last_val = val
            try:
                emp_data = extract_employee_data(ws, row, tarifas)
                emp_data["nombre"]     = val
                emp_data["forma_pago"] = metodo_actual
                k = val.strip().upper()
                emp_data["cedula"] = cedulas.get(k, "-")
                emp_data["aplica_seguro"] = seguro_por_nombre.get(k, 1)
                _enrich_seguro_si_formula_vacia(emp_data, tarifas_cfg)

                if emp_data["TotalSalario"] > 0 or emp_data["SalarioSemanal"] > 0 or emp_data.get("Bonificacion", 0) > 0:
                    ruta = os.path.join(out_dir, f"{val.replace(' ', '_')}.jpg")
                    generar_boleta_jpeg(
                        emp_data,
                        ruta,
                        nombre_semana,
                        periodo_line,
                        fecha_pago_str,
                        logo_path,
                        tipo_bono,
                        valor_sticker,
                    )
                    count += 1
            except Exception as ex:
                print(f"Error procesando {val}: {ex}")

    return True, f"Se generaron {count} boletas JPG en:\n{out_dir}", out_dir
