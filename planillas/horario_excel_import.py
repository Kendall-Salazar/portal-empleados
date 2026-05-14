"""
Parseo de workbooks tipo HORARIO 2026.xlsx (hojas con fila título, fechas, Colaborador + días).
Variantes: split_friday (Vie en col J) y linear_vie_jue (Vie..Jue en B–H).
"""
from __future__ import annotations

import datetime
import re
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

import horario_db

DAYS_ORDER = ("Vie", "Sáb", "Dom", "Lun", "Mar", "Mié", "Jue")

# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZACIÓN DE TAREAS DE LIMPIEZA
# Mapeo de palabras clave que aparecen en la sección "LIMPIEZA" del Excel
# a los nombres canónicos que el sistema usa en daily_tasks.
# ─────────────────────────────────────────────────────────────────────────────
_TASK_KEYWORD_MAP = {
    # Baños / Bathrooms
    "bano": "Baños",
    "baños": "Baños",
    "banos": "Baños",
    # Tanques / Tanks measurement
    "tanque": "Tanques",
    "tanques": "Tanques",
    "medir": "Tanques",
    # Oficina + Basureros + Baños
    "oficina": "Oficina + Basureros + Baños",
    "basurero": "Oficina + Basureros + Baños",
    "basureros": "Oficina + Basureros + Baños",
}

# Caracteres de flecha que el Excel puede contener (varios encodings)
_ARROW_UP_CHARS = {"↑", "\u2191", "▲", "\u25b2"}
_ARROW_DOWN_CHARS = {"↓", "\u2193", "▼", "\u25bc"}


def _normalize_task_label(raw: str) -> Optional[str]:
    """
    Recibe el texto de la columna A de la sección de limpieza y lo mapea
    a un nombre canónico. Retorna None si no reconoce ninguna tarea.
    """
    if not raw:
        return None
    cleaned = unicodedata.normalize("NFKD", raw.strip().lower())
    cleaned = "".join(c for c in cleaned if not unicodedata.combining(c))
    cleaned = re.sub(r"\s+", " ", cleaned)
    for keyword, canonical in _TASK_KEYWORD_MAP.items():
        if keyword in cleaned:
            return canonical
    return None


def _classify_shift_time(shift_code: str) -> Optional[str]:
    """
    Retorna 'AM' si el turno es principalmente de mañana (start < 12),
    'PM' si es de tarde, o None si no se puede determinar (OFF, etc.).
    """
    hours = horario_db.SHIFTS.get(shift_code)
    if not hours:
        # intenta con normalize_manual_shift_code
        norm = horario_db.normalize_manual_shift_code(shift_code)
        hours = horario_db.get_shift_hours_set(norm) if norm else set()
    if not hours:
        return None
    start = min(h % 24 for h in hours)
    return "AM" if start < 12 else "PM"


def _build_task_label(base_task: str, arrow_char: Optional[str], shift_code: Optional[str]) -> str:
    """
    Construye la etiqueta final de la tarea con el sufijo ↑AM / ↓PM si aplica.

    Lógica de precedencia:
    1. Si hay flecha explícita en el Excel → usar esa dirección.
    2. Si no hay flecha → inferir desde el código de turno del empleado asignado.
    3. Si no podemos determinar ninguna → retornar la tarea sin sufijo.
    """
    suffix = None
    if arrow_char and arrow_char in _ARROW_UP_CHARS:
        suffix = " ↑AM"
    elif arrow_char and arrow_char in _ARROW_DOWN_CHARS:
        suffix = " ↓PM"
    elif shift_code:
        time_class = _classify_shift_time(shift_code)
        if time_class == "AM":
            suffix = " ↑AM"
        elif time_class == "PM":
            suffix = " ↓PM"
    return f"{base_task}{suffix}" if suffix else base_task


# --- format_shift_code (mismo criterio que backend/main.py para inversa estable) ---


def _format_time_range(time_range: str) -> str:
    try:
        start_s, end_s = time_range.split("-")
        start_h = int(start_s)
        end_h = int(end_s)
        start_dt = datetime.time(start_h % 24, 0)
        end_dt = datetime.time(end_h % 24, 0)
        start_str = start_dt.strftime("%I:%M %p")
        end_str = end_dt.strftime("%I:%M %p")
        return f"{start_str} - {end_str}"
    except (ValueError, TypeError, AttributeError):
        return time_range


def format_shift_code_local(code: str) -> str:
    normalized = horario_db.normalize_manual_shift_code(code) or code
    if not normalized or normalized == "OFF":
        return "LIBRE"
    if normalized == "VAC":
        return "VACACIONES"
    if normalized == "PERM":
        return "PERMISO"
    if "+" in normalized:
        parts = normalized.split("_")
        if len(parts) > 1:
            times = parts[1].split("+")
            readable_times = [_format_time_range(t) for t in times]
            return " / ".join(readable_times)
    parts = normalized.split("_")
    if len(parts) > 1:
        return _format_time_range(parts[1])
    return normalized


def normalize_readable_key(text: str) -> str:
    if text is None:
        return ""
    s = str(text).strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def build_inverse_shift_map() -> Tuple[Dict[str, str], List[str]]:
    """texto_normalizado -> código; lista de advertencias por colisiones."""
    inv: Dict[str, str] = {}
    collisions: List[str] = []
    for code in sorted(horario_db.SHIFTS.keys()):
        if code in ("OFF", "VAC", "PERM"):
            continue
        readable = format_shift_code_local(code)
        key = normalize_readable_key(readable)
        if not key:
            continue
        if key in inv and inv[key] != code:
            collisions.append(f"{key}: {inv[key]} vs {code}")
        elif key not in inv:
            inv[key] = code
    inv[normalize_readable_key("LIBRE")] = "OFF"
    inv[normalize_readable_key("VACACIONES")] = "VAC"
    inv[normalize_readable_key("PERMISO")] = "PERM"
    return inv, collisions


_INVERSE_MAP_CACHE: Optional[Dict[str, str]] = None


def get_inverse_shift_map() -> Dict[str, str]:
    global _INVERSE_MAP_CACHE
    if _INVERSE_MAP_CACHE is None:
        _INVERSE_MAP_CACHE, _ = build_inverse_shift_map()
    return _INVERSE_MAP_CACHE


def readable_cell_to_code(text: Any, warnings: List[str]) -> str:
    if text is None or (isinstance(text, str) and not str(text).strip()):
        return "OFF"
    raw = str(text).strip()
    key = normalize_readable_key(raw)
    if key == "LIBRE" or key == "":
        return "OFF"
    if key == "VACACIONES":
        return "VAC"
    if key == "PERMISO":
        return "PERM"
    inv = get_inverse_shift_map()
    if key in inv:
        return inv[key]
    # Intento MANUAL_: quitar " / " por "+" para normalize_manual_shift_code estilo rangos
    manual_candidate = raw.replace(" / ", "+").replace("/", "-")
    norm = horario_db.normalize_manual_shift_code(manual_candidate)
    if norm and norm.startswith(horario_db.MANUAL_SHIFT_PREFIX):
        return norm
    # ─── NUEVO: Turno Excepcional / Libre texto ───────────────────────────────
    # Si el texto no coincide con ningún código conocido pero parece un horario
    # (contiene dígitos y am/pm o rangos), lo salvamos como MANUAL_ textual para
    # no perder información. Esto cubre casos como "5am-11pm + 5pm + 11pm".
    if re.search(r"\d", raw) and re.search(r"(am|pm|-)", raw, re.I):
        # Intentar normalización flexible: reemplazar separadores comunes
        flexible_candidate = (
            raw
            .replace(" + ", "+")
            .replace(", ", "+")
            .replace(" - ", "-")
        )
        norm2 = horario_db.normalize_manual_shift_code(flexible_candidate)
        if norm2 and norm2.startswith(horario_db.MANUAL_SHIFT_PREFIX):
            return norm2
    warnings.append(f"Sin código para texto: {raw[:80]}")
    return "OFF"


def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def parse_day_header(cell_value: Any) -> Optional[str]:
    if cell_value is None:
        return None
    s = _strip_accents(str(cell_value).strip().lower())
    if not s:
        return None
    if s.startswith("vie"):
        return "Vie"
    if s.startswith("sab") or s.startswith("sáb"):
        return "Sáb"
    if s.startswith("dom"):
        return "Dom"
    if s.startswith("lun"):
        return "Lun"
    if s.startswith("mar"):
        return "Mar"
    if s.startswith("mie") or s.startswith("mié"):
        return "Mié"
    if s.startswith("jue"):
        return "Jue"
    return None


def _is_colaborador_a1(cell_a: Any) -> bool:
    if cell_a is None:
        return False
    t = _strip_accents(str(cell_a).strip().lower())
    return t.startswith("colaborador")


def _is_noise_header(cell_value: Any) -> bool:
    if cell_value is None:
        return True
    s = _strip_accents(str(cell_value).strip().lower())
    if not s:
        return True
    if s.startswith("formato"):
        return True
    if s.startswith("horas") and len(s) < 8:
        return True
    if s.startswith("colaborador"):
        return True
    return False


def _is_cleaning_section_header(cell_value: Any) -> bool:
    """Detecta si una celda es la cabecera de la sección de LIMPIEZA/OBLIGACIONES."""
    if cell_value is None:
        return False
    s = _strip_accents(str(cell_value).strip().lower())
    return any(tok in s for tok in ("limpieza", "obligaciones"))


def find_header_and_column_map(ws) -> Tuple[Optional[int], Dict[int, str], List[str]]:
    """
    Retorna (header_row_1based, {col: día}, errors).
    """
    errors: List[str] = []
    max_scan = min(ws.max_row or 0, 40)
    for r in range(1, max_scan + 1):
        a = ws.cell(row=r, column=1).value
        if not _is_colaborador_a1(a):
            continue
        col_map: Dict[int, str] = {}
        max_c = min(ws.max_column or 15, 20)
        for c in range(2, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if _is_noise_header(v):
                continue
            day = parse_day_header(v)
            if day:
                col_map[c] = day
        days_found = set(col_map.values())
        if len(col_map) != len(days_found):
            continue
        if days_found == set(DAYS_ORDER) and len(col_map) == len(DAYS_ORDER):
            return r, col_map, []
    errors.append("No se encontró fila con Colaborador en A y los 7 días (Vie…Jue) en columnas distintas")
    return None, {}, errors


def parse_date_cell(val: Any) -> Optional[datetime.date]:
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def build_week_dates_from_row(ws, date_row: int, col_map: Dict[int, str]) -> Tuple[Dict[str, str], List[str], List[str]]:
    """week_dates DD/MM/YYYY, errors críticos, warnings."""
    errors: List[str] = []
    warnings: List[str] = []
    wd: Dict[str, str] = {}
    for col, day in col_map.items():
        val = ws.cell(row=date_row, column=col).value
        d = parse_date_cell(val)
        if d is None:
            errors.append(f"Fecha inválida columna {col} día {day}")
            continue
        wd[day] = f"{d.day:02d}/{d.month:02d}/{d.year}"
    for d in DAYS_ORDER:
        if d not in wd:
            errors.append(f"Falta fecha para {d}")
    if not errors and wd.get("Vie") and wd.get("Jue"):
        try:
            vie = datetime.datetime.strptime(wd["Vie"], "%d/%m/%Y").date()
            jue = datetime.datetime.strptime(wd["Jue"], "%d/%m/%Y").date()
            if (jue - vie).days != 6:
                warnings.append(f"Vie→Jue no son 6 días ({vie} → {jue})")
            if vie.weekday() != 4:
                warnings.append(f"La fecha en Vie no es viernes ({wd['Vie']})")
        except ValueError:
            pass
    return wd, errors, warnings


def extract_title_name(ws, header_row: int, sheet_title: str) -> str:
    if header_row > 1:
        v = ws.cell(row=1, column=1).value
        if v and isinstance(v, str):
            m = re.search(r"semana\s*(\d+)", v, re.I)
            if m:
                return f"Semana {m.group(1)}"
            s = v.strip()
            if s:
                return s[:120]
    st = (sheet_title or "").strip()
    if st.isdigit():
        return f"Semana {st}"
    return f"Semana {sheet_title}"[:120] if st else "Horario importado"


def _match_employee_name(raw: str, schedule: Dict[str, Dict[str, str]]) -> Optional[str]:
    """Fuzzy match de nombre parcial contra los empleados del schedule."""
    raw_upper = raw.upper()
    for emp in schedule:
        emp_upper = emp.upper()
        if raw_upper in emp_upper or emp_upper in raw_upper:
            return emp
    raw_noacc = _strip_accents(raw_upper)
    for emp in schedule:
        emp_noacc = _strip_accents(emp.upper())
        if raw_noacc in emp_noacc or emp_noacc in raw_noacc:
            return emp
    return None


def _extract_inline_task(raw: str) -> Tuple[str, Optional[str]]:
    """
    Extrae el texto entre paréntesis de una celda de horario.
    Retorna (texto_limpio_sin_parentesis, texto_tarea_o_None).
    Ejemplo: "06:00 AM - 04:00 PM (Medir Tanques)" → ("06:00 AM - 04:00 PM", "Medir Tanques")
    """
    match = re.search(r'\(([^)]+)\)', raw)
    if match:
        task_text = match.group(1).strip()
        clean = (raw[:match.start()] + raw[match.end():]).strip()
        return clean, task_text
    return raw, None


def _parse_cleaning_section_emp_format(
    ws,
    start_row: int,
    col_map: Dict[int, str],
    schedule: Dict[str, Dict[str, str]],
    warnings: List[str],
) -> Dict[str, Dict[str, Optional[str]]]:
    """
    Parsea la sección OBLIGACIONES/LIMPIEZA en formato "empleado en col A".

    Formato (semanas 10+):
      Col A = Nombre del colaborador
      Cols B-H = Tarea asignada ese día (ej: "Tanques", "Baños", "Ø" = sin tarea)

    Retorna daily_tasks: {empleado: {día: etiqueta | None}}
    """
    daily_tasks: Dict[str, Dict[str, Optional[str]]] = {
        emp: {d: None for d in DAYS_ORDER} for emp in schedule
    }

    max_r = ws.max_row or 0
    r = start_row
    stop_after = start_row + 35

    while r <= max_r and r <= stop_after:
        emp_cell = ws.cell(row=r, column=1).value
        if emp_cell is None or str(emp_cell).strip() == "":
            r += 1
            continue

        raw_emp = str(emp_cell).strip()

        # Parar si encontramos otro bloque de Colaborador (formato viejo) o sección nueva
        if _is_colaborador_a1(emp_cell):
            r += 1
            continue  # puede ser el header del bloque, skip y continuar

        # Ignorar filas de ruido
        low = _strip_accents(raw_emp.lower())
        if any(tok in low for tok in ("formato", "obligaciones", "limpieza", "aseo")):
            r += 1
            continue

        matched_emp = _match_employee_name(raw_emp, schedule)
        if not matched_emp:
            # No coincide con ningún empleado → puede ser fin de sección
            r += 1
            continue

        for col, day in col_map.items():
            task_val = ws.cell(row=r, column=col).value
            if task_val is None:
                continue
            task_str = str(task_val).strip()
            # "Ø", "0", vacío, guión → sin tarea
            if not task_str or task_str in ("Ø", "0", "-", "–"):
                continue
            canonical = _normalize_task_label(task_str)
            if canonical:
                if daily_tasks[matched_emp][day] is None:
                    shift_code = schedule[matched_emp].get(day)
                    label = _build_task_label(canonical, None, shift_code)
                    daily_tasks[matched_emp][day] = label
            else:
                warnings.append(
                    f"Limpieza: tarea no reconocida '{task_str[:40]}' para {matched_emp} en {day}"
                )

        r += 1

    return daily_tasks


def _detect_cleaning_format(
    ws,
    cleaning_start: int,
    col_map: Dict[int, str],
    schedule: Dict[str, Dict[str, str]],
) -> Tuple[str, int]:
    """
    Detecta el formato de la sección de limpieza y devuelve
    ('emp_in_col_a', data_start_row) o ('task_in_col_a', data_start_row).

    Lógica:
    - Si col A de la primera fila no-vacía coincide con un empleado → nuevo formato.
    - Si col A de la primera fila no-vacía es "Colaborador" → hay header interno,
      avanzar y re-detectar con la siguiente fila no-vacía.
    - Si col A coincide con una tarea conocida → formato viejo.
    - Default → nuevo formato (más común en este Excel).
    """
    max_r = ws.max_row or 0
    r = cleaning_start
    stop = cleaning_start + 5

    while r <= max_r and r <= stop:
        a_val = ws.cell(row=r, column=1).value
        if a_val is None or str(a_val).strip() == "":
            r += 1
            continue
        raw = str(a_val).strip()
        if _is_colaborador_a1(a_val):
            # Header interno de la sección → los datos empiezan en la siguiente fila
            return "emp_in_col_a", r + 1
        if _match_employee_name(raw, schedule):
            return "emp_in_col_a", r
        if _normalize_task_label(raw) is not None:
            return "task_in_col_a", r
        r += 1

    # Por defecto asumir formato nuevo
    return "emp_in_col_a", cleaning_start


def _parse_cleaning_section(
    ws,
    start_row: int,
    col_map: Dict[int, str],
    schedule: Dict[str, Dict[str, str]],
    warnings: List[str],
) -> Dict[str, Dict[str, Optional[str]]]:
    """
    Analiza la sección de LIMPIEZA/OBLIGACIONES del Excel y construye
    el daily_tasks en el mismo formato que assign_tasks() del motor.

    Detecta automáticamente si es formato viejo (tarea en col A) o
    nuevo (empleado en col A, semanas 10+) y delega al parser correcto.
    """
    fmt, data_start = _detect_cleaning_format(ws, start_row, col_map, schedule)

    if fmt == "emp_in_col_a":
        return _parse_cleaning_section_emp_format(ws, data_start, col_map, schedule, warnings)

    # ── Formato viejo: col A = Tarea, cols B-H = Empleado o flecha ──────────
    daily_tasks: Dict[str, Dict[str, Optional[str]]] = {
        emp: {d: None for d in DAYS_ORDER} for emp in schedule
    }

    max_r = ws.max_row or 0
    r = data_start
    stop_after = data_start + 30

    while r <= max_r and r <= stop_after:
        task_cell = ws.cell(row=r, column=1).value
        if task_cell is None or str(task_cell).strip() == "":
            r += 1
            continue

        raw_task = str(task_cell).strip()

        if _is_colaborador_a1(task_cell):
            break

        canonical_task = _normalize_task_label(raw_task)
        if canonical_task is None:
            r += 1
            continue

        for col, day in col_map.items():
            cell_val = ws.cell(row=r, column=col).value
            if cell_val is None:
                continue

            raw_cell = str(cell_val).strip()
            if not raw_cell:
                continue

            # ── Caso 1: flecha ───────────────────────────────────────────────
            arrow_char = None
            if any(c in raw_cell for c in _ARROW_UP_CHARS):
                arrow_char = "↑"
            elif any(c in raw_cell for c in _ARROW_DOWN_CHARS):
                arrow_char = "↓"

            if arrow_char is not None:
                target_time = "AM" if arrow_char == "↑" else "PM"
                assigned_emp = None
                for emp, emp_days in schedule.items():
                    code = emp_days.get(day, "OFF")
                    if code in ("OFF", "VAC", "PERM"):
                        continue
                    if _classify_shift_time(code) == target_time:
                        if daily_tasks.get(emp, {}).get(day) is None:
                            assigned_emp = emp
                            break
                if not assigned_emp:
                    for emp, emp_days in schedule.items():
                        code = emp_days.get(day, "OFF")
                        if code in ("OFF", "VAC", "PERM"):
                            continue
                        if _classify_shift_time(code) == target_time:
                            assigned_emp = emp
                            break

                if assigned_emp:
                    if daily_tasks.get(assigned_emp, {}).get(day) is None:
                        shift_code = schedule[assigned_emp].get(day)
                        label = _build_task_label(canonical_task, arrow_char, shift_code)
                        daily_tasks.setdefault(assigned_emp, {d: None for d in DAYS_ORDER})[day] = label
                else:
                    warnings.append(
                        f"Limpieza '{canonical_task}' flecha {arrow_char} en {day}: "
                        f"no encontré empleado con turno {'AM' if arrow_char=='↑' else 'PM'}"
                    )
                continue

            # ── Caso 2: nombre de empleado ───────────────────────────────────
            matched_emp = _match_employee_name(raw_cell, schedule)
            if matched_emp:
                if daily_tasks.get(matched_emp, {}).get(day) is None:
                    shift_code = schedule[matched_emp].get(day)
                    label = _build_task_label(canonical_task, None, shift_code)
                    daily_tasks.setdefault(matched_emp, {d: None for d in DAYS_ORDER})[day] = label
            else:
                warnings.append(
                    f"Limpieza '{canonical_task}' en {day}: "
                    f"celda '{raw_cell[:40]}' no corresponde a ningún empleado"
                )

        r += 1

    return daily_tasks


def parse_horario_sheet(ws, sheet_name: str) -> Dict[str, Any]:
    """
    Devuelve dict con keys: sheet, name_sugerido, week_dates, schedule, daily_tasks, warnings, errors.
    """
    warnings: List[str] = []
    errors: List[str] = []
    header_row, col_map, ferr = find_header_and_column_map(ws)
    errors.extend(ferr)
    if header_row is None or not col_map:
        return {
            "sheet": sheet_name,
            "name_sugerido": extract_title_name(ws, 2, sheet_name),
            "week_dates": {},
            "schedule": {},
            "daily_tasks": {},
            "warnings": warnings,
            "errors": errors,
        }
    date_row = header_row - 1
    if date_row < 1:
        errors.append("No hay fila de fechas sobre el encabezado")
        return {
            "sheet": sheet_name,
            "name_sugerido": extract_title_name(ws, header_row, sheet_name),
            "week_dates": {},
            "schedule": {},
            "daily_tasks": {},
            "warnings": warnings,
            "errors": errors,
        }
    week_dates, date_errors, date_warn = build_week_dates_from_row(ws, date_row, col_map)
    errors.extend(date_errors)
    warnings.extend(date_warn)

    schedule: Dict[str, Dict[str, str]] = {}
    # Tareas inline extraídas de las celdas de horario (ej: "06:00 AM (Medir Tanques)")
    inline_tasks: Dict[str, Dict[str, Optional[str]]] = {}

    # ── Fase 1: leer empleados ────────────────────────────────────────────────
    r = header_row + 1
    max_r = ws.max_row or 0
    cleaning_section_start: Optional[int] = None

    # Tokens que identifican el inicio de la sección de limpieza
    cleaning_tokens = ("obligaciones", "limpieza", "aseo")

    while r <= max_r:
        name_cell = ws.cell(row=r, column=1).value
        if name_cell is None or str(name_cell).strip() == "":
            r += 1
            continue
        name = str(name_cell).strip()
        low = _strip_accents(name.lower())

        # Detectar inicio de la sección de limpieza
        if any(low.startswith(t) for t in cleaning_tokens) or any(t in low for t in cleaning_tokens):
            cleaning_section_start = r + 1  # las tareas empiezan en la fila siguiente
            break

        # Detectar otro encabezado Colaborador (segunda tabla en la misma hoja)
        if _is_colaborador_a1(name_cell):
            break

        # Detectar token "FORMATO" (ruido)
        if low.startswith("formato"):
            r += 1
            continue

        # Leer el horario del empleado — extrayendo tareas inline si las hay
        row_sched: Dict[str, str] = {}
        row_inline: Dict[str, Optional[str]] = {}
        for col, day in col_map.items():
            cell_val = ws.cell(row=r, column=col).value
            raw_str = str(cell_val).strip() if cell_val is not None else ""
            # Extraer tarea entre paréntesis antes de parsear el código de turno
            clean_str, inline_task_text = _extract_inline_task(raw_str)
            row_sched[day] = readable_cell_to_code(clean_str if clean_str else cell_val, warnings)
            if inline_task_text:
                canonical = _normalize_task_label(inline_task_text)
                row_inline[day] = canonical if canonical else inline_task_text
        if any(v != "OFF" for v in row_sched.values()) or name:
            schedule[name] = {d: row_sched.get(d, "OFF") for d in DAYS_ORDER}
            if any(v for v in row_inline.values()):
                inline_tasks[name] = {d: row_inline.get(d) for d in DAYS_ORDER}
        r += 1

    if not schedule:
        errors.append("No se encontraron filas de empleados con datos")

    # ── Fase 2: leer tareas de limpieza ──────────────────────────────────────
    daily_tasks: Dict[str, Dict[str, Optional[str]]] = {}
    if cleaning_section_start is not None and schedule:
        daily_tasks = _parse_cleaning_section(
            ws,
            cleaning_section_start,
            col_map,
            schedule,
            warnings,
        )
    else:
        # Sin sección separada de limpieza → usar tareas inline si las hay
        daily_tasks = {emp: {d: None for d in DAYS_ORDER} for emp in schedule}

    # Merge: las tareas de sección separada tienen prioridad;
    # las inline rellenan donde la sección separada no asignó nada.
    for emp, days_inline in inline_tasks.items():
        emp_dt = daily_tasks.setdefault(emp, {d: None for d in DAYS_ORDER})
        for day, task in days_inline.items():
            if task and emp_dt.get(day) is None:
                emp_dt[day] = task

    return {
        "sheet": sheet_name,
        "name_sugerido": extract_title_name(ws, header_row, sheet_name),
        "week_dates": week_dates,
        "schedule": schedule,
        "daily_tasks": daily_tasks,
        "warnings": warnings,
        "errors": errors,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PARSER FORMATO VERTICAL (Semanas 1-8)
# Formato: Col A = rango de horas, Cols B-H = nombres de empleados asignados
# ─────────────────────────────────────────────────────────────────────────────

_TIME_RANGE_RE = re.compile(
    r"(\d{1,2})(?::(\d{2}))?\s*(?:am|pm|a\.m\.?|p\.m\.?|hrs?)?\.?"
    r"\s*(?:a|al?|-|–)\s*"
    r"(\d{1,2})(?::(\d{2}))?\s*(?:am|pm|p\.m\.?|a\.m\.?|hrs?)?\.?",
    re.IGNORECASE,
)

_AMPM_RE = re.compile(r"(am|pm)", re.IGNORECASE)


def _parse_time_token(h_str: str, m_str: Optional[str], context_raw: str, position: str) -> int:
    """Convierte hora + minutos a entero 0-23. Maneja am/pm según posición en el string."""
    h = int(h_str)
    # Determinar si el contexto menciona am/pm alrededor del número
    # Usar la mitad izquierda/derecha del string original para inferir am/pm
    half = len(context_raw) // 2
    left_part = context_raw[:half].lower()
    right_part = context_raw[half:].lower()
    relevant = left_part if position == "start" else right_part
    is_pm = "pm" in relevant or "p.m" in relevant
    is_am = "am" in relevant or "a.m" in relevant
    if is_pm and h != 12:
        h += 12
    elif is_am and h == 12:
        h = 0
    return h % 24


def _time_range_to_shift_code(raw: str, warnings: List[str]) -> Optional[str]:
    """
    Convierte un texto como "5:00 am a 1:00 pm" a un código de turno normalizado.
    Retorna el código MANUAL_ si no hay match exacto en SHIFTS.
    """
    raw_clean = raw.strip()
    m = _TIME_RANGE_RE.search(raw_clean)
    if not m:
        return None

    h1_str, m1_str, h2_str, m2_str = m.group(1), m.group(2), m.group(3), m.group(4)
    h1 = _parse_time_token(h1_str, m1_str, raw_clean, "start")
    h2 = _parse_time_token(h2_str, m2_str, raw_clean, "end")

    # Construir candidato de texto normalizado para buscar en catálogo
    candidate = f"{h1:02d}:00 - {h2:02d}:00"
    norm = horario_db.normalize_manual_shift_code(f"{h1}-{h2}")
    if norm:
        return norm

    warnings.append(f"Rango de hora sin match en catálogo: '{raw_clean[:60]}' → usando MANUAL_{h1}-{h2}")
    manual = horario_db.normalize_manual_shift_code(f"{h1:02d}:{m1_str or '00'}-{h2:02d}:{m2_str or '00'}")
    return manual or f"MANUAL_{h1}-{h2}"


def _is_time_range_cell(val: Any) -> bool:
    """Retorna True si la celda parece contener un rango de horas de trabajo."""
    if val is None:
        return False
    s = _strip_accents(str(val).strip().lower())
    return bool(_TIME_RANGE_RE.search(s))


def _is_horas_header(val: Any) -> bool:
    if val is None:
        return False
    s = _strip_accents(str(val).strip().lower())
    return s.startswith("hora") or s.startswith("turno")


def _split_employee_names(cell_text: str) -> List[str]:
    """Separa múltiples nombres en una celda (por '/', '+', 'y', salto de línea)."""
    parts = re.split(r"[/\+\n]|\s+y\s+", cell_text, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]


def parse_horario_sheet_vertical(ws, sheet_name: str) -> Dict[str, Any]:
    """
    Parser para el formato VERTICAL (Semanas 1-8):
      - Col A: Rangos de hora o vacío
      - Cols B-H: Días de la semana en encabezado; empleados en celdas de datos
    """
    warnings: List[str] = []
    errors: List[str] = []
    max_scan = min(ws.max_row or 0, 50)

    # ── Buscar fila con días en cols B-H ─────────────────────────────────────
    header_row = None
    col_map: Dict[int, str] = {}
    for r in range(1, max_scan + 1):
        tmp: Dict[int, str] = {}
        max_c = min(ws.max_column or 15, 20)
        for c in range(2, max_c + 1):
            day = parse_day_header(ws.cell(row=r, column=c).value)
            if day:
                tmp[c] = day
        if set(tmp.values()) == set(DAYS_ORDER) and len(tmp) == len(DAYS_ORDER):
            header_row = r
            col_map = tmp
            break

    if header_row is None:
        errors.append("Formato vertical: no se encontró fila de días (Vie…Jue)")
        return {
            "sheet": sheet_name,
            "name_sugerido": extract_title_name(ws, 2, sheet_name),
            "week_dates": {},
            "schedule": {},
            "daily_tasks": {},
            "warnings": warnings,
            "errors": errors,
        }

    # ── Leer fechas (fila justo antes del header de días, si existe) ──────────
    date_row = header_row - 1
    week_dates: Dict[str, str] = {}
    if date_row >= 1:
        wd, date_errors, date_warn = build_week_dates_from_row(ws, date_row, col_map)
        if not date_errors:
            week_dates = wd
            warnings.extend(date_warn)

    # ── Leer filas de datos: Col A = rango de hora, Cols B-H = empleados ─────
    schedule: Dict[str, Dict[str, str]] = {}
    current_shift_code: Optional[str] = None
    max_r = ws.max_row or 0

    for r in range(header_row + 1, max_r + 1):
        a_val = ws.cell(row=r, column=1).value
        a_str = str(a_val).strip() if a_val is not None else ""

        # Detectar fin de la tabla de horarios (sección de limpieza u otro)
        low_a = _strip_accents(a_str.lower())
        if any(tok in low_a for tok in ("obligaciones", "limpieza", "aseo")):
            break

        # Si col A tiene rango de hora, actualizar el turno vigente
        if _is_time_range_cell(a_val):
            code = _time_range_to_shift_code(a_str, warnings)
            current_shift_code = code

        # Si col A tiene "LIBRE", "VACACIONES", etc. mapearlo también
        elif a_str.upper() in ("LIBRE", "FREE", "OFF"):
            current_shift_code = "OFF"
        elif a_str.upper() in ("VACACIONES", "VAC"):
            current_shift_code = "VAC"
        elif a_str.upper() == "PERMISO":
            current_shift_code = "PERM"

        if current_shift_code is None:
            continue  # fila sin turno identificado

        # Leer empleados de cada columna de día
        for col, day in col_map.items():
            cell_val = ws.cell(row=r, column=col).value
            if cell_val is None:
                continue
            cell_str = str(cell_val).strip()
            if not cell_str or cell_str in ("-", "–", "x", "X"):
                continue

            names = _split_employee_names(cell_str)
            for emp_name in names:
                if not emp_name:
                    continue
                if emp_name not in schedule:
                    schedule[emp_name] = {d: "OFF" for d in DAYS_ORDER}
                # Solo asignar si aún es OFF (el primero que aparece en la hoja gana)
                if schedule[emp_name][day] == "OFF":
                    schedule[emp_name][day] = current_shift_code

    if not schedule:
        errors.append("Formato vertical: no se encontraron empleados con datos")

    daily_tasks = {emp: {d: None for d in DAYS_ORDER} for emp in schedule}

    return {
        "sheet": sheet_name,
        "name_sugerido": extract_title_name(ws, header_row, sheet_name),
        "week_dates": week_dates,
        "schedule": schedule,
        "daily_tasks": daily_tasks,
        "warnings": warnings,
        "errors": errors,
    }


def _sheet_has_horizontal_format(ws) -> bool:
    """Retorna True si la hoja tiene el formato horizontal (Colaborador en col A)."""
    max_scan = min(ws.max_row or 0, 40)
    for r in range(1, max_scan + 1):
        if _is_colaborador_a1(ws.cell(row=r, column=1).value):
            return True
    return False


def _sheet_has_vertical_format(ws) -> bool:
    """Retorna True si la hoja parece tener el formato vertical (rangos de hora en col A)."""
    max_scan = min(ws.max_row or 0, 40)
    time_count = 0
    for r in range(1, max_scan + 1):
        if _is_time_range_cell(ws.cell(row=r, column=1).value):
            time_count += 1
            if time_count >= 2:
                return True
    return False


def list_sheet_names(path: str) -> List[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def parse_workbook_sheets(path: str, sheet_names: List[str]) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        out = []
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                out.append({
                    "sheet": sn,
                    "name_sugerido": sn,
                    "week_dates": {},
                    "schedule": {},
                    "daily_tasks": {},
                    "warnings": [],
                    "errors": [f"Hoja '{sn}' no existe en el archivo"],
                })
                continue
            ws = wb[sn]
            if _sheet_has_horizontal_format(ws):
                out.append(parse_horario_sheet(ws, sn))
            elif _sheet_has_vertical_format(ws):
                out.append(parse_horario_sheet_vertical(ws, sn))
            else:
                # Intentar horizontal como fallback (puede tener errores)
                result = parse_horario_sheet(ws, sn)
                if result["errors"]:
                    result["warnings"].insert(0, "Formato de hoja no reconocido — parseo parcial")
                out.append(result)
        return out
    finally:
        wb.close()
