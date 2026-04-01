from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl

import database as plan_db


AUTO_NOTE_PREFIX = "Rebajo automatico desde planilla"


def build_semana_planilla_key(mes_id, num_semana, viernes_str):
    return f"mes:{mes_id}|sem:{num_semana}|vie:{viernes_str}"


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_amount(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = str(value).strip()
    if not text:
        return 0.0

    cleaned = (
        text.replace("₡", "")
        .replace(",", "")
        .replace("$", "")
        .strip()
    )
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return 0.0


def _build_auto_note(semana):
    start = _parse_iso_date(semana["viernes"])
    if not start:
        return f"{AUTO_NOTE_PREFIX} - Semana {semana['num_semana']}"

    end = start + timedelta(days=6)
    return (
        f"{AUTO_NOTE_PREFIX} - Semana {semana['num_semana']} "
        f"({start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')})"
    )


def _select_prestamo_for_week(loans, semana):
    if not loans:
        return None

    week_start = _parse_iso_date(semana["viernes"])
    if not week_start:
        return loans[0]
    week_end = week_start + timedelta(days=6)

    candidates = []
    fallback = []
    for loan in loans:
        start = _parse_iso_date(loan.get("fecha_inicio"))
        if start and start > week_end:
            continue

        fallback.append(loan)
        liquidacion = _parse_iso_date(loan.get("fecha_liquidacion"))
        if liquidacion and liquidacion < week_start:
            continue
        candidates.append(loan)

    pool = candidates or fallback
    if not pool:
        return None

    pool.sort(
        key=lambda loan: (
            1 if loan.get("estado") == "activo" else 0,
            loan.get("fecha_inicio") or "",
            loan.get("id") or 0,
        )
    )
    return pool[-1]


def _scan_week_sheet(ws, emp_by_name):
    found = defaultdict(float)
    for row in range(5, ws.max_row + 1):
        emp_name = ws.cell(row=row, column=1).value
        if not isinstance(emp_name, str):
            continue

        emp_name = emp_name.strip()
        empleado_id = emp_by_name.get(emp_name)
        if not empleado_id:
            continue

        # Columna N (14) = Préstamo; M (13) = Bonificaciones
        monto = _parse_amount(ws.cell(row=row, column=14).value)
        if monto > 0:
            found[empleado_id] += monto

    return found


def _recalculate_affected_loans(conn, loan_ids):
    for loan_id in sorted(loan_ids):
        plan_db._recalcular_prestamo_conn(conn, loan_id)


def sync_rebajos_mes(mes, archivo_path, semanas=None):
    result = {
        "created": 0,
        "updated": 0,
        "deleted": 0,
        "skipped": 0,
        "affected_prestamos": 0,
    }
    if not mes or not archivo_path:
        return result

    semanas = semanas if semanas is not None else plan_db.get_semanas_del_mes(mes["id"])
    prefix = f"mes:{mes['id']}|"

    conn = plan_db.get_conn()
    try:
        emp_rows = conn.execute("SELECT id, nombre FROM empleados").fetchall()
        emp_by_name = {
            row["nombre"].strip(): row["id"]
            for row in emp_rows
            if row["nombre"]
        }

        loan_rows = conn.execute(
            "SELECT * FROM prestamos ORDER BY empleado_id, fecha_inicio, id"
        ).fetchall()
        loans_by_emp = defaultdict(list)
        for row in loan_rows:
            loans_by_emp[row["empleado_id"]].append(dict(row))

        existing_rows = conn.execute(
            """
            SELECT pa.id, pa.prestamo_id, pa.monto, pa.fecha, pa.semana_planilla, pa.notas,
                   p.empleado_id
            FROM prestamo_abonos pa
            JOIN prestamos p ON p.id = pa.prestamo_id
            WHERE pa.tipo='planilla' AND pa.semana_planilla LIKE ?
            ORDER BY pa.id
            """,
            (f"{prefix}%",),
        ).fetchall()
        existing_by_key = defaultdict(list)
        for row in existing_rows:
            existing_by_key[(row["semana_planilla"], row["empleado_id"])].append(dict(row))

        desired = {}
        if archivo_path and semanas:
            wb = openpyxl.load_workbook(archivo_path, data_only=True)
            for semana in semanas:
                week_key = build_semana_planilla_key(
                    mes["id"], semana["num_semana"], semana["viernes"]
                )
                sheet_name = f"Semana {semana['num_semana']}"
                if sheet_name not in wb.sheetnames:
                    continue

                scan = _scan_week_sheet(wb[sheet_name], emp_by_name)
                if not scan:
                    continue

                note = _build_auto_note(semana)
                for empleado_id, monto in scan.items():
                    desired[(week_key, empleado_id)] = {
                        "monto": round(monto, 2),
                        "fecha": semana["viernes"],
                        "notas": note,
                        "semana": semana,
                    }

        affected_loans = set()

        for key, data in desired.items():
            existing_group = existing_by_key.pop(key, [])
            primary = existing_group[0] if existing_group else None

            for duplicate in existing_group[1:]:
                conn.execute("DELETE FROM prestamo_abonos WHERE id=?", (duplicate["id"],))
                affected_loans.add(duplicate["prestamo_id"])
                result["deleted"] += 1

            if primary:
                updates = []
                params = []
                if round(float(primary["monto"] or 0), 2) != data["monto"]:
                    updates.append("monto=?")
                    params.append(data["monto"])
                if primary["fecha"] != data["fecha"]:
                    updates.append("fecha=?")
                    params.append(data["fecha"])
                if (primary["notas"] or "") != data["notas"]:
                    updates.append("notas=?")
                    params.append(data["notas"])

                if updates:
                    params.append(primary["id"])
                    conn.execute(
                        f"UPDATE prestamo_abonos SET {', '.join(updates)} WHERE id=?",
                        params,
                    )
                    result["updated"] += 1
                affected_loans.add(primary["prestamo_id"])
                continue

            prestamo = _select_prestamo_for_week(
                loans_by_emp.get(key[1], []),
                data["semana"],
            )
            if not prestamo:
                result["skipped"] += 1
                continue

            conn.execute(
                """
                INSERT INTO prestamo_abonos
                (prestamo_id, monto, tipo, fecha, semana_planilla, notas, fecha_registro)
                VALUES (?, ?, 'planilla', ?, ?, ?, ?)
                """,
                (
                    prestamo["id"],
                    data["monto"],
                    data["fecha"],
                    key[0],
                    data["notas"],
                    datetime.now().isoformat(),
                ),
            )
            affected_loans.add(prestamo["id"])
            result["created"] += 1

        for stale_group in existing_by_key.values():
            for stale in stale_group:
                conn.execute("DELETE FROM prestamo_abonos WHERE id=?", (stale["id"],))
                affected_loans.add(stale["prestamo_id"])
                result["deleted"] += 1

        _recalculate_affected_loans(conn, affected_loans)
        conn.commit()
        result["affected_prestamos"] = len(affected_loans)
        return result
    finally:
        conn.close()


def clear_auto_rebajos_mes(mes_id):
    prefix = f"mes:{mes_id}|"
    result = {"deleted": 0, "affected_prestamos": 0}
    conn = plan_db.get_conn()
    try:
        rows = conn.execute(
            """
            SELECT id, prestamo_id
            FROM prestamo_abonos
            WHERE tipo='planilla' AND semana_planilla LIKE ?
            """,
            (f"{prefix}%",),
        ).fetchall()
        if not rows:
            return result

        affected_loans = {row["prestamo_id"] for row in rows}
        conn.execute(
            "DELETE FROM prestamo_abonos WHERE tipo='planilla' AND semana_planilla LIKE ?",
            (f"{prefix}%",),
        )
        _recalculate_affected_loans(conn, affected_loans)
        conn.commit()
        result["deleted"] = len(rows)
        result["affected_prestamos"] = len(affected_loans)
        return result
    finally:
        conn.close()
