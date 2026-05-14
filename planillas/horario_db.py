"""
horario_db.py - Puente entre el generador de horarios y la planilla de pago
============================================================================
Funciones para:
  - Migrar datos de database.json a SQLite
  - CRUD de empleados/config/horarios del scheduler
  - Clasificar horas según turno (diurna/nocturna/mixta)
  - Rellenar automáticamente las horas en el Excel de planilla
"""
import json
import os
import re
import sqlite3
from datetime import date as date_cls
from datetime import datetime, timedelta

import database as db

# ── DEFINICIONES DE TURNOS (copiado de scheduler_engine.py) ─────────────────
SHIFTS = {
    "OFF": set(), "VAC": set(), "PERM": set(),
    "N_22-05": set([22, 23, 24, 25, 26, 27, 28]),
    "T1_05-13": set(range(5, 13)), "T2_06-14": set(range(6, 14)),
    "T3_07-15": set(range(7, 15)), "T4_08-16": set(range(8, 16)),
    "T8_13-20": set(range(13, 20)), "T9_14-21": set(range(14, 21)),
    "T10_15-22": set(range(15, 22)), "T11_12-20": set(range(12, 20)),
    "T12_14-22": set(range(14, 22)), "T13_16-22": set(range(16, 22)),
    "T16_05-14": set(range(5, 14)),
    "J_06-16": set(range(6, 16)), "J_07-17": set(range(7, 17)),
    "J_08-18": set(range(8, 18)), "J_09-19": set(range(9, 19)),
    "J_10-20": set(range(10, 20)),
    "E1_07-18": set(range(7, 18)), "E2_08-19": set(range(8, 19)),
    "T17_16-23": set(range(16, 23)),
    "X_07-19": set(range(7, 19)), "X_08-20": set(range(8, 20)),
    "D1_05-13": set(range(5, 13)), "D2_14-22": set(range(14, 22)),
    "D3_15-23": set(range(15, 23)), "D4_13-22": set(range(13, 22)),
    "R1_07-11": set(range(7, 11)), "R2_16-20": set(range(16, 20)),
    "Q1_05-11+17-20": set(range(5, 11)) | set(range(17, 20)),
    "Q2_07-11+17-20": set(range(7, 11)) | set(range(17, 20)),
    "Q3_05-11+17-22": set(range(5, 11)) | set(range(17, 22)),
}

MANUAL_SHIFT_PREFIX = "MANUAL_"
DIAS_MAP = {"Vie": 0, "Sáb": 1, "Dom": 2, "Lun": 3, "Mar": 4, "Mié": 5, "Jue": 6}
DIAS_EXCEL = ["Viernes", "Sabado", "Domingo", "Lunes", "Martes", "Miercoles", "Jueves"]


# ═════════════════════════════════════════════════════════════════════════════
# CLASIFICACIÓN DE HORAS
# ═════════════════════════════════════════════════════════════════════════════

def _parse_manual_time_token(token):
    raw = str(token or "").strip().lower().replace(".", "")
    raw = re.sub(r"\s+", "", raw)
    if not raw:
        return None

    match = re.fullmatch(r"(\d{1,2})(?::(\d{2}))?(am|pm)", raw)
    if match:
        hour = int(match.group(1))
        minutes = int(match.group(2) or 0)
        suffix = match.group(3)
        if minutes != 0 or hour < 1 or hour > 12:
            return None
        if suffix == "am":
            return 0 if hour == 12 else hour
        return hour if hour == 12 else hour + 12

    match = re.fullmatch(r"(\d{1,2})(?::(\d{2}))?", raw)
    if not match:
        return None

    hour = int(match.group(1))
    minutes = int(match.group(2) or 0)
    if minutes != 0 or hour < 0 or hour > 29:
        return None
    return hour


def _split_manual_range_segment(segment):
    cleaned = str(segment or "").strip().replace("–", "-").replace("—", "-")
    if not cleaned:
        return None

    for pattern in (
        re.compile(r"\s*-\s*"),
        re.compile(r"\s+a\s+", re.IGNORECASE),
        re.compile(r"\s+to\s+", re.IGNORECASE),
    ):
        parts = pattern.split(cleaned, maxsplit=1)
        if len(parts) == 2:
            return parts[0], parts[1]

    return None


def normalize_manual_shift_code(shift_code):
    if not isinstance(shift_code, str):
        return None

    raw = shift_code.strip()
    if not raw:
        return None

    upper = raw.upper()
    if upper in SHIFTS:
        return upper
    if upper in ("OFF", "LIBRE", "DESCANSO"):
        return "OFF"
    if upper in ("VAC", "VACACIONES"):
        return "VAC"
    if upper in ("PERM", "PERMISO"):
        return "PERM"

    candidate = raw
    if upper.startswith(MANUAL_SHIFT_PREFIX):
        candidate = raw[len(MANUAL_SHIFT_PREFIX):]

    segments = [seg for seg in re.split(r"\s*(?:\+|/|,)\s*", candidate) if seg]
    if not segments:
        return None

    normalized_segments = []
    for segment in segments:
        split = _split_manual_range_segment(segment)
        if not split:
            return None

        start = _parse_manual_time_token(split[0])
        end = _parse_manual_time_token(split[1])
        if start is None or end is None:
            return None

        normalized_segments.append(f"{start:02d}-{end:02d}")

    return f"{MANUAL_SHIFT_PREFIX}{'+'.join(normalized_segments)}"


def get_shift_hours_set(shift_code):
    normalized = normalize_manual_shift_code(shift_code)
    if normalized in SHIFTS:
        return set(SHIFTS[normalized])

    if not normalized or not normalized.startswith(MANUAL_SHIFT_PREFIX):
        return set()

    horas = set()
    for segment in normalized[len(MANUAL_SHIFT_PREFIX):].split("+"):
        start_raw, end_raw = segment.split("-")
        start = int(start_raw)
        end = int(end_raw)
        if end <= start:
            end += 24
        horas.update(range(start, end))
    return horas


def clasificar_turno(turno_code):
    """
    Clasifica un turno y devuelve (horas_diurnas, horas_nocturnas, horas_mixtas).

    Reglas:
      - Turno inicia antes de las 12:00 -> DIURNA
      - Turno inicia >=12:00 y <22:00 -> MIXTA
      - Turno N_22-05 -> NOCTURNA
      - OFF/VAC/PERM -> (0, 0, 0)
    """
    normalized = normalize_manual_shift_code(turno_code)
    if normalized in ("OFF", "VAC", "PERM"):
        return (0, 0, 0)

    horas_set = get_shift_hours_set(turno_code)
    if not horas_set:
        return (0, 0, 0)

    total_horas = len(horas_set)
    min_hora = min(h % 24 for h in horas_set)  # Normalize 22-28 -> 22-4

    if normalized == "N_22-05":
        return (0, total_horas, 0)  # 100% nocturna
    elif normalized == "T17_16-23":
        return (0, 1, 6)  # Especial: 1h nocturna, 6h mixtas
    elif normalized and normalized.startswith(MANUAL_SHIFT_PREFIX):
        nocturnas = sum(1 for h in horas_set if (h % 24) >= 22 or (h % 24) < 5)
        regulares = total_horas - nocturnas
        if min_hora >= 22 or min_hora < 5:
            return (0, total_horas, 0)
        if min_hora < 12:
            return (regulares, nocturnas, 0)
        return (0, nocturnas, regulares)
    elif min_hora < 12:
        return (total_horas, 0, 0)  # Inicia antes de mediodía -> diurna
    else:
        return (0, 0, total_horas)  # Inicia a mediodía o después -> mixta


def procesar_horario_semana(horario_dict):
    """
    Toma un dict {empleado: {dia: turno}} y devuelve un dict
    {empleado: {dia: (h_diurnas, h_nocturnas, h_mixtas)}}.
    """
    resultado = {}
    for emp, dias in horario_dict.items():
        resultado[emp] = {}
        for dia, turno in dias.items():
            resultado[emp][dia] = clasificar_turno(turno)
    return resultado


# Orden de días en hoja semanal (Vie → Jue)
DIAS_SEMANA_PLANILLA = ("Vie", "Sáb", "Dom", "Lun", "Mar", "Mié", "Jue")

# Horas estándar para feriado no laborado / permiso en feriado (fila ★)
HORAS_FERIADO_SIN_LABOR = 8
# Horas diurnas cargadas en planilla por día de vacaciones
HORAS_VAC_DIURNAS = 8


def turno_categoria_planilla(turno_code):
    """VAC | PERM | OFF | WORK según código de turno del horario."""
    n = normalize_manual_shift_code(turno_code)
    if n == "VAC":
        return "VAC"
    if n == "PERM":
        return "PERM"
    if n == "OFF":
        return "OFF"
    return "WORK"


def horas_base_dia_planilla(turno_code):
    """
    Horas (diurnas, nocturnas, mixtas) antes del cap de jefe.
    VAC → 8 diurnas; PERM/OFF → 0; resto → clasificar_turno.
    """
    cat = turno_categoria_planilla(turno_code)
    if cat == "VAC":
        return (HORAS_VAC_DIURNAS, 0, 0)
    if cat in ("PERM", "OFF"):
        return (0, 0, 0)
    return clasificar_turno(turno_code)


# Topes de jornada ordinaria (horas); el excedente es extraordinario del mismo tipo.
CAP_ORD_DIURNA = 8.0
CAP_ORD_MIXTA = 7.0
CAP_ORD_NOCTURNA = 6.0


def split_jornada_ordinaria_extra(h_diurna, h_nocturna, h_mixta):
    """
    Parte las horas del día en ordinarias vs extraordinarias por tipo
    (ley: diurna 8 h, mixta 7 h, nocturna 6 h).
    Retorna (od, om, on, ed, em, en).
    """
    h_d = float(h_diurna or 0)
    h_n = float(h_nocturna or 0)
    h_m = float(h_mixta or 0)
    od = min(h_d, CAP_ORD_DIURNA)
    ed = max(0.0, h_d - CAP_ORD_DIURNA)
    om = min(h_m, CAP_ORD_MIXTA)
    em = max(0.0, h_m - CAP_ORD_MIXTA)
    on = min(h_n, CAP_ORD_NOCTURNA)
    en = max(0.0, h_n - CAP_ORD_NOCTURNA)
    return od, om, on, ed, em, en


def aplicar_cap_jefe_pista(h_diurna, h_nocturna, h_mixta, es_jefe):
    """Excedente de 8h por tipo → horas extra. Devuelve (h_d, h_n, h_m, h_extra)."""
    h_extra = 0
    if es_jefe:
        if h_diurna > 8:
            h_extra += h_diurna - 8
            h_diurna = 8
        if h_mixta > 8:
            h_extra += h_mixta - 8
            h_mixta = 8
        if h_nocturna > 8:
            h_extra += h_nocturna - 8
            h_nocturna = 8
    return h_diurna, h_nocturna, h_mixta, h_extra


def tarifa_tipo_desde_totales_semana(total_d, total_m, total_n):
    """
    Igual criterio que _calcular_tarifa_dominante: mixta si todo 0;
    si no, tipo con más horas.
    Retorna (tarifa_tipo_str|None, es_todo_cero).
    """
    if total_d == 0 and total_m == 0 and total_n == 0:
        return None, True
    max_hours = max(total_d, total_m, total_n)
    if total_d == max_hours:
        return "diurna", False
    if total_m == max_hours:
        return "mixta", False
    return "nocturna", False


def feriado_celda_horas(cat, _tiene_horas_jornada=None):
    """
    Horas a escribir en fila ★ Feriado para un día que es feriado (columna del día).
    VAC → 0. Cualquier otro caso (OFF, PERM, WORK con o sin jornada) → recargo
    estándar HORAS_FERIADO_SIN_LABOR; las horas trabajadas siguen en D/M/N/extras.
    El segundo argumento se ignora (compat. llamadas antiguas).
    """
    if cat == "VAC":
        return 0
    return HORAS_FERIADO_SIN_LABOR


def _valor_tarifa_tipo(tarifa_tipo, tarifas):
    if tarifa_tipo == "diurna":
        return tarifas.get("tarifa_diurna", 0) or 0
    if tarifa_tipo == "mixta":
        return tarifas.get("tarifa_mixta", 0) or 0
    if tarifa_tipo == "nocturna":
        return tarifas.get("tarifa_nocturna", 0) or 0
    return 0


# ═════════════════════════════════════════════════════════════════════════════
# MIGRACIÓN JSON → SQLITE
# ═════════════════════════════════════════════════════════════════════════════

def migrar_json_a_sqlite(json_path):
    """Migra database.json del scheduler a las tablas SQLite."""
    if not os.path.exists(json_path):
        return False, f"No se encontró {json_path}"

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    conn = db.get_conn()

    # 1. Migrar empleados
    for emp in data.get("employees", []):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO horario_empleados
                (nombre, genero, puede_nocturno, allow_no_rest, forced_libres,
                 forced_quebrado, es_jefe_pista, turnos_fijos)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                emp["name"], emp.get("gender", "M"),
                1 if emp.get("can_do_night", True) else 0,
                1 if emp.get("allow_no_rest", False) else 0,
                1 if emp.get("forced_libres", False) else 0,
                1 if emp.get("forced_quebrado", False) else 0,
                1 if emp.get("is_jefe_pista", False) else 0,
                json.dumps(emp.get("fixed_shifts", {}), ensure_ascii=False),
            ))
        except sqlite3.IntegrityError:
            pass  # Ya existe

    # 2. Migrar config
    cfg = data.get("config", {})
    conn.execute("DELETE FROM horario_config")
    conn.execute("""
        INSERT INTO horario_config
        (id, night_mode, fixed_night_person, allow_long_shifts, use_refuerzo,
         refuerzo_type, allow_collision_quebrado, collision_peak_priority,
         sunday_cycle_index, sunday_rotation_queue, use_history, holidays)
        VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        cfg.get("night_mode", "rotation"),
        cfg.get("fixed_night_person"),
        1 if cfg.get("allow_long_shifts", False) else 0,
        1 if cfg.get("use_refuerzo", False) else 0,
        cfg.get("refuerzo_type", "diurno"),
        1 if cfg.get("allow_collision_quebrado", False) else 0,
        cfg.get("collision_peak_priority", "pm"),
        cfg.get("sunday_cycle_index", 0),
        json.dumps(cfg.get("sunday_rotation_queue")) if cfg.get("sunday_rotation_queue") else None,
        1 if cfg.get("use_history", True) else 0,
        json.dumps(cfg.get("holidays", [])),
    ))

    # 3. Migrar historial de horarios
    for entry in data.get("history_log", []):
        try:
            conn.execute("""
                INSERT INTO horarios_generados (nombre, horario, tareas, metadata, timestamp)
                VALUES (?, ?, ?, ?, ?)
            """, (
                entry.get("name", "SIN NOMBRE"),
                json.dumps(entry.get("schedule", {}), ensure_ascii=False),
                json.dumps(entry.get("daily_tasks", {}), ensure_ascii=False),
                json.dumps({
                    k: v for k, v in entry.items()
                    if k not in ("name", "schedule", "daily_tasks", "timestamp")
                }, ensure_ascii=False),
                entry.get("timestamp", datetime.now().isoformat()),
            ))
        except Exception:
            pass

    # 4. Migrar last_result como horario si existe
    last = data.get("last_result", {})
    if last.get("schedule"):
        conn.execute("""
            INSERT INTO horarios_generados (nombre, horario, tareas, metadata, timestamp)
            VALUES (?, ?, ?, ?, ?)
        """, (
            "ÚLTIMO GENERADO",
            json.dumps(last["schedule"], ensure_ascii=False),
            json.dumps(last.get("daily_tasks", {}), ensure_ascii=False),
            json.dumps(last.get("metadata", {}), ensure_ascii=False),
            datetime.now().isoformat(),
        ))

    conn.commit()
    conn.close()
    return True, "Migración completada"

# ═════════════════════════════════════════════════════════════════════════════
# CRUD HORARIOS
# ═════════════════════════════════════════════════════════════════════════════

def get_horarios_generados():
    """Listado para planillas: todas las filas activas (mismo nombre puede repetirse)."""
    conn = db.get_conn()
    rows = conn.execute(
        """
        SELECT id, nombre, timestamp
        FROM horarios_generados
        WHERE IFNULL(deleted, 0) = 0
        ORDER BY id DESC
        """
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]
def get_horario_por_id(horario_id):
    conn = db.get_conn()
    row = conn.execute(
        "SELECT * FROM horarios_generados WHERE id=?", (horario_id,)
    ).fetchone()
    conn.close()
    if row:
        d = dict(row)
        d["horario"] = json.loads(d["horario"])
        d["tareas"] = json.loads(d["tareas"]) if d["tareas"] else {}
        d["metadata"] = json.loads(d["metadata"]) if d["metadata"] else {}
        return d
    return None


def get_horario_por_nombre(nombre):
    """Si hay varios con el mismo nombre, devuelve el más reciente (mayor id)."""
    conn = db.get_conn()
    row = conn.execute(
        "SELECT * FROM horarios_generados WHERE nombre=? AND IFNULL(deleted, 0) = 0 ORDER BY id DESC LIMIT 1",
        (nombre,),
    ).fetchone()
    conn.close()
    if row:
        d = dict(row)
        d["horario"] = json.loads(d["horario"])
        d["tareas"] = json.loads(d["tareas"]) if d["tareas"] else {}
        d["metadata"] = json.loads(d["metadata"]) if d["metadata"] else {}
        return d
    return None


def guardar_horario(nombre, schedule_dict, tasks_dict=None, metadata_dict=None):
    """Siempre crea una fila nueva (no sobrescribe por nombre)."""
    conn = db.get_conn()
    horario_json = json.dumps(schedule_dict, ensure_ascii=False)
    tareas_json = json.dumps(tasks_dict or {}, ensure_ascii=False)
    metadata_json = json.dumps(metadata_dict or {}, ensure_ascii=False)
    ts = datetime.now().isoformat()
    conn.execute("""
        INSERT INTO horarios_generados (nombre, horario, tareas, metadata, timestamp, deleted, deleted_at)
        VALUES (?, ?, ?, ?, ?, 0, NULL)
    """, (nombre, horario_json, tareas_json, metadata_json, ts))
    conn.commit()
    conn.close()


def eliminar_horario(horario_id):
    conn = db.get_conn()
    conn.execute("DELETE FROM horarios_generados WHERE id=?", (horario_id,))
    conn.commit()
    conn.close()


def update_horario_tareas(horario_id, tareas_dict):
    """Actualiza las tareas de un horario existente."""
    conn = db.get_conn()
    tareas_json = json.dumps(tareas_dict or {}, ensure_ascii=False)
    conn.execute("UPDATE horarios_generados SET tareas=? WHERE id=?", (tareas_json, horario_id))
    conn.commit()
    conn.close()


# ═════════════════════════════════════════════════════════════════════════════
# RELLENAR EXCEL CON HORAS DEL HORARIO
# ═════════════════════════════════════════════════════════════════════════════

def _parse_viernes_planilla_cell(val):
    """Celda del viernes de inicio de período (fila 2 col C): date/datetime/serial/texto."""
    if val is None:
        return None
    if isinstance(val, date_cls) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(val).date()
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(s[:10] if len(s) > 10 else s, fmt).date()
            except ValueError:
                continue
    return None


def _holiday_entry_iso(h):
    """Normaliza h['date'] a YYYY-MM-DD para comparar con iso de la semana."""
    if not h:
        return None
    d = h.get("date")
    if d is None:
        return None
    if hasattr(d, "isoformat") and callable(getattr(d, "isoformat")) and not isinstance(d, str):
        return d.isoformat()[:10]
    s = str(d).strip()
    if "T" in s:
        s = s[:10]
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return f"{s[:4]}-{s[5:7]}-{s[8:10]}"
    parts = s.split("/")
    if len(parts) == 3:
        dd, mm, yy = parts[0].zfill(2), parts[1].zfill(2), parts[2]
        if len(yy) == 2:
            yy = "20" + yy if int(yy) < 70 else "19" + yy
        return f"{yy}-{mm}-{dd}"
    return None


def rellenar_horas_en_excel(excel_path, nombre_hoja, horario_dict, holidays=None):
    """
    Abre el Excel de planilla, busca la hoja semanal y rellena las horas
    de cada empleado según el horario generado.

    nombre_hoja: nombre de la pestaña en el Excel (e.g. "Semana 10")
    horario_dict: {empleado: {dia_es: turno_code}} con dias en español
                  (Vie, Sáb, Dom, Lun, Mar, Mié, Jue)
    holidays: list of {"date": "YYYY-MM-DD", "name": "..."} from config
    """
    import openpyxl

    if not os.path.exists(excel_path):
        return False, f"No se encontró {excel_path}", {}

    wb = openpyxl.load_workbook(excel_path)
    if nombre_hoja not in wb.sheetnames:
        return False, f"No se encontró la hoja '{nombre_hoja}'", {}

    ws = wb[nombre_hoja]

    # Layout: columnas C–I = Vie..Jue. Filas por empleado = bloque variable
    # (3 ordinarias + 0–3 extras + fila feriado opcional); ver planilla_layout.
    DIAS_COL = {
        "Vie": 3, "Sáb": 4, "Dom": 5,
        "Lun": 6, "Mar": 7, "Mié": 8, "Jue": 9,
    }

    import planilla_layout as pllay  # noqa: PLC0415 — evita import circular al cargar módulo

    conn = db.get_conn()
    empleados_planilla = {
        row["nombre"]: dict(row)
        for row in conn.execute(
            "SELECT nombre, tipo_pago, salario_fijo, "
            "COALESCE(NULLIF(TRIM(periodo_salario_fijo), ''), 'mensual') AS periodo_salario_fijo "
            "FROM empleados"
        ).fetchall()
    }

    # Obtener préstamos activos para rellenar la deducción de planilla (col 14 = Préstamo; 13 = Bonific.)
    prestamos_activos = {}
    prest_rows = conn.execute("""
        SELECT e.nombre, p.pago_semanal, p.saldo
        FROM prestamos p
        JOIN empleados e ON p.empleado_id = e.id
        WHERE p.estado = 'activo'
    """).fetchall()
    for pr in prest_rows:
        prestamos_activos[pr["nombre"]] = min(pr["pago_semanal"], pr["saldo"])

    conn.close()

    # ── Detectar feriados de esta semana (viernes en fila 2 col C) ──
    viernes_date = _parse_viernes_planilla_cell(ws.cell(row=2, column=3).value)
    semana_feriados = []  # list of (dia_es, col_idx, holiday_name)
    if holidays and viernes_date:
        for i, dia_es in enumerate(["Vie", "Sáb", "Dom", "Lun", "Mar", "Mié", "Jue"]):
            dia_date = viernes_date + timedelta(days=i)
            iso_date = dia_date.isoformat()
            for h in holidays:
                if _holiday_entry_iso(h) == iso_date:
                    col_idx = DIAS_COL.get(dia_es)
                    if col_idx:
                        semana_feriados.append((dia_es, col_idx, h.get("name") or ""))
                    break

    # Si hay feriados, desocultar filas de feriado para todos los empleados
    has_holidays_this_week = len(semana_feriados) > 0

    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment

    NO_FILL = PatternFill(fill_type=None)
    VAC_FILL = PatternFill(fill_type="solid", fgColor="FEF9C3")

    recap_horas = {}

    def _buscar_fila_empleado(emp_nombre, tipo_pago):
        for r in range(5, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if not (val and isinstance(val, str) and val.strip() == emp_nombre):
                continue

            jornada_val = ws.cell(row=r, column=2).value
            es_bloque_por_horas = jornada_val == "Hrs. Diurnas"

            if tipo_pago == "fijo":
                if not es_bloque_por_horas:
                    return r
                continue

            if es_bloque_por_horas:
                return r

        return None

    for emp_nombre in horario_dict:
        emp_info = empleados_planilla.get(emp_nombre, {})
        tipo_pago = emp_info.get("tipo_pago")
        salario_fijo = emp_info.get("salario_fijo") or 0
        periodo_sf = emp_info.get("periodo_salario_fijo") or "mensual"
        emp_row = _buscar_fila_empleado(emp_nombre, tipo_pago)

        if emp_row is None:
            continue  # Empleado no está en esta hoja

        # Inyectar abono de préstamo activo (si tiene) en la columna 14 (N = Préstamo; no en 13 Bonific.)
        pago_prestamo = prestamos_activos.get(emp_nombre)
        if pago_prestamo:
            ws.cell(row=emp_row, column=14).value = pago_prestamo

        if tipo_pago == "fijo":
            recap_horas[emp_nombre] = {
                "diurnas": 0,
                "mixtas": 0,
                "nocturnas": 0,
                "extra_diurnas": 0,
                "extra_mixtas": 0,
                "extra_nocturnas": 0,
                "extra": 0,
                "horas_feriado_no_labor": 0,
                "recargo_feriado": 0,
                "salario_bruto": db.salario_fijo_a_bruto_semanal(salario_fijo, periodo_sf) if salario_fijo else 0,
            }
            continue

        emp_schedule = horario_dict.get(emp_nombre) or {}

        scn = pllay.scan_jornada_block_rows(ws, emp_row)
        last_clear = pllay.jornada_block_last_row(scn)
        for col_clear in range(3, 10):
            for rr in range(emp_row, last_clear + 1):
                c_clear = ws.cell(row=rr, column=col_clear)
                c_clear.value = None
                c_clear.fill = NO_FILL
                c_clear.comment = None

        total_emp_d = 0
        total_emp_m = 0
        total_emp_n = 0
        total_emp_ed = 0
        total_emp_em = 0
        total_emp_en = 0

        for dia_sched in DIAS_SEMANA_PLANILLA:
            col = DIAS_COL.get(dia_sched)
            if col is None:
                continue

            turno = emp_schedule.get(dia_sched, "OFF")
            cat = turno_categoria_planilla(turno)
            h_diurna, h_nocturna, h_mixta = horas_base_dia_planilla(turno)
            od, om, on, ed, em, en = split_jornada_ordinaria_extra(
                h_diurna, h_nocturna, h_mixta
            )

            r_d = scn["rd"]
            r_m = scn["rm"]
            r_n = scn["rn"]
            r_ed = scn.get("ed")
            r_em = scn.get("em")
            r_en = scn.get("en")

            ws.cell(row=r_d, column=col).value = od if od > 0 else None
            ws.cell(row=r_m, column=col).value = om if om > 0 else None
            ws.cell(row=r_n, column=col).value = on if on > 0 else None
            if r_ed:
                ws.cell(row=r_ed, column=col).value = ed if ed > 0 else None
            if r_em:
                ws.cell(row=r_em, column=col).value = em if em > 0 else None
            if r_en:
                ws.cell(row=r_en, column=col).value = en if en > 0 else None

            if cat == "VAC":
                vac_rows = [r_d, r_m, r_n]
                if r_ed:
                    vac_rows.append(r_ed)
                if r_em:
                    vac_rows.append(r_em)
                if r_en:
                    vac_rows.append(r_en)
                fer_r = scn.get("fer")
                if fer_r:
                    vac_rows.append(fer_r)
                for rel_row in vac_rows:
                    vc = ws.cell(row=rel_row, column=col)
                    vc.fill = VAC_FILL
                ws.cell(row=r_d, column=col).comment = Comment("VAC", "planilla")

            total_emp_d += od
            total_emp_m += om
            total_emp_n += on
            total_emp_ed += ed
            total_emp_em += em
            total_emp_en += en

        total_hrs_feriado = 0
        recargo_feriado = 0.0

        if has_holidays_this_week:
            holiday_row = scn.get("fer")
            if holiday_row is None:
                for cand in range(emp_row + 3, min(emp_row + 12, ws.max_row + 1)):
                    if "Feriado" in str(ws.cell(row=cand, column=2).value or ""):
                        holiday_row = cand
                        break
            if holiday_row is not None and "Feriado" in str(
                ws.cell(row=int(holiday_row), column=2).value or ""
            ):
                hr_i = int(holiday_row)
                ws.row_dimensions[hr_i].hidden = False

                # Tarifas desde BD (las celdas D3/F3/H3 pueden ser fórmulas y openpyxl no evalúa)
                tf = db.get_tarifas()
                tarifas = {
                    "tarifa_diurna": float(tf.get("tarifa_diurna") or 0),
                    "tarifa_mixta": float(tf.get("tarifa_mixta") or 0),
                    "tarifa_nocturna": float(tf.get("tarifa_nocturna") or 0),
                }

                tarifa_tipo, es_off = tarifa_tipo_desde_totales_semana(
                    total_emp_d + total_emp_ed,
                    total_emp_m + total_emp_em,
                    total_emp_n + total_emp_en,
                )
                if es_off:
                    tarifa_usar = float(tarifas["tarifa_mixta"] or 0)
                else:
                    tarifa_usar = float(_valor_tarifa_tipo(tarifa_tipo, tarifas) or 0)

                for dia_es, col_idx, _hn in semana_feriados:
                    turno_f = emp_schedule.get(dia_es, "OFF")
                    cat_f = turno_categoria_planilla(turno_f)
                    hd, hn, hm = horas_base_dia_planilla(turno_f)
                    fh = feriado_celda_horas(cat_f)
                    ws.cell(row=hr_i, column=col_idx).value = fh if fh else None
                    total_hrs_feriado += fh

                # Columna L (bruto feriado): fórmula en la planilla (horas × tarifa dominante).
                # No sobrescribir: el Excel recalcula al cargar horas en C–I de la fila ★.
                recargo_feriado = round(total_hrs_feriado * tarifa_usar, 2)

        recap_horas[emp_nombre] = {
            "diurnas": total_emp_d,
            "mixtas": total_emp_m,
            "nocturnas": total_emp_n,
            "extra_diurnas": total_emp_ed,
            "extra_mixtas": total_emp_em,
            "extra_nocturnas": total_emp_en,
            "extra": total_emp_ed + total_emp_em + total_emp_en,
            "horas_feriado_no_labor": total_hrs_feriado,
            "recargo_feriado": recargo_feriado,
        }

    wb.save(excel_path)
    wb.close()
    return True, "Horas rellenadas exitosamente", recap_horas


# ═════════════════════════════════════════════════════════════════════════════
# SINCRONIZAR EMPLEADOS DE HORARIO → PLANILLA
# ═════════════════════════════════════════════════════════════════════════════

def sincronizar_empleados_a_planilla():
    """
    Toma los empleados de horario_empleados que no existen en la tabla
    empleados de planilla y los agrega con tipo_pago='tarjeta' por defecto.
    Retorna lista de empleados agregados.
    """
    conn = db.get_conn()
    hor_emps = conn.execute(
        "SELECT nombre FROM horario_empleados WHERE activo=1"
    ).fetchall()
    plan_emps = conn.execute(
        "SELECT nombre FROM empleados"
    ).fetchall()
    plan_nombres = {r["nombre"] for r in plan_emps}

    agregados = []
    for r in hor_emps:
        if r["nombre"] not in plan_nombres:
            conn.execute(
                "INSERT INTO empleados (nombre, tipo_pago) VALUES (?, 'tarjeta')",
                (r["nombre"],)
            )
            agregados.append(r["nombre"])

    conn.commit()
    conn.close()
    return agregados
